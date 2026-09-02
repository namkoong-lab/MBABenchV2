"""Workbook/sheet properties block (judge v6, item D — "serve the blind evidence").

The 2026-09-01 evidence sweep found ~34 rubric checks graded on properties
the judge is never shown: true tab order (the file listing was sorted
alphabetically), hidden sheets/rows/columns (served as ordinary content),
data validation, column widths and row heights, cell comments, conditional
formats, hyperlinks, defined names, calc mode, print setup. This module
extracts all of that ONCE at CSV-extraction time, saves it beside the CSVs
(`_workbook_properties.json`, so the per-workbook cache carries it and repeat
gradings reuse identical facts), and renders it as a compact, deterministic
text block for the judge's seed prompt.

Where a property cannot be read the block says so explicitly ("unknown"),
so the model can tell "absent" from "not provided".

Cache generation: files written here ride in `*_csv_cache_v3` — a v2 cache
has no properties file and the loaders degrade to the old behaviour
(alphabetical listing, no block), which is why the generation was bumped.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Optional

from openpyxl.utils import get_column_letter

try:
    from .logger import logger
except ImportError:  # imported as a bare module (utils/ on sys.path)
    from logger import logger

FILENAME = "_workbook_properties.json"
SCHEMA_VERSION = 1
_MAX_LIST = 25          # per-list cap in the rendered text (JSON keeps everything)
_MAX_COMMENT_CHARS = 160


def _safe(fn, default="unknown"):
    try:
        v = fn()
        return default if v is None else v
    except Exception:  # noqa: BLE001 — property reads must never break extraction
        return default


def _runs(pairs: list[tuple[int, Any]]) -> list[dict]:
    """[(index, value)] sorted by index -> [{first, last, value}] runs of
    consecutive indexes with equal values."""
    runs: list[dict] = []
    for idx, val in sorted(pairs, key=lambda p: p[0]):
        if runs and runs[-1]["last"] == idx - 1 and runs[-1]["value"] == val:
            runs[-1]["last"] = idx
        else:
            runs.append({"first": idx, "last": idx, "value": val})
    return runs


def _col_runs_text(runs: list[dict], fmt=lambda v: v) -> str:
    parts = []
    for r in runs:
        a, b = get_column_letter(r["first"]), get_column_letter(r["last"])
        parts.append(f"{a}{'' if a == b else ':' + b}={fmt(r['value'])}")
    return ", ".join(parts)


def _row_runs_text(runs: list[dict], fmt=lambda v: v) -> str:
    parts = []
    for r in runs:
        a, b = r["first"], r["last"]
        parts.append(f"{a}{'' if a == b else '-' + str(b)}={fmt(r['value'])}")
    return ", ".join(parts)


def _ranges_text(indexes: list[int], col: bool) -> str:
    if not indexes:
        return "none"
    runs = _runs([(i, True) for i in indexes])
    parts = []
    for r in runs:
        if col:
            a, b = get_column_letter(r["first"]), get_column_letter(r["last"])
        else:
            a, b = str(r["first"]), str(r["last"])
        parts.append(a if a == b else f"{a}:{b}" if col else f"{a}-{b}")
    return ", ".join(parts)


# ---------------------------------------------------------------------------
# Extraction
# ---------------------------------------------------------------------------


def _color_text(color) -> Optional[str]:
    """openpyxl Color -> 'rgb:FFC000' / 'theme:4' / 'indexed:12' / None.

    `.rgb` on a theme colour is a descriptor ERROR STRING in openpyxl
    ("Values must be of type <class 'str'>"), so read the type first.
    """
    if color is None:
        return None
    ctype = getattr(color, "type", None)
    if ctype == "rgb":
        rgb = getattr(color, "rgb", None)
        return f"rgb:{rgb}" if isinstance(rgb, str) and len(rgb) in (6, 8) else None
    if ctype == "theme":
        tint = getattr(color, "tint", 0) or 0
        return f"theme:{getattr(color, 'theme', '?')}" + (f" tint {tint:+.2f}" if tint else "")
    if ctype == "indexed":
        return f"indexed:{getattr(color, 'indexed', '?')}"
    return None


# Defined names that add-ins/system tooling plant in workbooks (Capital IQ,
# @RISK, Palisade, print areas, solver) — counted, not listed.
_SYSTEM_NAME_PREFIXES = ("IQ_", "IQB_", "Risk", "Pal_", "_xlnm", "solver_", "_xlfn", "Slicer_")


def _sheet_properties(ws, index: int, output_name: Optional[str]) -> dict:
    kind = "chartsheet" if type(ws).__name__ == "Chartsheet" else "worksheet"
    props: dict[str, Any] = {
        "name": ws.title,
        "output_name": output_name,
        "index": index,
        "kind": kind,
        "state": _safe(lambda: ws.sheet_state, "visible"),
        "tab_color": _safe(lambda: _color_text(ws.sheet_properties.tabColor), None),
    }
    if kind == "chartsheet":
        return props

    props.update({
        "max_row": _safe(lambda: ws.max_row, 0),
        "max_column": _safe(lambda: ws.max_column, 0),
        "zoom": _safe(lambda: ws.sheet_view.zoomScale, None),
        "gridlines": _safe(lambda: ws.sheet_view.showGridLines, None),
        "freeze_panes": _safe(lambda: ws.freeze_panes, None),
        "merged_ranges": _safe(lambda: sorted(str(r) for r in ws.merged_cells.ranges), []),
        "protected": _safe(lambda: bool(ws.protection.sheet), None),
        "print_area": _safe(lambda: ws.print_area, None),
        "print_title_rows": _safe(lambda: ws.print_title_rows, None),
        "print_title_cols": _safe(lambda: ws.print_title_cols, None),
        "page_orientation": _safe(lambda: ws.page_setup.orientation, None),
        "fit_to_page": _safe(
            lambda: bool(ws.sheet_properties.pageSetUpPr.fitToPage)
            if ws.sheet_properties.pageSetUpPr else False,
            None,
        ),
    })

    # used range + comments + counts in one pass over the cells
    min_r = min_c = None
    max_r = max_c = 0
    n_values = n_formulas = 0
    comments = []
    try:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                r, c = cell.row, cell.column
                min_r = r if min_r is None or r < min_r else min_r
                min_c = c if min_c is None or c < min_c else min_c
                max_r, max_c = max(max_r, r), max(max_c, c)
                n_values += 1
                if isinstance(v, str) and v.startswith("="):
                    n_formulas += 1
                elif type(v).__name__ in ("ArrayFormula", "DataTableFormula"):
                    n_formulas += 1
            for cell in row:
                cm = getattr(cell, "comment", None)
                if cm is not None:
                    comments.append({
                        "ref": cell.coordinate,
                        "author": getattr(cm, "author", None),
                        "text": (getattr(cm, "text", "") or "")[:_MAX_COMMENT_CHARS],
                    })
        props["used_range"] = (
            f"{get_column_letter(min_c)}{min_r}:{get_column_letter(max_c)}{max_r}"
            if min_r is not None else None
        )
        props["n_values"] = n_values
        props["n_formulas"] = n_formulas
        props["comments"] = comments
    except Exception:  # noqa: BLE001
        props["used_range"] = "unknown"
        props["comments"] = "unknown"

    # hidden rows / cols, widths / heights
    try:
        hidden_rows, heights = [], []
        for idx, dim in ws.row_dimensions.items():
            if getattr(dim, "hidden", False):
                hidden_rows.append(int(idx))
            h = getattr(dim, "height", None)
            if h is not None and getattr(dim, "customHeight", None) is not False:
                heights.append((int(idx), round(float(h), 1)))
        props["hidden_rows"] = sorted(hidden_rows)
        props["row_heights"] = _runs(heights)
        props["default_row_height"] = _safe(lambda: ws.sheet_format.defaultRowHeight, None)
    except Exception:  # noqa: BLE001
        props["hidden_rows"] = "unknown"
        props["row_heights"] = "unknown"
    try:
        hidden_cols, widths = [], []
        for key, dim in ws.column_dimensions.items():
            lo = int(getattr(dim, "min", None) or 0) or None
            hi = int(getattr(dim, "max", None) or 0) or None
            if lo is None:
                from openpyxl.utils import column_index_from_string
                lo = hi = column_index_from_string(key)
            hi = hi or lo
            if hi - lo > 200:  # openpyxl sometimes reports a 16384-wide default dim
                hi = lo
            if getattr(dim, "hidden", False):
                hidden_cols.extend(range(lo, hi + 1))
            w = getattr(dim, "width", None)
            if w is not None and getattr(dim, "customWidth", True):
                for c in range(lo, hi + 1):
                    widths.append((c, round(float(w), 2)))
        props["hidden_cols"] = sorted(set(hidden_cols))
        props["column_widths"] = _runs(widths)
        props["default_col_width"] = _safe(lambda: ws.sheet_format.defaultColWidth, None)
    except Exception:  # noqa: BLE001
        props["hidden_cols"] = "unknown"
        props["column_widths"] = "unknown"

    # data validation / conditional formatting / hyperlinks
    try:
        dvs = []
        for dv in (ws.data_validations.dataValidation if ws.data_validations else []):
            dvs.append({
                "sqref": str(dv.sqref),
                "type": dv.type,
                "operator": dv.operator,
                "formula1": dv.formula1,
                "formula2": dv.formula2,
                "allow_blank": dv.allowBlank,
            })
        props["data_validations"] = sorted(dvs, key=lambda d: d["sqref"])
    except Exception:  # noqa: BLE001
        props["data_validations"] = "unknown"
    try:
        cfs = []
        for cf in ws.conditional_formatting:
            cfs.append({
                "sqref": str(cf.sqref),
                "rules": [
                    {"type": r.type, "operator": getattr(r, "operator", None),
                     "formula": list(getattr(r, "formula", []) or [])}
                    for r in cf.rules
                ],
            })
        props["conditional_formats"] = sorted(cfs, key=lambda d: d["sqref"])
    except Exception:  # noqa: BLE001
        props["conditional_formats"] = "unknown"
    try:
        links = []
        for hl in ws._hyperlinks:
            links.append({"ref": hl.ref, "target": hl.target or hl.location})
        props["hyperlinks"] = sorted(links, key=lambda d: d["ref"])
    except Exception:  # noqa: BLE001
        props["hyperlinks"] = "unknown"
    return props


def extract_workbook_properties(workbook, excel_file_path, name_map: dict | None = None) -> dict:
    """Everything the rubric grades that is not a cell value.

    `name_map` maps original sheet names to the output (filtered/safe) names
    the CSVs were saved under; sheets skipped by the filter map to None.
    """
    path = Path(excel_file_path)
    wb: dict[str, Any] = {
        "filename": path.name,
        "bytes": _safe(lambda: path.stat().st_size, None),
        "has_vba": _safe(lambda: workbook.vba_archive is not None, None),
        "calc_mode": _safe(lambda: workbook.calculation.calcMode, None),
        "full_calc_on_load": _safe(lambda: workbook.calculation.fullCalcOnLoad, None),
        "iterative_calc": _safe(lambda: workbook.calculation.iterate, None),
        "defined_names": [],
        "external_links": [],
        "active_sheet": _safe(lambda: workbook.active.title if workbook.active else None, None),
    }
    try:
        names = []
        dn = workbook.defined_names
        items = dn.items() if hasattr(dn, "items") else [(d.name, d) for d in dn.definedName]
        for name, d in items:
            names.append({"name": name, "refers_to": getattr(d, "attr_text", None), "scope": None})
        for ws in workbook.worksheets:
            local = getattr(ws, "defined_names", None)
            if local and hasattr(local, "items"):
                for name, d in local.items():
                    names.append({"name": name, "refers_to": getattr(d, "attr_text", None),
                                  "scope": ws.title})
        wb["defined_names"] = sorted(names, key=lambda d: (d["scope"] or "", d["name"]))
    except Exception:  # noqa: BLE001
        wb["defined_names"] = "unknown"
    try:
        wb["external_links"] = sorted(
            str(getattr(getattr(l, "file_link", None), "Target", None) or "?")
            for l in workbook._external_links
        )
    except Exception:  # noqa: BLE001
        wb["external_links"] = "unknown"

    sheets = []
    name_map = name_map or {}
    for i, ws in enumerate(workbook._sheets if hasattr(workbook, "_sheets") else workbook.worksheets, 1):
        sheets.append(_sheet_properties(ws, i, name_map.get(ws.title, ws.title)))
    return {"schema": SCHEMA_VERSION, "workbook": wb, "sheets": sheets}


def save_properties(output_dir: Path, props: dict) -> Path:
    p = Path(output_dir) / FILENAME
    p.write_text(json.dumps(props, indent=2, default=str), encoding="utf-8")
    return p


def load_properties(directory) -> Optional[dict]:
    if not directory:
        return None
    p = Path(directory) / FILENAME
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as e:
        logger.warning(f"  [properties] unreadable {p}: {e}")
        return None


# ---------------------------------------------------------------------------
# Ordering + rendering
# ---------------------------------------------------------------------------


def order_file_list(file_list: list[str], props: Optional[dict]) -> list[str]:
    """Order `*_full.csv` names by the workbook's TRUE tab order.

    Names not present in the properties (older caches, filtered sheets) keep
    an alphabetical order after the known ones, so a v2 cache degrades to
    exactly the old listing.
    """
    if not props:
        return sorted(file_list)
    rank = {}
    for s in props.get("sheets", []):
        out_name = s.get("output_name")
        if out_name:
            from .excel_utils import create_safe_filename  # local import: avoid a cycle at import time

            rank[f"{create_safe_filename(out_name)}_full.csv"] = s["index"]
    known = sorted((f for f in file_list if f in rank), key=lambda f: rank[f])
    unknown = sorted(f for f in file_list if f not in rank)
    return known + unknown


def _fmt_list(items, limit=_MAX_LIST, fn=str) -> str:
    if items == "unknown":
        return "unknown"
    if not items:
        return "none"
    shown = [fn(x) for x in items[:limit]]
    more = len(items) - limit
    return "; ".join(shown) + (f"; (+{more} more)" if more > 0 else "")


def render_properties_text(props: Optional[dict], listed_files: Optional[set] = None) -> str:
    """Compact deterministic text for the seed prompt.

    `listed_files` (the `*_full.csv` names actually served) marks sheets
    whose CSV was dropped (ignored/filtered) so the judge is not sent
    looking for a file that is not there.
    """
    if not props:
        return "  (workbook properties not available — older extraction cache)"
    wb = props.get("workbook", {})
    lines = []
    size = wb.get("bytes")
    size_txt = f"{size / 1024:.0f} KB" if isinstance(size, (int, float)) else "unknown size"
    calc_mode = wb.get("calc_mode")
    lines.append(
        f"Workbook {wb.get('filename', '?')} ({size_txt}); calc mode: "
        f"{calc_mode if calc_mode and calc_mode != 'unknown' else 'auto (Excel default, none set)'}"
        f"{' (full calc on load)' if wb.get('full_calc_on_load') else ''}; iterative calc: "
        f"{'on' if wb.get('iterative_calc') else ('off' if wb.get('iterative_calc') is not None else 'unknown')}; "
        f"VBA: {'yes' if wb.get('has_vba') else ('no' if wb.get('has_vba') is not None else 'unknown')}; "
        f"active sheet: {wb.get('active_sheet') or 'unknown'}"
    )
    dn = wb.get("defined_names")
    if isinstance(dn, list):
        user_names = [d for d in dn if not str(d.get("name", "")).startswith(_SYSTEM_NAME_PREFIXES)]
        n_sys = len(dn) - len(user_names)
        lines.append(
            "Defined names: " + _fmt_list(
                user_names,
                fn=lambda d: f"{d['name']}{' (' + d['scope'] + ')' if d.get('scope') else ''} -> {d.get('refers_to')}",
            )
            + (f" [+{n_sys} add-in/system names not listed]" if n_sys else "")
        )
    else:
        lines.append("Defined names: unknown")
    lines.append("External links: " + _fmt_list(wb.get("external_links")))
    lines.append("Sheets in TRUE TAB ORDER (index. name [state]):")
    from .excel_utils import create_safe_filename

    for s in props.get("sheets", []):
        state = s.get("state", "visible")
        tag = f" [{state.upper()}]" if state and state != "visible" else ""
        if s.get("kind") == "chartsheet":
            lines.append(f"  {s['index']}. {s['name']}{tag} — chart sheet (no cell data)")
            continue
        out = s.get("output_name")
        served = ""
        if out is None:
            served = " — not served (filtered out)"
        elif listed_files is not None and f"{create_safe_filename(out)}_full.csv" not in listed_files:
            served = " — not served (ignored sheet)"
        elif out != s["name"]:
            served = f" — served as {out}_full.csv"
        head = (
            f"  {s['index']}. {s['name']}{tag}{served}: "
            f"{s.get('max_row')}x{s.get('max_column')} (used {s.get('used_range') or 'empty'}; "
            f"{s.get('n_values', '?')} values, {s.get('n_formulas', '?')} formulas)"
        )
        lines.append(head)
        fp = s.get("freeze_panes")
        detail = [
            f"freeze panes: {fp or 'none'}",
            f"gridlines: {'on' if s.get('gridlines') in (None, True) else 'off'}",
            f"zoom: {s.get('zoom') or 100}",
            f"tab color: {s.get('tab_color') or 'none'}",
            f"protected: {'yes' if s.get('protected') else 'no'}",
            f"merged ranges: {len(s.get('merged_ranges') or []) if s.get('merged_ranges') != 'unknown' else 'unknown'}",
        ]
        lines.append("     " + "; ".join(detail))
        hr, hc = s.get("hidden_rows", []), s.get("hidden_cols", [])
        lines.append(
            "     hidden rows: " + (_ranges_text(hr, col=False) if hr != "unknown" else "unknown")
            + "; hidden cols: " + (_ranges_text(hc, col=True) if hc != "unknown" else "unknown")
        )
        cw = s.get("column_widths")
        dcw = s.get("default_col_width")
        lines.append(
            "     column widths: "
            + (_col_runs_text(cw[:_MAX_LIST]) + (f", (+{len(cw) - _MAX_LIST} more runs)" if len(cw) > _MAX_LIST else "")
               if isinstance(cw, list) and cw else ("unknown" if cw == "unknown" else "all default"))
            + f" (default {dcw if dcw else 8.43})"
        )
        rh = s.get("row_heights")
        lines.append(
            "     custom row heights: "
            + (_row_runs_text(rh[:_MAX_LIST]) + (f", (+{len(rh) - _MAX_LIST} more runs)" if len(rh) > _MAX_LIST else "")
               if isinstance(rh, list) and rh else ("unknown" if rh == "unknown" else "none"))
        )
        lines.append(
            "     data validation: " + _fmt_list(
                s.get("data_validations"),
                fn=lambda d: f"{d['sqref']} {d.get('type') or '?'}"
                             f"{' ' + str(d.get('formula1')) if d.get('formula1') else ''}",
            )
        )
        lines.append(
            "     conditional formats: " + _fmt_list(
                s.get("conditional_formats"),
                fn=lambda d: f"{d['sqref']} ({', '.join(r.get('type') or '?' for r in d.get('rules', []))})",
            )
        )
        lines.append(
            "     comments/notes: " + _fmt_list(
                s.get("comments"),
                fn=lambda d: f"{d['ref']}: \"{(d.get('text') or '').strip()[:80]}\"",
            )
        )
        lines.append("     hyperlinks: " + _fmt_list(s.get("hyperlinks"), fn=lambda d: f"{d['ref']} -> {d.get('target')}"))
        lines.append(
            f"     print: area {s.get('print_area') or 'none'}; title rows {s.get('print_title_rows') or 'none'}; "
            f"orientation {s.get('page_orientation') or 'default'}; fit to page: "
            f"{'yes' if s.get('fit_to_page') else 'no'}"
        )
    return "\n".join(lines)
