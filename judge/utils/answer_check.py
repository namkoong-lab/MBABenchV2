"""Harness answer check (judge v6) — deterministic Questions-sheet grading.

Two jobs:

1. **Measure** every Questions-sheet answer of the attempt against the golden
   solution with the shared rulebook (utils/answer_rules.py) — the full
   per-question artifact rides with the grading as `answer_check.json`.
2. **Decide** the two harness-decidable Accuracy checks and hand the verdicts
   to the judge's scoring layer (`harness_verdicts`), where the grading's
   `--accuracy-check harness|llm` flag chooses which engine's decision lands
   in the recorded total. BOTH engines' verdicts are always recorded, so the
   comparison never needs a re-run.

Finding the golden's answers (67/68 goldens follow one convention; task 54
has TWO sheets, "Questions Task 1"/"Questions Task 2", and both are read):
sheets whose name starts with "questions"; a header row within the first ten
rows holding an "Answers" cell; questions in the "Questions..." column
(column A everywhere); optional "Unit" column; the header phrase carries the
requested precision ("round your answers to two decimal places").

Finding the ATTEMPT's answers is name-agnostic (the canary showed 2 of 4
attempts named the sheet "Answers" and one had an "Answer Map" decoy):
  1. every sheet is indexed by the golden question texts it contains
     (normalised exact match, then a conservative fuzzy pass);
  2. a sheet named Questions*/Answers* wins ties (name is a hint, not proof —
     it still has to contain the question texts);
  3. per-question rows are paired by TEXT, so reordering, inserted rows and
     phantom rows do not matter; an unmatched question is "unanswered",
     never guessed;
  4. the answer column is the cell in the question row's header row whose
     text starts with "answer" (a 297-style layout has Path/Metric/Year/
     "Model value" between the question and the "Answer [$] (2 dp)" column,
     so "adjacent column" would pick the wrong one); no such header => the
     sheet FAILS CLOSED to the LLM verdict, with the reason recorded;
  5. uncached formula results are recalculated once through LibreOffice.

Per answer the checker also records whether the cell holds a live formula;
a numeric answer typed as a constant is `hardcoded` — the rubric's Final
calculation accuracy check rejects those even when the value is right.

Never raises out of run_answer_check(): any failure is status "error" and the
judge grades on with the LLM verdicts.
"""

from __future__ import annotations

import difflib
import json
import re
import shutil
import subprocess
import tempfile
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, time as dt_time
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.utils import get_column_letter

try:
    from . import answer_rules as rules
    from .logger import logger
    from .misc_utils import load_env_var
except ImportError:  # imported as a bare module (utils/ on sys.path)
    import answer_rules as rules
    from logger import logger
    from misc_utils import load_env_var

HEADER_SCAN_ROWS = 10
HEADER_SCAN_COLS = 8  # A-H
FUZZY_MIN_RATIO = 0.90
FUZZY_MIN_LEN = 15
# A sheet "validates" as an answer sheet when it contains at least this many
# golden question texts (absolute floor) — protects against a lone stray
# match in a model sheet.
VALIDATE_MIN_HITS = 1
# The harness measures Final calculation accuracy only when it could locate
# at least this share of the golden questions in the attempt; below it, the
# layout is too foreign to trust and the check falls back to the LLM.
MEASURE_MIN_LOCATED_SHARE = 0.5

CHECK_FINAL_ACCURACY = ("Accuracy", "Final calculation accuracy")
CHECK_COMPLETENESS = ("Accuracy", "Deliverable completeness")


def _is_number(v) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _display(v) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, (datetime, date, dt_time)):
        return v.isoformat()
    return str(v)


def _jsonable(v):
    if v is None or _is_number(v) or isinstance(v, (str, bool)):
        return v
    return _display(v)


def _is_formula(v) -> bool:
    if isinstance(v, str) and v.startswith("="):
        return True
    return type(v).__name__ in ("ArrayFormula", "DataTableFormula")


def _answered(v) -> bool:
    return v is not None and not (isinstance(v, str) and v.strip() == "")


# ---------------------------------------------------------------------------
# Golden side
# ---------------------------------------------------------------------------


@dataclass
class GoldenQuestion:
    qid: int
    sheet: str
    row: int
    label: str
    norm_label: str
    value: Any
    formula: Any
    number_format: Any
    unit: Any
    answer_col: int


@dataclass
class GoldenAnswers:
    status: str
    reason: Optional[str] = None
    sheets: list = field(default_factory=list)      # [{name, header_row, question_col, answer_col, unit_col, header_text}]
    questions: list = field(default_factory=list)   # [GoldenQuestion]
    precision: dict = field(default_factory=dict)   # per sheet name -> Precision dict
    percent_directive: dict = field(default_factory=dict)  # sheet -> bool
    recalc_used: bool = False
    recalc_error: Optional[str] = None


def _questions_sheets(wb) -> list[str]:
    return [n for n in wb.sheetnames if n.strip().casefold().startswith("questions")]


def _find_header(ws) -> Optional[tuple[int, int, Optional[int], Optional[int], str]]:
    """(header_row, answer_col, question_col, unit_col, question_header_text)."""
    max_r = min(HEADER_SCAN_ROWS, ws.max_row or 1)
    for r in range(1, max_r + 1):
        answer_col = question_col = unit_col = None
        qtext = ""
        for c in range(1, HEADER_SCAN_COLS + 1):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str):
                continue
            t = rules.norm_text(v)
            if t in ("answers", "answer") and answer_col is None:
                answer_col = c
            elif t.startswith("question") and question_col is None:
                question_col = c
                qtext = v
            elif t.startswith("unit") and unit_col is None:
                unit_col = c
        if answer_col is not None:
            return r, answer_col, (question_col or 1), unit_col, qtext
    return None


def extract_golden(xlsx_path: Path, allow_recalc: bool = True) -> GoldenAnswers:
    xlsx_path = Path(xlsx_path)
    wb_f = openpyxl.load_workbook(str(xlsx_path), data_only=False)
    wb_v = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    out = GoldenAnswers(status="ok")
    try:
        sheets = _questions_sheets(wb_v)
        if not sheets:
            out.status, out.reason = "convention_not_found", "no sheet named Questions*"
            return out
        anchors = {}
        for name in sheets:
            hdr = _find_header(wb_v[name])
            if hdr:
                anchors[name] = hdr
        if not anchors:
            out.status = "convention_not_found"
            out.reason = "no 'Answers' header in rows 1-10, columns A-H"
            return out

        # Recalc once if any golden answer cell is a formula with no cached value.
        targets = []
        for name, (hrow, acol, qcol, ucol, _) in anchors.items():
            ws = wb_v[name]
            for r in range(hrow + 1, (ws.max_row or hrow) + 1):
                targets.append((name, r, acol))
        if allow_recalc and _uncached_count(wb_f, wb_v, targets):
            wb_v, out.recalc_used, out.recalc_error = _recalc_reload(xlsx_path, wb_v, "golden")

        qid = 0
        for name, (hrow, acol, qcol, ucol, qtext) in anchors.items():
            ws_v, ws_f = wb_v[name], wb_f[name]
            prec = rules.parse_precision_directive(qtext)
            out.precision[name] = asdict(prec)
            out.percent_directive[name] = bool(
                re.search(r"decimal proportion|as a decimal|stored as a decimal",
                          rules.norm_text(qtext))
            )
            out.sheets.append({
                "name": name, "header_row": hrow, "question_col": get_column_letter(qcol),
                "answer_col": get_column_letter(acol),
                "unit_col": get_column_letter(ucol) if ucol else None,
                "header_text": qtext,
            })
            for r in range(hrow + 1, (ws_v.max_row or hrow) + 1):
                label = ws_v.cell(row=r, column=qcol).value
                value = ws_v.cell(row=r, column=acol).value
                if (label is None or str(label).strip() == "") and value is None:
                    continue  # blank / phantom row
                if label is None or str(label).strip() == "":
                    continue  # value without a question — not gradable by text
                qid += 1
                fcell = ws_f.cell(row=r, column=acol)
                out.questions.append(GoldenQuestion(
                    qid=qid, sheet=name, row=r, label=str(label).strip(),
                    norm_label=rules.norm_text(label), value=value,
                    formula=fcell.value if _is_formula(fcell.value) else None,
                    number_format=fcell.number_format,
                    unit=ws_v.cell(row=r, column=ucol).value if ucol else None,
                    answer_col=acol,
                ))
        if not out.questions:
            out.status = "convention_not_found"
            out.reason = "Answers header found but no question rows below it"
        return out
    finally:
        wb_f.close()
        wb_v.close()


# ---------------------------------------------------------------------------
# Attempt side (name-agnostic)
# ---------------------------------------------------------------------------


@dataclass
class LocatedAnswer:
    qid: int
    sheet: str
    row: int
    col: int
    value: Any
    is_formula: bool
    formula: Any
    number_format: Any
    match_kind: str  # "exact" | "fuzzy"


@dataclass
class AttemptAnswers:
    status: str
    reason: Optional[str] = None
    sheets_used: list = field(default_factory=list)   # [{name, header_row, answer_col, hits, priority}]
    answers: dict = field(default_factory=dict)       # qid -> LocatedAnswer
    unlocated_qids: list = field(default_factory=list)
    sheets_without_answer_column: list = field(default_factory=list)
    recalc_used: bool = False
    recalc_error: Optional[str] = None


def _sheet_priority(name: str) -> int:
    t = name.strip().casefold()
    if t.startswith("questions"):
        return 0
    if t.startswith("question"):
        return 1
    if t.startswith("answers"):
        return 2
    if t.startswith("answer"):
        return 3
    return 9


def _index_text_cells(wb) -> dict[str, list[tuple[str, int, int, str]]]:
    """norm text -> [(sheet, row, col, raw)] over every string cell."""
    index: dict[str, list] = {}
    for ws in wb.worksheets:
        if type(ws).__name__ == "Chartsheet":
            continue
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            for c_idx, v in enumerate(row, 1):
                if isinstance(v, str) and v.strip():
                    index.setdefault(rules.norm_text(v), []).append((ws.title, r_idx, c_idx, v))
    return index


def _pick(cands: list[tuple[str, int, int, str]]) -> tuple[str, int, int, str]:
    """Best candidate location for a question text: named sheets first."""
    return sorted(cands, key=lambda t: (_sheet_priority(t[0]), t[0], t[1]))[0]


def _find_answer_col(ws, question_row: int, question_col: int,
                     cache: dict) -> tuple[Optional[int], Optional[int]]:
    """(answer_col, header_row) for the block containing *question_row*.

    The header row is the nearest row at or above the question (within 60
    rows) that holds a cell starting with "question", else the row directly
    above the first question of the block. In that row the answer column is
    the first cell whose text starts with "answer" (excluding the question
    column itself). Cached per sheet.
    """
    key = ws.title
    if key in cache:
        return cache[key]
    result = (None, None)
    top = max(1, question_row - 60)
    for r in range(question_row, top - 1, -1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column or 1, 40) + 1)]
        norms = [rules.norm_text(v) if isinstance(v, str) else "" for v in row_vals]
        has_q_header = any(n.startswith("question") for n in norms)
        if not has_q_header and r != question_row - 1:
            continue
        for c_idx, n in enumerate(norms, 1):
            if c_idx == question_col:
                continue
            if n.startswith("answer") and not n.startswith("answer map"):
                result = (c_idx, r)
                break
        if result[0] is not None or has_q_header:
            break
    cache[key] = result
    return result


def extract_attempt(xlsx_path: Path, golden: GoldenAnswers,
                    allow_recalc: bool = True) -> AttemptAnswers:
    xlsx_path = Path(xlsx_path)
    out = AttemptAnswers(status="ok")
    wb_f = openpyxl.load_workbook(str(xlsx_path), data_only=False)
    wb_v = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    try:
        index = _index_text_cells(wb_v)
        located: dict[int, tuple[str, int, int, str]] = {}
        # Pass 1 — normalised exact text.
        for q in golden.questions:
            cands = index.get(q.norm_label)
            if cands:
                located[q.qid] = (*_pick(cands)[:3], "exact")
        # Pass 2 — conservative fuzzy on the leftovers, restricted to sheets
        # that already matched something (or named sheets if nothing did).
        leftovers = [q for q in golden.questions if q.qid not in located]
        if leftovers:
            allowed = {loc[0] for loc in located.values()} or {
                n for n in wb_v.sheetnames if _sheet_priority(n) < 9
            }
            pool = [(t, locs) for t, locs in index.items()
                    if len(t) >= FUZZY_MIN_LEN and any(l[0] in allowed for l in locs)]
            used = {(l[0], l[1]) for l in located.values()}
            for q in leftovers:
                if len(q.norm_label) < FUZZY_MIN_LEN:
                    continue
                best, best_ratio = None, 0.0
                for t, locs in pool:
                    ratio = difflib.SequenceMatcher(None, q.norm_label, t).ratio()
                    if ratio > best_ratio:
                        best, best_ratio = locs, ratio
                if best and best_ratio >= FUZZY_MIN_RATIO:
                    loc = _pick([l for l in best if l[0] in allowed] or best)
                    if (loc[0], loc[1]) not in used:
                        located[q.qid] = (loc[0], loc[1], loc[2], "fuzzy")
                        used.add((loc[0], loc[1]))

        if not located:
            out.status = "answers_not_found"
            out.reason = "no sheet contains the golden question texts"
            return out

        # Answer column per sheet, then the target cells.
        col_cache: dict = {}
        sheet_hits: dict[str, int] = {}
        targets: list[tuple[int, str, int, int]] = []
        for qid, (sname, r, c, kind) in located.items():
            sheet_hits[sname] = sheet_hits.get(sname, 0) + 1
            acol, hrow = _find_answer_col(wb_v[sname], r, c, col_cache)
            if acol is None:
                continue
            targets.append((qid, sname, r, acol))
        for sname, hits in sorted(sheet_hits.items(), key=lambda kv: (_sheet_priority(kv[0]), kv[0])):
            acol, hrow = col_cache.get(sname, (None, None))
            entry = {"name": sname, "hits": hits, "priority": _sheet_priority(sname),
                     "header_row": hrow,
                     "answer_col": get_column_letter(acol) if acol else None}
            out.sheets_used.append(entry)
            if acol is None:
                out.sheets_without_answer_column.append(sname)

        if allow_recalc and targets and _uncached_count(
            wb_f, wb_v, [(s, r, c) for _, s, r, c in targets]
        ):
            wb_v, out.recalc_used, out.recalc_error = _recalc_reload(xlsx_path, wb_v, "attempt")

        for qid, sname, r, acol in targets:
            fcell = wb_f[sname].cell(row=r, column=acol)
            out.answers[qid] = LocatedAnswer(
                qid=qid, sheet=sname, row=r, col=acol,
                value=wb_v[sname].cell(row=r, column=acol).value,
                is_formula=_is_formula(fcell.value),
                formula=fcell.value if _is_formula(fcell.value) else None,
                number_format=fcell.number_format,
                match_kind=located[qid][3],
            )
        out.unlocated_qids = [q.qid for q in golden.questions if q.qid not in out.answers]
        if not out.answers:
            out.status = "answers_not_found"
            out.reason = (
                "question texts found but no 'Answer' header column in: "
                + ", ".join(out.sheets_without_answer_column)
            )
        return out
    finally:
        wb_f.close()
        wb_v.close()


# ---------------------------------------------------------------------------
# Recalc helpers
# ---------------------------------------------------------------------------


def _uncached_count(wb_f, wb_v, targets) -> int:
    n = 0
    for sheet, row, col in targets:
        v = wb_f[sheet].cell(row=row, column=col).value
        if _is_formula(v) and wb_v[sheet].cell(row=row, column=col).value is None:
            n += 1
    return n


def _recalculate_copy(xlsx_path: Path, outdir: Path) -> Path:
    soffice = load_env_var("PATHS_LIBREOFFICE_PATH", required=True)
    subprocess.run(
        [soffice, "--headless", "--calc", "--convert-to", "xlsx",
         "--outdir", str(outdir), str(xlsx_path)],
        check=True, capture_output=True, timeout=300,
    )
    out = outdir / xlsx_path.name
    if not out.exists():
        raise FileNotFoundError(f"LibreOffice produced no output for {xlsx_path}")
    return out


def _recalc_reload(xlsx_path: Path, wb_v, side: str):
    """Recalculate through LibreOffice and return (values_wb, used, error)."""
    logger.info(f"  [answer_check] {side} {xlsx_path.name}: answer cells lack cached "
                f"values — recalculating via LibreOffice")
    tmpdir = Path(tempfile.mkdtemp(prefix="answer_check_recalc_"))
    try:
        recalced = _recalculate_copy(xlsx_path, tmpdir)
        wb_v.close()
        return openpyxl.load_workbook(str(recalced), data_only=True), True, None
    except Exception as e:  # noqa: BLE001 — best-effort
        logger.warning(f"  [answer_check] recalc failed: {e}")
        return wb_v, False, str(e)
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


# ---------------------------------------------------------------------------
# Comparison
# ---------------------------------------------------------------------------


def compare_answer_sets(golden: GoldenAnswers, attempt: AttemptAnswers) -> dict:
    questions = []
    for q in golden.questions:
        prec = rules.Precision(**golden.precision.get(q.sheet, {"dp": None}))
        ctx = rules.AnswerContext(
            label=q.label, unit=q.unit, precision=prec,
            expected_formula=q.formula, expected_number_format=q.number_format,
            percent_directive=golden.percent_directive.get(q.sheet, False),
        )
        loc = attempt.answers.get(q.qid)
        got = loc.value if loc else None
        if loc is not None and isinstance(loc.formula, str):
            ctx.got_formula = loc.formula
        cmp = rules.compare(q.value, got, ctx)
        item = {
            "qid": q.qid,
            "label": q.label,
            "golden_cell": f"{q.sheet}!{get_column_letter(q.answer_col)}{q.row}",
            "attempt_cell": f"{loc.sheet}!{get_column_letter(loc.col)}{loc.row}" if loc else None,
            "located": loc is not None,
            "match_kind": loc.match_kind if loc else None,
            "expected": _jsonable(q.value),
            "got": _jsonable(got),
            "unit": _jsonable(q.unit),
            "verdict": cmp["verdict"],
            "rule": cmp["rule"],
            "flags": cmp["flags"],
            "detail": cmp["detail"],
            "tolerance": cmp.get("tolerance"),
            "abs_delta": cmp.get("abs_delta"),
            "attempt_is_formula": bool(loc and loc.is_formula),
            "attempt_formula": (loc.formula if loc and isinstance(loc.formula, str) else None),
            "hardcoded": False,
        }
        if loc is None:
            item["detail"] = "question text not found in the attempt (unanswered)"
        # A numeric golden answered with a numeric CONSTANT is hardcoded —
        # but only where the golden itself computes the answer with a live
        # formula (task 54's second sheet hardcodes 23 scenario counts in
        # the golden; holding attempts to a standard the golden fails is
        # not a measurement).
        if (loc is not None and _answered(got) and not loc.is_formula
                and q.formula is not None
                and cmp.get("expected_kind") == "number"
                and cmp.get("got_kind") == "number"):
            item["hardcoded"] = True
        questions.append(item)

    verdicts = [q["verdict"] for q in questions]
    n_q = len(questions)
    n_match = verdicts.count("match")
    n_answered = sum(1 for q in questions if q["located"] and q["verdict"] != "missing")
    result = {
        "status": "ok",
        "rules_version": rules.RULES_VERSION,
        "n_questions": n_q,
        "n_match": n_match,
        "n_mismatch": verdicts.count("mismatch"),
        "n_missing": verdicts.count("missing"),
        "n_unlocated": len(attempt.unlocated_qids),
        "n_answered": n_answered,
        "n_hardcoded": sum(1 for q in questions if q["hardcoded"]),
        "n_fuzzy_matched": sum(1 for q in questions if q["match_kind"] == "fuzzy"),
        "fraction_correct": (n_match / n_q) if n_q else None,
        "rules_fired": _count(q["rule"] for q in questions if q["rule"]),
        "flags": _count(f for q in questions for f in q["flags"]),
        "golden_sheets": golden.sheets,
        "attempt_sheets": attempt.sheets_used,
        "attempt_sheets_without_answer_column": attempt.sheets_without_answer_column,
        "questions": questions,
    }
    return result


def _count(it) -> dict:
    out: dict = {}
    for k in it:
        out[k] = out.get(k, 0) + 1
    return dict(sorted(out.items()))


# ---------------------------------------------------------------------------
# Harness verdicts for the scoring layer
# ---------------------------------------------------------------------------


def harness_verdicts(result: dict, hardcoded_counts: bool = True) -> dict:
    """Translate an answer-check result into per-check harness decisions.

    Returns {"<Category>/<check name>": {engine: "harness"|"llm", decision,
    summary, mistakes[], fallback_reason, ...stats}} for the two checks the
    harness can speak to. `engine: "llm"` means "could not measure — the
    judge's own verdict stands", with the reason recorded.
    """
    fa_key = "/".join(CHECK_FINAL_ACCURACY)
    dc_key = "/".join(CHECK_COMPLETENESS)
    base = {
        fa_key: {"engine": "llm", "fallback_reason": None},
        dc_key: {"engine": "llm", "fallback_reason": None},
    }
    status = result.get("status")
    if status != "ok":
        reason = f"answer check status {status!r}: {result.get('reason') or result.get('error') or ''}".strip()
        for k in base:
            base[k]["fallback_reason"] = reason
        return base

    n_q = result.get("n_questions") or 0
    if n_q == 0:
        for k in base:
            base[k]["fallback_reason"] = "golden has no questions"
        return base

    located = n_q - (result.get("n_unlocated") or 0)
    share = located / n_q
    stats = {
        "n_questions": n_q,
        "n_match": result["n_match"],
        "n_mismatch": result["n_mismatch"],
        "n_missing": result["n_missing"],
        "n_unlocated": result["n_unlocated"],
        "n_answered": result["n_answered"],
        "n_hardcoded": result["n_hardcoded"],
        "fraction_correct": result["fraction_correct"],
        "rules_fired": result["rules_fired"],
        "flags": result["flags"],
        "hardcoded_counts": hardcoded_counts,
        "rules_version": result.get("rules_version"),
    }

    # --- Final calculation accuracy -----------------------------------
    fa = dict(base[fa_key], **stats)
    if share < MEASURE_MIN_LOCATED_SHARE:
        fa["fallback_reason"] = (
            f"only {located}/{n_q} golden questions located in the attempt "
            f"(< {MEASURE_MIN_LOCATED_SHARE:.0%}); layout not trusted"
        )
    else:
        mistakes = []
        for q in result["questions"]:
            loc = q["attempt_cell"] or q["golden_cell"]
            if q["verdict"] == "missing":
                mistakes.append({
                    "location": loc,
                    "description": f"Unanswered: \"{_short(q['label'])}\" — expected {q['expected']!r}.",
                    "severity": "major",
                })
            elif q["verdict"] == "mismatch":
                extra = ""
                if "possible_unit_scale_difference" in q["flags"]:
                    extra = " Possible unit-scale difference (x1000/x1e6) — review."
                if "sign_flip_not_outflow" in q["flags"]:
                    extra += " Sign flipped on a non-outflow row."
                mistakes.append({
                    "location": loc,
                    "description": (
                        f"Wrong answer to \"{_short(q['label'])}\": expected "
                        f"{q['expected']!r}, got {q['got']!r} ({q['detail']}).{extra}"
                    ),
                    "severity": "major",
                })
            if hardcoded_counts and q["hardcoded"]:
                mistakes.append({
                    "location": loc,
                    "description": (
                        f"Hardcoded answer to \"{_short(q['label'])}\": the cell holds "
                        f"the constant {q['got']!r} instead of a live formula."
                    ),
                    "severity": "minor",
                })
        fa["engine"] = "harness"
        fa["decision"] = "pass" if not mistakes else "fail"
        fa["mistakes"] = mistakes
        fa["summary"] = (
            f"Harness answer check: {result['n_match']}/{n_q} answers match the golden "
            f"within the rulebook ({result['n_mismatch']} wrong, {result['n_missing']} "
            f"unanswered, {result['n_hardcoded']} hardcoded"
            f"{'' if hardcoded_counts else ' (not counted)'}). "
            f"Rules fired: {result['rules_fired'] or 'none'}."
        )
    base[fa_key] = fa

    # --- Deliverable completeness: harness only decides the zero case --
    dc = dict(base[dc_key], n_questions=n_q, n_answered=result["n_answered"])
    if result["n_answered"] == 0:
        dc["engine"] = "harness"
        dc["decision"] = "fail"
        dc["summary"] = f"Harness: 0 of {n_q} Questions-sheet answers are present."
        dc["mistakes"] = [{
            "location": result["golden_sheets"][0]["name"] if result.get("golden_sheets") else "Questions",
            "description": f"None of the {n_q} required question answers is present in the attempt.",
            "severity": "major",
        }]
    else:
        dc["fallback_reason"] = (
            f"{result['n_answered']}/{n_q} answers present; completeness beyond the "
            f"Questions sheet is the judge's call"
        )
    base[dc_key] = dc
    return base


def _short(s: Any, n: int = 90) -> str:
    s = str(s or "")
    return s if len(s) <= n else s[: n - 1] + "…"


# ---------------------------------------------------------------------------
# Entry points
# ---------------------------------------------------------------------------


def run_answer_check(attempt_xlsx, solution_xlsx, output_json_path=None,
                     hardcoded_counts: bool = True) -> dict:
    """Full check: extract both sides, compare, derive harness verdicts,
    optionally write the artifact. Never raises."""
    try:
        golden = extract_golden(Path(solution_xlsx), allow_recalc=True)
        if golden.status != "ok":
            result = {
                "status": "convention_not_found",
                "side": "golden",
                "reason": golden.reason,
                "recalc_used": golden.recalc_used,
            }
        else:
            attempt = extract_attempt(Path(attempt_xlsx), golden, allow_recalc=True)
            if attempt.status != "ok":
                result = {
                    "status": attempt.status,
                    "side": "attempt",
                    "reason": attempt.reason,
                    "golden_sheets": golden.sheets,
                    "n_questions": len(golden.questions),
                    "attempt_sheets": attempt.sheets_used,
                    "recalc_used": golden.recalc_used or attempt.recalc_used,
                }
            else:
                result = compare_answer_sets(golden, attempt)
                result["recalc_used"] = bool(golden.recalc_used or attempt.recalc_used)
                result["solution_recalc_used"] = golden.recalc_used
                result["attempt_recalc_used"] = attempt.recalc_used
                if golden.recalc_error or attempt.recalc_error:
                    result["recalc_error"] = golden.recalc_error or attempt.recalc_error
        result["harness_verdicts"] = harness_verdicts(result, hardcoded_counts)
    except Exception as e:  # noqa: BLE001 — deliberately never blocks grading
        logger.warning(f"  [answer_check] unexpected failure: {e}")
        result = {"status": "error", "error": str(e)}
        result["harness_verdicts"] = harness_verdicts(result, hardcoded_counts)

    if output_json_path:
        try:
            Path(output_json_path).parent.mkdir(parents=True, exist_ok=True)
            Path(output_json_path).write_text(json.dumps(result, indent=2, default=str))
        except OSError as e:
            logger.warning(f"  [answer_check] could not write artifact: {e}")
    return result


def summary_block(result: dict) -> dict:
    """The compact block recorded in scored_results.answer_check."""
    keys = ("status", "reason", "n_questions", "n_match", "n_mismatch", "n_missing",
            "n_unlocated", "n_answered", "n_hardcoded", "fraction_correct",
            "rules_fired", "flags", "recalc_used", "rules_version")
    out = {k: result.get(k) for k in keys if k in result}
    hv = result.get("harness_verdicts") or {}
    out["harness"] = {
        k: {kk: v.get(kk) for kk in ("engine", "decision", "fallback_reason")}
        for k, v in hv.items()
    }
    return out
