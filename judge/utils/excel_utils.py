import argparse
import base64
import colorsys
import csv
import io
import math
import os
import re
import shutil
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

from .logger import logger
from .misc_utils import load_env_var

NOGRADE_PREFIX = "_nograde_"


def get_worksheet_info(workbook: openpyxl.Workbook, sheet_name: str) -> Dict[str, Any]:
    """Get high level information about a worksheet."""
    worksheet = workbook[sheet_name]
    return {
        "name": sheet_name,
        "max_row": worksheet.max_row,
        "max_column": worksheet.max_column,
        "dimensions": f"{worksheet.max_row} rows x {worksheet.max_column} columns",
    }


def load_workbooks(excel_file_path: str) -> Tuple[openpyxl.Workbook, openpyxl.Workbook]:
    """Load both formula and data-only workbooks."""

    # read_only does not keep merged cell, so we can't use the flag for workbook
    workbook = openpyxl.load_workbook(excel_file_path, data_only=False)
    workbook_data_only = openpyxl.load_workbook(
        excel_file_path, data_only=True, read_only=True
    )
    return workbook, workbook_data_only


### Cell processing functions - start
### Excel-style number rendering ############################################
# These helpers render a numeric cell value the way Excel itself would
# display it (e.g. -179166.67 under '\$#,##0;"($"#,##0\);\-' becomes
# '($179,167)'). They are best-effort: _render_number_format returns
# None for any format it does not handle, so _get_formatted_value can
# fall back to the legacy '<raw> [FORMAT:<pattern>]' representation.


def _split_format_segments(format_string: str) -> List[str]:
    """Split an Excel format string on top-level ';' separators.
    Semicolons inside quoted strings or directly after a backslash are
    treated as literal."""
    segments: List[str] = []
    current: List[str] = []
    in_quotes = False
    i = 0
    while i < len(format_string):
        c = format_string[i]
        if c == "\\" and i + 1 < len(format_string):
            current.append(c)
            current.append(format_string[i + 1])
            i += 2
            continue
        if c == '"':
            in_quotes = not in_quotes
            current.append(c)
            i += 1
            continue
        if c == ";" and not in_quotes:
            segments.append("".join(current))
            current = []
            i += 1
            continue
        current.append(c)
        i += 1
    segments.append("".join(current))
    return segments


def _is_date_or_time_format(format_string: str) -> bool:
    """True when the format contains d/y/h/s outside literal sections.
    Used to defer date/time rendering to the legacy fallback."""
    cleaned: List[str] = []
    in_quotes = False
    i = 0
    while i < len(format_string):
        c = format_string[i]
        if c == "\\" and i + 1 < len(format_string):
            i += 2
            continue
        if c == '"':
            in_quotes = not in_quotes
            i += 1
            continue
        if not in_quotes:
            cleaned.append(c.lower())
        i += 1
    return bool(re.search(r"[dymhs]", "".join(cleaned)))


def _tokenize_format_segment(segment: str) -> List[Tuple[str, str]]:
    """Tokenize a segment into ('number', text) and ('literal', text) parts.
    Number tokens are contiguous runs of #, 0, ., , and %. Everything
    else (escaped chars, quoted strings, $, parens, dashes, letters)
    becomes a literal."""
    tokens: List[Tuple[str, str]] = []
    cur_lit: List[str] = []
    cur_num: List[str] = []

    def flush_lit() -> None:
        if cur_lit:
            tokens.append(("literal", "".join(cur_lit)))
            cur_lit.clear()

    def flush_num() -> None:
        if cur_num:
            tokens.append(("number", "".join(cur_num)))
            cur_num.clear()

    i = 0
    while i < len(segment):
        c = segment[i]
        if c == "\\" and i + 1 < len(segment):
            flush_num()
            cur_lit.append(segment[i + 1])
            i += 2
            continue
        if c == '"':
            flush_num()
            i += 1
            while i < len(segment) and segment[i] != '"':
                cur_lit.append(segment[i])
                i += 1
            i += 1
            continue
        if c in "#0.,%":
            flush_lit()
            cur_num.append(c)
            i += 1
            continue
        flush_num()
        cur_lit.append(c)
        i += 1

    flush_num()
    flush_lit()
    return tokens


def _round_half_away_from_zero(value: float, decimals: int) -> float:
    """Round to *decimals* places using half-away-from-zero (Excel's rule)."""
    if value == 0:
        return 0.0
    multiplier = 10 ** decimals
    if value > 0:
        return math.floor(value * multiplier + 0.5) / multiplier
    return -math.floor(-value * multiplier + 0.5) / multiplier


def _format_number_template(value: float, template: str) -> str:
    """Render *value* using a number-only template like '#,##0.00' or '0.00%'."""
    has_percent = "%" in template
    template_no_pct = template.replace("%", "")
    rendered_value = value * 100 if has_percent else value

    if "." in template_no_pct:
        int_part, frac_part = template_no_pct.split(".", 1)
        decimals = sum(1 for c in frac_part if c in "0#")
    else:
        int_part = template_no_pct
        decimals = 0

    use_thousands = "," in int_part
    rounded = _round_half_away_from_zero(rendered_value, decimals)

    if decimals > 0:
        formatted = (
            f"{rounded:,.{decimals}f}" if use_thousands else f"{rounded:.{decimals}f}"
        )
    else:
        int_value = int(rounded)
        formatted = f"{int_value:,d}" if use_thousands else f"{int_value:d}"

    if has_percent:
        formatted += "%"
    return formatted


def _apply_format_segment(value: float, segment: str) -> str:
    """Render *value* using one segment (i.e. one of pos/neg/zero)."""
    tokens = _tokenize_format_segment(segment)

    number_template = next((t for k, t in tokens if k == "number"), None)
    if number_template is None:
        # Pure-literal segment, e.g. '\-' for zero -> just emit the literals.
        return "".join(t for _, t in tokens)

    formatted_number = _format_number_template(value, number_template)

    out: List[str] = []
    replaced = False
    for kind, text in tokens:
        if kind == "number":
            if not replaced:
                out.append(formatted_number)
                replaced = True
            # additional number tokens (rare) are dropped
        else:
            out.append(text)
    return "".join(out)


def _render_number_format(value, format_string: str) -> Optional[str]:
    """Render *value* as Excel would display it, or None to defer to fallback.

    Each early return below is a path we deliberately do NOT render;
    returning None hands control back to the legacy
    '<raw> [FORMAT:<pattern>]' representation in _get_formatted_value.
    """

    # Path A: no format / default format / text format -> defer.
    if format_string is None or format_string == "":
        return None
    if format_string in ("General", "@"):
        return None

    # Path B: color codes ([Red], [Blue], ...) or conditional ([>0]) -> defer.
    if "[" in format_string:
        return None

    # Path C: scientific notation ('0.00E+00') -> defer.
    if "E+" in format_string or "E-" in format_string:
        return None

    # Path D: date/time formats (any d/y/h/s outside literals) -> defer.
    if _is_date_or_time_format(format_string):
        return None

    # Path E: render. Pick the segment matching the value's sign.
    segments = _split_format_segments(format_string)
    if value < 0 and len(segments) >= 2 and segments[1].strip():
        # Negative-specific segment exists; it usually carries its own
        # sign indication (parens or literal '-'), so render |value|.
        return _apply_format_segment(abs(value), segments[1])
    if value == 0 and len(segments) >= 3 and segments[2].strip():
        return _apply_format_segment(value, segments[2])
    # Positive segment also handles negatives when no neg segment exists
    # (Excel prefixes a '-' sign in that case; _format_number_template
    # produces a signed string for negative inputs).
    return _apply_format_segment(value, segments[0])


### End Excel-style number rendering ########################################


def _get_formatted_value(cell, cell_data_only, _cached_config=None) -> str:
    """Get the formatted display value of a cell, preserving number formatting.

    Two paths:
      Path 1 (render-as-displayed): for numeric cells whose format is
        recognised by _render_number_format, return what Excel would
        display, e.g. '($179,167)' or '$40,000,000' or '3.38%'.
      Path 2 (legacy fallback): otherwise emit the original
        '<raw_value> [FORMAT:<pattern>]' form. Reached for None, text,
        dates, General/@, conditional/color formats, scientific
        notation, and anything else the renderer punts on.

    _cached_config: optional (do_rounding, float_rounding, percentage_rounding) tuple
    to skip per-cell env/YAML reads when the caller has already loaded them.
    """

    # --- Path 1: render-as-displayed (numeric values only) ----------------
    raw_value_for_render = cell_data_only.value
    if isinstance(raw_value_for_render, (int, float)):
        fmt = getattr(cell, "number_format", None)
        if fmt:
            rendered = _render_number_format(raw_value_for_render, fmt)
            if rendered is not None:
                return rendered

    # --- Path 2: legacy fallback (original implementation, unchanged) -----

    def _default_value(raw_value, float_rounding: int, do_rounding: bool) -> str:
        if isinstance(raw_value, float):
            if do_rounding:
                rounded_value = round(raw_value, float_rounding)
                formatted_value = f"{rounded_value}"
            else:
                formatted_value = str(raw_value)
        else:
            formatted_value = str(raw_value)
        return formatted_value

    if cell_data_only.value is None:
        return ""
    # Obtain parameters
    if _cached_config is not None:
        do_rounding, float_rounding, percentage_rounding = _cached_config
    else:
        do_rounding = load_env_var("JUDGE_DO_ROUNDING", "true").lower() == "true"
        if do_rounding:
            float_rounding = int(load_env_var("JUDGE_FLOAT_ROUNDING", 6))
            percentage_rounding = int(load_env_var("JUDGE_PERCENTAGE_ROUNDING", 8))

    # Get the raw value
    raw_value = cell_data_only.value

    # Check if this is a number with formatting
    if (
        isinstance(raw_value, (int, float))
        and hasattr(cell, "number_format")
        and cell.number_format
    ):
        # Get the number format from the cell
        number_format = cell.number_format

        # If it's not the default format, include format information
        if number_format != "General" and number_format != "@":

            # Get the number format from the cell
            number_format = cell.number_format
            parts = number_format.split(";")
            positive_format = parts[0] if len(parts) > 0 else None  # "$"#,##0_)
            negative_format = (
                parts[1] if len(parts) > 1 else None
            )  # [Red]\\("$"#,##0\\)
            zero_format = parts[2] if len(parts) > 2 else None  # (optional)

            if raw_value < 0 and negative_format:
                number_format = negative_format
                FORMAT_PREFIX = "NEG FORMAT"
            elif raw_value == 0 and zero_format:
                number_format = zero_format
                FORMAT_PREFIX = "ZERO FORMAT"
            else:
                number_format = positive_format
                FORMAT_PREFIX = "FORMAT"

            # For currency, percentage, and other special formats, show the format
            if "$" in number_format or "%" in number_format or "#,##0" in number_format:
                if do_rounding:
                    if "%" in number_format:
                        rounded_value = round(raw_value, percentage_rounding)
                    else:
                        rounded_value = round(
                            raw_value, float_rounding
                        )  # Round for context saving purposes
                    formatted_value = (
                        f"{rounded_value} [{FORMAT_PREFIX}:{number_format}]"
                    )
                else:
                    formatted_value = f"{raw_value} [{FORMAT_PREFIX}:{number_format}]"

    # If formatted value is not assigned for some reason, fallback to raw value as string
    if "formatted_value" not in locals() or formatted_value is None:
        formatted_value = _default_value(raw_value, float_rounding, do_rounding)
    return formatted_value


def _get_excel_cell_reference(row_idx: int, col_idx: int) -> str:
    """Convert row and column indices to Excel cell reference (e.g., A1, B19)."""
    if type(col_idx) != int:
        return f"Col {col_idx} Row {row_idx}"
    # Convert column index to Excel column letters
    col_letters = ""
    col_num = col_idx
    while col_num > 0:
        col_num -= 1
        col_letters = chr(65 + (col_num % 26)) + col_letters
        col_num //= 26

    return f"{col_letters}{row_idx}"


def _cell_has_value_or_color(cell, cell_data_only) -> bool:
    """Check if cell has value or color formatting."""
    # Check for actual value
    display_value = (
        str(cell_data_only.value) if cell_data_only.value is not None else ""
    )
    if display_value.strip():
        return True

    # Check for font color - RGB first, then indexed converted to RGB
    if cell.font and cell.font.color:
        try:
            if hasattr(cell.font.color, "rgb") and cell.font.color.rgb:
                color_value = str(cell.font.color.rgb)
                if (
                    color_value != "00000000"
                    and color_value != "Values must be of type <class 'str'>"
                ):
                    return True
            elif (
                hasattr(cell.font.color, "indexed")
                and cell.font.color.indexed is not None
            ):
                return True  # We'll convert indexed to RGB
        except (AttributeError, TypeError):
            pass

    # Check for background color - RGB first, then indexed converted to RGB
    if cell.fill:
        try:
            if hasattr(cell.fill, "start_color") and cell.fill.start_color:
                if hasattr(cell.fill.start_color, "rgb") and cell.fill.start_color.rgb:
                    bg_color = str(cell.fill.start_color.rgb)
                    if bg_color != "00000000":
                        return True
                elif (
                    hasattr(cell.fill.start_color, "indexed")
                    and cell.fill.start_color.indexed is not None
                ):
                    return True  # We'll convert indexed to RGB
        except (AttributeError, TypeError):
            pass

    return False


def _rgb_to_color_name(rgb_value: str) -> str:
    """Convert RGB value to human-readable color name using complete HSV analysis."""
    # Clean up RGB value - remove FF prefix if it's 8 characters
    if len(rgb_value) == 8 and rgb_value.startswith("FF"):
        rgb_clean = rgb_value[2:]
    elif len(rgb_value) == 8:
        # Handle ARGB format - take last 6 chars for RGB
        rgb_clean = rgb_value[2:]
    else:
        rgb_clean = rgb_value

    # Ensure we have exactly 6 characters
    if len(rgb_clean) != 6:
        return rgb_clean

    try:
        # Parse RGB values
        r = int(rgb_clean[0:2], 16)
        g = int(rgb_clean[2:4], 16)
        b = int(rgb_clean[4:6], 16)

        # Convert to HSV using colorsys
        h, s, v = colorsys.rgb_to_hsv(r / 255, g / 255, b / 255)
        h = h * 360  # Convert to degrees

        # First check saturation and value for achromatic colors
        if s < 0.1:
            if v < 0.2:
                color_name = "black"
            elif v > 0.9:
                color_name = "white"
            else:
                color_name = "gray"

        # Low value = dark colors
        elif v < 0.4:  # Increased threshold for better dark color detection
            if h < 30 or h >= 330:
                color_name = "dark_red"
            elif h < 75:  # Extended range for brown detection
                color_name = "brown"
            elif h < 150:
                color_name = "dark_green"
            elif h < 270:
                color_name = "dark_blue"
            else:
                color_name = "dark_purple"

        # Low saturation but not achromatic = muted colors
        elif s < 0.4:  # Slightly increased threshold
            if h < 60:
                color_name = "beige"
            elif h < 150:
                color_name = "olive"
            elif h < 210:
                color_name = "slate_blue"
            else:
                color_name = "muted_purple"

        # High saturation and value = bright colors
        elif s > 0.7 and v > 0.7:
            if h < 15 or h >= 345:
                color_name = "bright_red"
            elif h < 45:
                color_name = "bright_orange"
            elif h < 65:
                color_name = "bright_yellow"
            elif h < 170:
                color_name = "bright_green"
            elif h < 200:
                color_name = "bright_cyan"
            elif h < 260:
                color_name = "bright_blue"
            elif h < 310:
                color_name = "bright_purple"
            else:
                color_name = "bright_magenta"

        # Medium to high saturation with high value = pastel colors
        elif s < 0.7 and v > 0.7:  # Better pastel detection
            if h < 15 or h >= 330:
                color_name = "pink"
            elif h < 45:
                color_name = "peach"
            elif h < 65:
                color_name = "pale_yellow"
            elif h < 170:
                color_name = "pale_green"
            elif h < 200:
                color_name = "pale_cyan"
            elif h < 260:
                color_name = "pale_blue"
            elif h < 310:
                color_name = "lavender"
            else:
                color_name = "pale_magenta"

        # Finally use hue for normal saturated colors
        else:
            if h < 15 or h >= 345:
                color_name = "red"
            elif h < 45:
                color_name = "orange"
            elif h < 65:
                color_name = "yellow"
            elif h < 170:
                color_name = "green"
            elif h < 200:
                color_name = "cyan"
            elif h < 260:
                color_name = "blue"
            elif h < 310:
                color_name = "purple"
            else:
                color_name = "magenta"

        return f"{color_name} ({rgb_clean})"

    except ValueError:
        # Fallback to just returning the hex code if parsing fails
        return rgb_clean


def _convert_indexed_to_rgb(indexed_val: int) -> str:
    """Convert indexed color to RGB with custom overrides for specific indices."""
    # Custom indexed color overrides
    CUSTOM_INDEXED_COLORS = {
        13: "00FF00",  # Override index 13 (was FFFF00/yellow) -> green for A26
        14: "FF6600",  # Override index 14 (was FF00FF/magenta) -> orange for A27
    }

    # Check for custom override first
    if indexed_val in CUSTOM_INDEXED_COLORS:
        return CUSTOM_INDEXED_COLORS[indexed_val]

    # Otherwise use standard COLOR_INDEX
    from openpyxl.styles.colors import COLOR_INDEX

    if indexed_val < len(COLOR_INDEX):
        argb_hex = COLOR_INDEX[indexed_val]
        if len(argb_hex) == 8:
            return argb_hex[2:]  # Skip alpha (AA) part, take RRGGBB

    return "000000"  # Default to black if conversion fails


def extract_cell_formatting(cell) -> str:
    """Extract formatting information from a cell."""
    formatting_parts = []

    try:
        # Font formatting
        if cell.font:
            font = cell.font
            if font.name:
                formatting_parts.append(f"font:{font.name}")
            if font.size:
                formatting_parts.append(f"size:{font.size}")
            if font.bold:
                formatting_parts.append("bold")
            if font.italic:
                formatting_parts.append("italic")
            if font.underline:
                formatting_parts.append("underline")

            # Handle font color safely with explicit textcolor label and color names - RGB FIRST
            if font.color:
                try:
                    color_added = False

                    # RGB first priority - direct RGB colors
                    if hasattr(font.color, "rgb") and font.color.rgb:
                        color_value = str(font.color.rgb)
                        # Only use RGB if it's actually valid (not error or default)
                        if (
                            color_value != "00000000"
                            and color_value != "None"
                            and "Values must be of type" not in color_value
                        ):
                            color_name = _rgb_to_color_name(color_value)
                            formatting_parts.append(f"textcolor:{color_name}")
                            color_added = True

                    # If no RGB available or RGB invalid, convert indexed to RGB first, then apply HSV
                    if (
                        not color_added
                        and hasattr(font.color, "indexed")
                        and font.color.indexed is not None
                    ):
                        try:
                            indexed_val = int(
                                font.color.indexed
                            )  # Convert to int to handle openpyxl Integer objects
                            # Use custom conversion function with overrides
                            rgb_hex = _convert_indexed_to_rgb(indexed_val)
                            # Now apply HSV analysis to the converted RGB
                            color_name = _rgb_to_color_name(rgb_hex)
                            formatting_parts.append(f"textcolor:{color_name}")
                        except (ValueError, TypeError):
                            pass  # Skip if conversion fails

                except (AttributeError, TypeError):
                    pass  # Skip problematic color values

        # Fill/background color with explicit bgcolor label and color names - RGB FIRST
        if cell.fill:
            try:
                if hasattr(cell.fill, "start_color") and cell.fill.start_color:
                    bg_color_added = False

                    # RGB first priority - direct RGB colors
                    if (
                        hasattr(cell.fill.start_color, "rgb")
                        and cell.fill.start_color.rgb
                    ):
                        bg_color = str(cell.fill.start_color.rgb)
                        # Only use RGB if it's actually valid (not error or default)
                        if (
                            bg_color != "00000000"
                            and "Values must be of type" not in bg_color
                        ):
                            color_name = _rgb_to_color_name(bg_color)
                            formatting_parts.append(f"bgcolor:{color_name}")
                            bg_color_added = True

                    # If no RGB available or RGB invalid, convert indexed to RGB first, then apply HSV
                    if not bg_color_added and (
                        hasattr(cell.fill.start_color, "indexed")
                        and cell.fill.start_color.indexed is not None
                    ):
                        try:
                            indexed_val = int(
                                cell.fill.start_color.indexed
                            )  # Convert to int to handle openpyxl Integer objects
                            # Use custom conversion function with overrides
                            rgb_hex = _convert_indexed_to_rgb(indexed_val)
                            # Now apply HSV analysis to the converted RGB
                            color_name = _rgb_to_color_name(rgb_hex)
                            formatting_parts.append(f"bgcolor:{color_name}")
                        except (ValueError, TypeError):
                            pass  # Skip if conversion fails
            except (AttributeError, TypeError):
                pass  # Skip problematic fill values

        # Alignment (removed wrap_text to exclude 'wrap' from formatting)
        if cell.alignment:
            align = cell.alignment
            if align.horizontal and align.horizontal != "general":
                formatting_parts.append(f"halign:{align.horizontal}")
            if align.vertical and align.vertical != "bottom":
                formatting_parts.append(f"valign:{align.vertical}")
            # Removed: if align.wrap_text: formatting_parts.append("wrap")

        # Borders
        if cell.border:
            border_parts = []
            for side in ["top", "bottom", "left", "right"]:
                try:
                    border_side = getattr(cell.border, side)
                    if (
                        border_side
                        and hasattr(border_side, "style")
                        and border_side.style
                    ):
                        border_parts.append(f"{side}:{border_side.style}")
                except (AttributeError, TypeError):
                    continue
            if border_parts:
                formatting_parts.append(f"border:{','.join(border_parts)}")

    except Exception as e:
        # If any error occurs, return empty formatting rather than crash
        logger.info(f"Warning: Error extracting formatting for cell: {e}")
        return ""

    return ";".join(formatting_parts)


def create_enhanced_cell(
    cell,
    cell_data_only,
    row_idx: int,
    col_idx: int,
    _cached_config=None,
) -> str:
    """Create enhanced cell data with embedded formula and formatting."""
    # Start with formatted display value (includes number formatting)
    display_value = _get_formatted_value(cell, cell_data_only, _cached_config)

    # For non-empty cells, start with cell reference
    if display_value.strip():
        cell_ref = _get_excel_cell_reference(row_idx, col_idx)
        cell_parts = [f"[{cell_ref}]{display_value}"]
    else:
        cell_parts = [display_value]

    # Add formula if present. Plain formulas come back as strings starting
    # with "="; array formulas (Ctrl+Shift+Enter) come back as ArrayFormula
    # objects whose .text is the formula and .ref is the spilled range.
    formula_text = None
    if isinstance(cell.value, str) and cell.value.startswith("="):
        formula_text = cell.value
    elif isinstance(cell.value, openpyxl.worksheet.formula.ArrayFormula):
        formula_text = cell.value.text
    if formula_text is not None:
        cell_parts.append(f"FORMULA:{formula_text}")

    # Add formatting if present and cell has value or color formatting
    if _cell_has_value_or_color(cell, cell_data_only):
        formatting = extract_cell_formatting(cell)
        if formatting:
            cell_parts.append(f"FORMAT:{formatting}")

    return "|".join(cell_parts)


def list_to_csv(data: List[List[str]]) -> str:
    """Convert list of lists to CSV string."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerows(data)
    return output.getvalue()


def extract_all_cell_data(worksheet, worksheet_data) -> Dict[str, Any]:
    """Extract everything into one CSV.

    JUDGE_FAST_CELL_EXTRACT (default "true") selects between:
      - fast: parallel iter_rows() on both sheets + one-time env var load
      - slow: original per-cell worksheet_data.cell() lookup (kept for parity)
    """
    use_fast = load_env_var("JUDGE_FAST_CELL_EXTRACT", "true").lower() == "true"

    enhanced_data = []

    if use_fast:
        # In read-only mode, worksheet_data.cell(r, c) re-scans the XML from the
        # start on each call, and _get_formatted_value reloads project_configs.yaml
        # per cell. Walk both sheets in lockstep and load env vars once.
        do_rounding = load_env_var("JUDGE_DO_ROUNDING", "true").lower() == "true"
        float_rounding = (
            int(load_env_var("JUDGE_FLOAT_ROUNDING", 6)) if do_rounding else 6
        )
        percentage_rounding = (
            int(load_env_var("JUDGE_PERCENTAGE_ROUNDING", 8)) if do_rounding else 8
        )
        cached_config = (do_rounding, float_rounding, percentage_rounding)

        for row_idx, (row, row_data) in enumerate(
            zip(worksheet.iter_rows(), worksheet_data.iter_rows()), 1
        ):
            enhanced_row = []
            for col_idx, (cell, cell_data_only) in enumerate(zip(row, row_data), 1):
                enhanced_cell = create_enhanced_cell(
                    cell,
                    cell_data_only,
                    row_idx,
                    col_idx,
                    _cached_config=cached_config,
                )
                enhanced_row.append(enhanced_cell)
            enhanced_data.append(enhanced_row)
    else:
        for row_idx, row in enumerate(worksheet.iter_rows(), 1):
            enhanced_row = []
            for col_idx, cell in enumerate(row, 1):
                cell_data_only = worksheet_data.cell(row_idx, col_idx)
                enhanced_cell = create_enhanced_cell(
                    cell,
                    cell_data_only,
                    row_idx,
                    col_idx,
                )
                enhanced_row.append(enhanced_cell)
            enhanced_data.append(enhanced_row)

    return {
        "full": list_to_csv(enhanced_data),
        "metadata": {
            "format": "Enhanced with cell references, formulas and formatting",
            "separator": "|",
            "rows": len(enhanced_data),
            "columns": len(enhanced_data[0]) if enhanced_data else 0,
        },
    }


### Cell processing functions - end
### Additional Info Extraction functions - start
def create_safe_filename(sheet_name: str) -> str:
    """Create a safe filename from sheet name."""
    return "".join(c for c in sheet_name if c.isalnum() or c in (" ", "-", "_")).strip()


def extract_merged_cells_info(worksheet) -> List[str]:
    """Extract merged cell ranges from a worksheet."""
    merged_cells = []
    for merged_range in worksheet.merged_cells.ranges:
        # Convert merged range to string format like "A1:C3"
        merged_cells.append(str(merged_range))
    return merged_cells


def extract_frozen_panes_info(worksheet) -> Dict[str, Any]:
    """Extract frozen panes information from a worksheet."""
    frozen_info = {"has_frozen_panes": False, "freeze_panes": None, "split_panes": None}

    def safe_round(value):
        try:
            return round(value)
        except (TypeError, ValueError):
            return None

    try:
        # Extract freeze pane location from worksheet.sheet_view.pane
        if hasattr(worksheet, "sheet_view") and worksheet.sheet_view:
            sheet_view = worksheet.sheet_view
            if hasattr(sheet_view, "pane") and sheet_view.pane:
                pane = sheet_view.pane
                if pane.xSplit or pane.ySplit:
                    frozen_info["has_frozen_panes"] = True
                    frozen_info["freeze_panes"] = _get_excel_cell_reference(
                        safe_round(pane.ySplit), safe_round(pane.xSplit)
                    )
                    frozen_info["split_panes"] = {
                        "xSplit": pane.xSplit,
                        "ySplit": pane.ySplit,
                        "topLeftCell": pane.topLeftCell,
                        "activePane": pane.activePane,
                        "state": pane.state,
                    }
    except (AttributeError, TypeError):
        # Handle cases where properties don't exist or have unexpected types
        import traceback

        logger.info(
            f"Warning: Error extracting frozen panes info: {traceback.format_exc()}"
        )

        pass

    return frozen_info


### Additional Info Extraction functions - end


def save_additional_format_info(
    output_dir: Path, sheet_name: str, worksheet, quiet: bool = False
) -> str:
    """Save additional formatting information (merged cells, frozen panes) to a text file."""
    output_dir.mkdir(exist_ok=True)
    safe_sheet_name = create_safe_filename(sheet_name)

    # Create the additional format file
    format_file_path = output_dir / f"{safe_sheet_name}_additional_format.txt"

    # Extract information
    merged_cells = extract_merged_cells_info(worksheet)
    frozen_panes = extract_frozen_panes_info(worksheet)

    # Create content
    content_lines = []
    content_lines.append(f"Additional Format Information for Sheet: {sheet_name}")
    content_lines.append("=" * 60)
    content_lines.append("")

    # Merged Cells Section
    content_lines.append("MERGED CELLS:")
    content_lines.append("-" * 20)
    if merged_cells:
        for i, merged_range in enumerate(merged_cells, 1):
            content_lines.append(f"{i}. {merged_range}")
    else:
        content_lines.append("No merged cells found.")
    content_lines.append("")

    # Frozen Panes Section
    content_lines.append("FROZEN PANES:")
    content_lines.append("-" * 20)
    if frozen_panes["has_frozen_panes"]:
        if frozen_panes["freeze_panes"]:
            content_lines.append(f"Freeze Panes: {frozen_panes['freeze_panes']}")
        if frozen_panes["split_panes"]:
            content_lines.append("Split Panes Information:")
            split_info = frozen_panes["split_panes"]
            for key, value in split_info.items():
                if value is not None:
                    content_lines.append(f"  {key}: {value}")
    else:
        content_lines.append("No frozen panes found.")
    content_lines.append("")

    # Metadata
    content_lines.append("METADATA:")
    content_lines.append("-" * 20)
    content_lines.append(
        f"Sheet Dimensions: {worksheet.max_row} rows x {worksheet.max_column} columns"
    )
    content_lines.append(f"Total Merged Cell Ranges: {len(merged_cells)}")
    content_lines.append(f"Has Frozen Panes: {frozen_panes['has_frozen_panes']}")

    # Write to file
    content = "\n".join(content_lines)
    format_file_path.write_text(content, encoding="utf-8")

    if not quiet:
        logger.info(f"Saved additional format info: {format_file_path}")

    return str(format_file_path)


def save_sheet_csv_files(
    output_dir: Path,
    sheet_name: str,
    extraction_result: Dict[str, Any],
    worksheet=None,
    quiet: bool = False,
) -> List[str]:
    """Save CSV files and additional format info for a single sheet to directory."""
    output_dir.mkdir(exist_ok=True)
    safe_sheet_name = create_safe_filename(sheet_name)
    saved_files = []

    # Save based on extraction result content

    if "full" in extraction_result:
        enhanced_path = output_dir / f"{safe_sheet_name}_full.csv"
        enhanced_path.write_text(extraction_result["full"], encoding="utf-8")
        saved_files.append(str(enhanced_path))
        if not quiet:
            logger.info(f"Saved: {enhanced_path}")

    # Save additional format information (merged cells, frozen panes)
    if worksheet is not None:
        additional_format_path = save_additional_format_info(
            output_dir, sheet_name, worksheet, quiet=quiet
        )
        saved_files.append(additional_format_path)

    return saved_files


def process_all_worksheets(
    excel_file_path: str,
    output_dir: Path,
    quiet: bool = False,
    run_calculation: bool = False,
    sheet_name_filter: callable = None,
) -> Dict[str, Any]:
    """
    Process all worksheets in an Excel file and extract to CSV.

    Args:
        excel_file_path: Path to Excel file
        output_dir: Directory to save CSV files

    Returns:
        Dictionary with results for each sheet
    """

    def recalculate_xlsx(filepath: str, outdir: str = "."):
        """Re-save xlsx through LibreOffice to trigger formula calculation."""
        ### Note: This function assumes LibreOffice is installed and added to PATH, or the path is provided through environment variable. Otherwise, call the function with run_calculation=False to skip recalculation step.
        import subprocess

        subprocess.run(
            [
                load_env_var("PATHS_LIBREOFFICE_PATH", required=True),
                "--headless",
                "--calc",
                "--convert-to",
                "xlsx",
                "--outdir",
                outdir,
                filepath,
            ],
            check=True,
            cwd=os.getcwd(),
        )

    if run_calculation:
        # Sometimes. Excel files created by openpyxl won't calculate the cached values from the formulas. This process triggers a recalculation through Libreoffice.
        # Make recalulated file under same directory
        temp_output_dir = Path(excel_file_path).parent / "temp_recalculated"
        temp_output_dir.mkdir(exist_ok=True)
        logger.info(
            f"INFO: Recalculating Excel file using LibreOffice: {excel_file_path}. Putting recalculated file in {temp_output_dir}"
        )
        recalculate_xlsx(excel_file_path, str(temp_output_dir))
        recalculated_file = temp_output_dir / Path(excel_file_path).name
        if recalculated_file.exists():
            excel_file_path = str(recalculated_file)
        else:
            raise FileNotFoundError(
                f"Recalculated file not found: {recalculated_file}. Something may have gone wrong with LibreOffice conversion."
            )
    if not output_dir.exists():
        output_dir.mkdir(parents=True, exist_ok=True)

    # Load workbooks
    workbook, workbook_data_only = load_workbooks(excel_file_path)

    # Get all sheet names
    sheet_names = workbook.sheetnames
    if not quiet:
        logger.info(f"Found {len(sheet_names)} worksheets: {sheet_names}")

    results = {}
    all_saved_files = []

    # Process each worksheet
    for sheet_name in sheet_names:
        # Apply sheet name filter if provided
        output_sheet_name = sheet_name
        if sheet_name_filter is not None:
            filtered_name = sheet_name_filter(sheet_name)
            if filtered_name is None:
                if not quiet:
                    logger.info(
                        f"\n=== Skipping sheet (filtered out): {sheet_name} ==="
                    )
                continue
            output_sheet_name = filtered_name

        if not quiet:
            if output_sheet_name != sheet_name:
                logger.info(
                    f"\n=== Processing sheet: {sheet_name} -> {output_sheet_name} ==="
                )
            else:
                logger.info(f"\n=== Processing sheet: {sheet_name} ===")

        # Get worksheets
        worksheet = workbook[sheet_name]
        worksheet_data = workbook_data_only[sheet_name]

        # Get sheet info
        if type(worksheet) is openpyxl.chartsheet.chartsheet.Chartsheet:
            logger.info(f"Skipping ChartSheet: {sheet_name}")
            continue

        sheet_info = get_worksheet_info(workbook, sheet_name)
        if not quiet:
            logger.info(f"Dimensions: {sheet_info['dimensions']}")

        # Extract data
        extraction_result = extract_all_cell_data(worksheet, worksheet_data)

        # Save files
        saved_files = save_sheet_csv_files(
            output_dir, output_sheet_name, extraction_result, worksheet, quiet=quiet
        )

        # Store results
        results[output_sheet_name] = {
            "info": sheet_info,
            "extraction": extraction_result,
            "saved_files": saved_files,
        }

        all_saved_files.extend(saved_files)

    # Summary
    if not quiet:
        logger.info(f"\n=== Summary ===")
        logger.info(f"Processed {len(sheet_names)} worksheets")
        logger.info(f"Generated {len(all_saved_files)} CSV files")
        logger.info(f"Output directory: {output_dir.absolute()}")

    return {
        "sheets": results,
        "summary": {
            "total_sheets": len(sheet_names),
            "total_files": len(all_saved_files),
            "output_directory": str(output_dir.absolute()),
            "all_files": all_saved_files,
        },
    }


### File I/O Helper Functions ###############################################


def encode_file_to_base64(file_path: Path) -> tuple[str, str]:
    """Encode a file to base64 and return the encoded string and MIME type."""
    file_path = Path(file_path)
    with open(file_path, "rb") as f:
        file_bytes = f.read()
        base64_encoded = base64.b64encode(file_bytes).decode("utf-8")

    suffix = file_path.suffix.lower()
    mime_types = {
        ".csv": "text/csv",
        ".txt": "text/plain",
        ".pdf": "application/pdf",
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
    }
    mime_type = mime_types.get(suffix, "application/octet-stream")

    return base64_encoded, mime_type


def find_context_files_simple(source_dir: Path) -> Optional[Path]:
    """Find the first text or PDF file in the source directory or task subfolder."""
    context_files = []
    context_files.extend(source_dir.glob("*.pdf"))
    context_files.extend(source_dir.glob("*.txt"))

    task_dir = source_dir / "task"
    if task_dir.exists() and task_dir.is_dir():
        context_files.extend(task_dir.glob("*.pdf"))
        context_files.extend(task_dir.glob("*.txt"))

    if not context_files:
        return None

    return context_files[0]


def find_and_merge_context_files(
    source_dir: Path,
    output_dir: Path,
    merge_pdfs: bool = True,
) -> Optional[Path]:
    """Discover context files and write a single merged file to *output_dir*.

    Looks in *source_dir* and *source_dir*/"task" for .pdf and .txt files.
    PDFs take priority over TXTs when both exist. Files of the chosen type
    are merged in sorted-filename order: PDFs via pypdf page concatenation,
    TXTs via newline-separated concatenation with filename headers.

    With merge_pdfs=False, falls back to the legacy single-file behavior
    (pick the first file via find_context_files_simple).

    Returns the destination path (output_dir/context.pdf or
    output_dir/context.txt), or None if no context file was found.
    """
    if not merge_pdfs:
        context_file = find_context_files_simple(source_dir)
        if not context_file or not context_file.exists():
            return None
        ext = context_file.suffix.lower()
        if ext == ".pdf":
            dest = output_dir / "context.pdf"
        elif ext == ".txt":
            dest = output_dir / "context.txt"
        else:
            dest = output_dir / context_file.name
        shutil.copy(str(context_file), str(dest))
        logger.info(f"Context file: {context_file.name}")
        return dest

    pdfs = list(source_dir.glob("*.pdf"))
    txts = list(source_dir.glob("*.txt"))
    task_dir = source_dir / "task"
    if task_dir.exists() and task_dir.is_dir():
        pdfs.extend(task_dir.glob("*.pdf"))
        txts.extend(task_dir.glob("*.txt"))

    pdfs.sort(key=lambda p: p.name)
    txts.sort(key=lambda p: p.name)

    # Dedup: when an answer-containing PDF is present (e.g. "Questions with
    # Answers.pdf"), drop the question-only sibling (e.g. "Questions.pdf").
    # The answer variant is a superset, so including both wastes context.
    has_answer_pdf = any("answer" in p.name.lower() for p in pdfs)
    if has_answer_pdf:
        question_only = [
            p
            for p in pdfs
            if "question" in p.name.lower() and "answer" not in p.name.lower()
        ]
        if question_only:
            pdfs = [p for p in pdfs if p not in question_only]
            logger.info(
                f"  Dropping question-only PDF(s) in favor of answer variant: "
                f"{[p.name for p in question_only]}"
            )

    if pdfs:
        if txts:
            logger.warning(
                f"  Ignoring {len(txts)} .txt context file(s) because PDFs are "
                f"present: {[t.name for t in txts]}"
            )
        dest = output_dir / "context.pdf"
        if len(pdfs) == 1:
            shutil.copy(str(pdfs[0]), str(dest))
            logger.info(f"  Context PDF: {pdfs[0].name}")
            return dest

        from pypdf import PdfWriter

        writer = PdfWriter()
        included, skipped = [], []
        for p in pdfs:
            try:
                writer.append(str(p))
                included.append(p.name)
            except Exception as e:
                skipped.append((p.name, str(e)))
                logger.warning(f"  Skipping PDF {p.name} during merge: {e}")
        if not included:
            writer.close()
            logger.error("  All context PDFs failed to merge; no context written")
            return None
        with open(dest, "wb") as f:
            writer.write(f)
        writer.close()
        logger.info(
            f"  Merged {len(included)} context PDFs into {dest.name} "
            f"(order: {included})"
        )
        if skipped:
            logger.warning(f"  Skipped {len(skipped)} PDF(s) during merge: {skipped}")
        return dest

    if txts:
        dest = output_dir / "context.txt"
        if len(txts) == 1:
            shutil.copy(str(txts[0]), str(dest))
            logger.info(f"  Context TXT: {txts[0].name}")
            return dest

        parts = []
        included = []
        for t in txts:
            try:
                with open(t, "r", encoding="utf-8") as f:
                    parts.append(f"--- {t.name} ---\n\n{f.read()}")
                included.append(t.name)
            except Exception as e:
                logger.warning(f"  Skipping TXT {t.name} during merge: {e}")
        if not parts:
            return None
        with open(dest, "w", encoding="utf-8") as f:
            f.write("\n\n".join(parts))
        logger.info(
            f"  Merged {len(included)} context TXTs into {dest.name} "
            f"(order: {included})"
        )
        return dest

    return None


def find_golden_solution_file(source_dir: Path) -> Path:
    """Find the golden solution file in the source directory.

    When multiple Excel files exist, picks the one with the shortest filename.
    This typically selects the base solution file over versioned variants
    (e.g., 'Solution.xlsx' over 'Solution - Q19.xlsx').
    """
    solution_dir = source_dir / "solution"
    excel_exts = [".xlsx", ".xlsm", ".xls"]

    # Check dedicated solution folder first
    if solution_dir.exists() and solution_dir.is_dir():
        excel_files = [
            item
            for item in solution_dir.iterdir()
            if item.suffix.lower() in excel_exts and item.is_file()
        ]
        if excel_files:
            selected = min(excel_files, key=lambda p: len(p.name))
            if len(excel_files) > 1:
                logger.info(
                    f"  Multiple solution files, picking shortest: {selected.name}"
                )
            return selected

    # Look for files with "solution" or "answer" in name
    candidates = []
    for item in source_dir.iterdir():
        if item.is_file() and item.suffix.lower() in excel_exts:
            name = item.name.lower()
            if "solution" in name or "answer" in name:
                candidates.append(item)
    if candidates:
        selected = min(candidates, key=lambda p: len(p.name))
        if len(candidates) > 1:
            logger.info(f"  Multiple solution files, picking shortest: {selected.name}")
        return selected

    # Fallback: first Excel file found
    excel_files = [
        item
        for item in source_dir.iterdir()
        if item.is_file() and item.suffix.lower() in excel_exts
    ]
    if excel_files:
        return min(excel_files, key=lambda p: len(p.name))

    raise FileNotFoundError(
        "No golden solution Excel file found in the expected places."
    )


def copy_support_files(
    source_dir: Path,
    output_dir: Path,
    default_rubric_path: str = None,
    default_weights_path: str = None,
    merge_context_pdfs: bool = True,
):
    """Copy context, rubric, and perturbations files from source to output directory.

    Args:
        source_dir: Task source directory containing original files.
        output_dir: Destination directory for copied files.
        default_rubric_path: Fallback path for rubric.json if not found in source_dir.
        default_weights_path: Fallback path for rubric_weights.json if not found in source_dir.
        merge_context_pdfs: When True (default), concatenate all PDFs (or all
            TXTs) found in the source into a single context file. When False,
            only the first context file is used (legacy behavior).
    """
    copied_files = []

    context_dest = find_and_merge_context_files(
        source_dir, output_dir, merge_pdfs=merge_context_pdfs
    )
    if context_dest:
        copied_files.append(str(context_dest))
    else:
        logger.info("Context file not found")

    rubric_json_source = source_dir / "rubric.json"
    if rubric_json_source.exists():
        rubric_json_dest = output_dir / "rubric.json"
        shutil.copy(str(rubric_json_source), str(rubric_json_dest))
        copied_files.append(str(rubric_json_dest))
    elif default_rubric_path and Path(default_rubric_path).exists():
        rubric_json_dest = output_dir / "rubric.json"
        shutil.copy(str(default_rubric_path), str(rubric_json_dest))
        copied_files.append(str(rubric_json_dest))
    else:
        logger.error(
            "Required file rubric.json not found in source directory or default path."
        )

    perturbations_json_source = source_dir / "perturbations.json"
    if perturbations_json_source.exists():
        perturbations_json_dest = output_dir / "perturbations.json"
        shutil.copy(str(perturbations_json_source), str(perturbations_json_dest))
        copied_files.append(str(perturbations_json_dest))

    # Copy rubric weights file
    weights_json_source = source_dir / "rubric_weights.json"
    if weights_json_source.exists():
        weights_json_dest = output_dir / "rubric_weights.json"
        shutil.copy(str(weights_json_source), str(weights_json_dest))
        copied_files.append(str(weights_json_dest))
    elif default_weights_path and Path(default_weights_path).exists():
        weights_json_dest = output_dir / "rubric_weights.json"
        shutil.copy(str(default_weights_path), str(weights_json_dest))
        copied_files.append(str(weights_json_dest))

    return copied_files


def prepare_directory_files(directory_path: str) -> dict:
    """Prepare all CSV files and their corresponding TXT files in a directory."""
    directory = Path(directory_path)
    if not directory.exists() or not directory.is_dir():
        raise NotADirectoryError(f"Not a directory: {directory}")

    files_dict = {}
    csv_files = sorted(directory.glob("*.csv"), key=lambda f: f.stat().st_ctime)

    for csv_file in csv_files:
        files_dict[csv_file.name] = str(csv_file)

        base_name = csv_file.stem.replace("_full", "")
        txt_file = directory / f"{base_name}_additional_format.txt"

        if txt_file.exists():
            files_dict[txt_file.name] = str(txt_file)

    return files_dict


def _attempt_sheet_name_filter(sheet_name: str) -> Optional[str]:
    """Filter attempt sheets: keep those starting with 'answers_' or 'model_' (stripping
    the prefix) for grading.  All other sheets are kept with a '_nograde_' prefix so they
    are still included in the judge context but flagged as not-to-be-graded."""
    if sheet_name.startswith("answers_"):
        return sheet_name[len("answers_") :]
    if sheet_name.startswith("model_"):
        return sheet_name[len("model_") :]
    return f"{NOGRADE_PREFIX}{sheet_name}"


def process_case_files(
    file_paths,
    task_folder: str,
    use_existing: bool = False,
    run_calculation: bool = False,
    attempt_sheet_name_filter: bool = False,
    quiet: bool = False,
):
    """Process multiple Excel files into enhanced CSV files.

    Args:
        file_paths: List of Excel file paths to process
        task_folder: Path to the task folder
        use_existing: If True, skip processing if CSV files already exist in the output folder
        run_calculation: If True, run Excel formula calculations before extracting CSVs.
        attempt_sheet_name_filter: If True, only keep attempt sheets starting with
            'answers_' or 'model_', stripping the prefix from the output name.
    """
    workbook_dirs = {}

    # Create output directory in task_folder/judge_results
    task_path = Path(task_folder)
    file_output_dir = task_path / "judge_results"
    file_output_dir.mkdir(parents=True, exist_ok=True)

    for file_path in file_paths:
        if not quiet:
            logger.info(f"\nProcessing file: {file_path}")
        file_path = Path(file_path)

        workbook_stem = file_path.stem
        workbook_dir = file_output_dir / workbook_stem
        workbook_dir.mkdir(parents=True, exist_ok=True)

        # Check if CSV files already exist in the workbook directory
        existing_csvs = list(workbook_dir.glob("*.csv"))
        if use_existing and existing_csvs:
            if not quiet:
                logger.info(
                    f"    Skipping CSV extraction (files exist): {len(existing_csvs)} CSV files found in {workbook_dir}/"
                )
            workbook_dirs[workbook_stem] = str(workbook_dir)
            continue

        # Apply sheet name filter for attempt files if enabled
        sheet_filter = None
        if attempt_sheet_name_filter and file_path.name == "ai_attempt.xlsx":
            sheet_filter = _attempt_sheet_name_filter

        process_all_worksheets(
            excel_file_path=str(file_path),
            output_dir=workbook_dir,
            quiet=quiet,
            run_calculation=run_calculation and file_path.name == "ai_attempt.xlsx",
            sheet_name_filter=sheet_filter,
        )
        workbook_dirs[workbook_stem] = str(workbook_dir)

    return {
        "folder_name": task_path.name,
        "output_dir": file_output_dir,
        "workbook_dirs": workbook_dirs,
    }


### CSV Shortening Functions ################################################


def get_csv_dimensions(file_path: Path) -> tuple:
    """Get the number of rows and columns in a CSV file.

    Returns:
        tuple: (rows, cols) or (None, None) if unable to read
    """
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)
            if not rows:
                return (0, 0)
            num_rows = len(rows)
            num_cols = max(len(row) for row in rows) if rows else 0
            return (num_rows, num_cols)
    except Exception:
        return (None, None)


def _clean_csv_content(content: str) -> str:
    """Clean CSV content by removing duplicate commas and trailing commas."""
    lines = content.split("\n")
    processed_lines = []

    for line in lines:
        # Collapse any sequence of 2+ commas to single comma
        new_line = re.sub(r",{2,}", ",", line)
        # Remove trailing commas
        stripped_line = new_line.rstrip(",")
        processed_lines.append(stripped_line)

    return "\n".join(processed_lines)


def _shorten_csv_files_base(
    directory_path: Path,
    target_chars: int,
    steps: list,
    keep_first_rows: int = 60,
    keep_last_rows: int = 40,
    min_rows_threshold: int = 100,
    keep_first_cols: int = 50,
    keep_last_cols: int = 25,
    min_cols_threshold: int = 75,
    clean_csv: bool = True,
) -> dict:
    """Base function to shorten CSV files in a directory using a step-based approach.

    Steps are executed in order and stop early if total character count falls below target.

    Returns:
        dict with total_original, total_shortened, steps_executed, per_file info.
    """
    if not directory_path.exists():
        return {
            "total_original": 0,
            "total_shortened": 0,
            "steps_executed": 0,
            "per_file": {},
        }

    # Collect file info for full CSVs and additional format txt files
    file_infos: dict[str, dict] = {}
    file_contents: dict[str, str] = {}

    csv_files = sorted(directory_path.glob("*_full.csv"))
    for csv_file in csv_files:
        try:
            with open(csv_file, "r", encoding="utf-8") as f:
                content = f.read()
            rows, cols = get_csv_dimensions(csv_file)
            file_infos[csv_file.name] = {
                "chars": len(content),
                "rows": rows,
                "cols": cols,
                "rows_redacted": 0,
                "cols_redacted": 0,
                "shortened_chars": len(content),
                "original_rows": rows,
                "original_cols": cols,
            }
            file_contents[csv_file.name] = content
        except Exception:
            pass

    txt_files = sorted(directory_path.glob("*_additional_format.txt"))
    for txt_file in txt_files:
        try:
            with open(txt_file, "r", encoding="utf-8") as f:
                content = f.read()
            num_lines = content.count("\n") + (
                1 if content and not content.endswith("\n") else 0
            )
            file_infos[txt_file.name] = {
                "chars": len(content),
                "rows": num_lines,
                "cols": None,
                "rows_redacted": 0,
                "cols_redacted": 0,
                "shortened_chars": len(content),
                "original_rows": num_lines,
                "original_cols": None,
            }
            file_contents[txt_file.name] = content
        except Exception:
            pass

    if not file_infos:
        return {
            "total_original": 0,
            "total_shortened": 0,
            "steps_executed": 0,
            "per_file": {},
        }

    total_original = sum(info["chars"] for info in file_infos.values())

    def _calculate_total_chars() -> int:
        return sum(len(content) for content in file_contents.values())

    def _estimate_cleaned_chars() -> int:
        total = 0
        for content in file_contents.values():
            cleaned = _clean_csv_content(content)
            total += len(cleaned)
        return total

    def _is_under_target() -> bool:
        current = _calculate_total_chars()
        if clean_csv:
            estimated_after_clean = _estimate_cleaned_chars()
            logger.info(
                f"    Current total chars: {current:,}. "
                f"After cleaning: {estimated_after_clean:,}. Target: {target_chars:,}"
            )
            return estimated_after_clean < target_chars
        else:
            logger.info(
                f"    Current total chars: {current:,}. Target: {target_chars:,}"
            )
            if current < target_chars:
                logger.info("    Target met based on current chars.")
            return current < target_chars

    def _step_redact_middle_rows(
        threshold: int = None,
        first: int = None,
        last: int = None,
    ) -> None:
        _threshold = threshold if threshold is not None else min_rows_threshold
        _first = first if first is not None else keep_first_rows
        _last = last if last is not None else keep_last_rows

        eligible_files = [
            f
            for f in file_infos.keys()
            if file_infos[f].get("rows_redacted", 0) == 0 and file_contents.get(f, "")
        ]
        if not eligible_files:
            return

        largest_filename = max(
            eligible_files, key=lambda f: len(file_contents.get(f, ""))
        )
        content = file_contents.get(largest_filename, "")
        if not content:
            return

        logger.info(f"    Redacting rows from file: {largest_filename}")
        lines = content.split("\n")
        while lines and lines[-1] == "":
            lines.pop()
        num_lines = len(lines)

        if num_lines > _threshold:
            first_section = lines[:_first]
            last_section = lines[-_last:] if _last > 0 else []

            first_redacted = _first + 1
            last_redacted = num_lines - _last
            rows_redacted = last_redacted - first_redacted + 1

            redaction_msg = f"\n[Rows {first_redacted}-{last_redacted} redacted ({rows_redacted} rows)]\n"

            file_contents[largest_filename] = (
                "\n".join(first_section) + redaction_msg + "\n".join(last_section)
            )
            file_infos[largest_filename]["rows_redacted"] = rows_redacted

    def _step_redact_middle_columns(
        threshold: int = None,
        first: int = None,
        last: int = None,
    ) -> None:
        _threshold = threshold if threshold is not None else min_cols_threshold
        _first = first if first is not None else keep_first_cols
        _last = last if last is not None else keep_last_cols

        eligible_files = [
            f
            for f in file_infos.keys()
            if file_infos[f].get("cols_redacted", 0) == 0 and file_contents.get(f, "")
        ]
        if not eligible_files:
            return

        largest_filename = max(
            eligible_files, key=lambda f: len(file_contents.get(f, ""))
        )
        content = file_contents.get(largest_filename, "")
        if not content:
            return

        num_cols = file_infos[largest_filename].get("cols", 0)
        if num_cols is None or num_cols <= _threshold:
            return

        logger.info(f"    Redacting columns from file: {largest_filename}")

        lines = content.split("\n")
        new_lines = []

        first_redacted_col = _first + 1
        last_redacted_col = num_cols - _last
        cols_redacted = last_redacted_col - first_redacted_col + 1

        for line in lines:
            if line.startswith("[") and "redacted" in line:
                new_lines.append(line)
                continue

            cells = line.split(",")
            if len(cells) > _threshold:
                first_cells = cells[:_first]
                last_cells = cells[-_last:] if _last > 0 else []
                redaction_msg = (
                    f"[Columns {first_redacted_col}-{last_redacted_col} redacted]"
                )
                new_line = (
                    ",".join(first_cells)
                    + ","
                    + redaction_msg
                    + ","
                    + ",".join(last_cells)
                )
                new_lines.append(new_line)
            else:
                new_lines.append(line)

        file_contents[largest_filename] = "\n".join(new_lines)
        file_infos[largest_filename]["cols_redacted"] = cols_redacted

    def _step_clean_csv_content() -> None:
        for filename in file_contents:
            content = file_contents[filename]
            if content:
                file_contents[filename] = _clean_csv_content(content)

    step_functions = {
        "redact_rows": _step_redact_middle_rows,
        "redact_cols": _step_redact_middle_columns,
    }

    steps_executed = 0
    for step_type, step_name in steps:
        if _is_under_target():
            break
        logger.info(f"    Executing step: {step_name}")
        step_func = step_functions.get(step_type)
        if step_func:
            step_func()
        steps_executed += 1

    if clean_csv:
        _step_clean_csv_content()

    # Finalize: update file_infos and overwrite original files with shortened content
    for filename, info in file_infos.items():
        content = file_contents.get(filename, "")
        shortened_chars = len(content)
        info["shortened_chars"] = shortened_chars
        info["chars_saved"] = info.get("chars", 0) - shortened_chars

        if content:
            original_path = directory_path / filename
            with open(original_path, "w", encoding="utf-8") as f:
                f.write(content)

    total_shortened = sum(len(content) for content in file_contents.values())

    return {
        "total_original": total_original,
        "total_shortened": total_shortened,
        "steps_executed": steps_executed,
        "per_file": file_infos,
    }


# Default step sequences for different file types
SOLUTION_SHORTENING_STEPS = [
    ("redact_rows", "redact_rows"),
    ("redact_cols", "redact_cols"),
    ("redact_rows", "redact_rows"),
    ("redact_cols", "redact_cols"),
    ("redact_rows", "redact_rows"),
    ("redact_cols", "redact_cols"),
    ("redact_rows", "redact_rows"),
    ("redact_rows", "redact_rows"),
    ("redact_rows", "redact_rows"),
    ("redact_rows", "redact_rows"),
]

ATTEMPT_SHORTENING_STEPS = [
    ("redact_rows", "redact_rows"),
    ("redact_cols", "redact_cols"),
    ("redact_rows", "redact_rows"),
    ("redact_cols", "redact_cols"),
    ("redact_rows", "redact_rows"),
    ("redact_cols", "redact_cols"),
    ("redact_rows", "redact_rows"),
    ("redact_rows", "redact_rows"),
    ("redact_rows", "redact_rows"),
    ("redact_rows", "redact_rows"),
    ("redact_rows", "redact_rows"),
    ("redact_rows", "redact_rows"),
    ("redact_rows", "redact_rows"),
    ("redact_rows", "redact_rows"),
    ("redact_rows", "redact_rows"),
    ("redact_rows", "redact_rows"),
    ("redact_rows", "redact_rows"),
    ("redact_rows", "redact_rows"),
    ("redact_rows", "redact_rows"),
    ("redact_rows", "redact_rows"),
]


def shorten_solution_csv_files(
    directory_path: Path,
    target_chars: int,
) -> dict:
    """Shorten CSV files for golden solution context.

    Uses SOLUTION_SHORTENING_STEPS sequence optimized for solution files.
    Cleans CSV content by removing duplicate commas and trailing commas.
    """
    return _shorten_csv_files_base(
        directory_path=directory_path,
        target_chars=target_chars,
        steps=SOLUTION_SHORTENING_STEPS,
        clean_csv=True,
    )


def shorten_attempt_csv_files(
    directory_path: Path,
    target_chars: int,
) -> dict:
    """Shorten CSV files for AI attempt context.

    Uses ATTEMPT_SHORTENING_STEPS sequence optimized for attempt files.
    Does not clean CSV content (preserves original formatting).
    """
    return _shorten_csv_files_base(
        directory_path=directory_path,
        target_chars=target_chars,
        steps=ATTEMPT_SHORTENING_STEPS,
        clean_csv=False,
    )


def calculate_message_size_for_files(directory_path: Path) -> dict:
    """Calculate the message size for all CSV files in a directory.

    Returns:
        dict: {"total": int, "per_file": {filename: {"chars": int, "rows": int, "cols": int}}}
    """
    total_chars = 0
    per_file_counts = {}

    if not directory_path.exists():
        return {"total": 0, "per_file": {}}

    # Count full CSV files
    csv_files = sorted(directory_path.glob("*.csv"))
    for csv_file in csv_files:
        if csv_file.name.endswith("_full.csv"):
            try:
                with open(csv_file, "r", encoding="utf-8") as f:
                    content = f.read()
                    char_count = len(content)
                    total_chars += char_count
                    rows, cols = get_csv_dimensions(csv_file)

                    per_file_counts[csv_file.name] = {
                        "chars": char_count,
                        "rows": rows,
                        "cols": cols,
                    }
            except UnicodeDecodeError:
                base64_content, _ = encode_file_to_base64(csv_file)
                char_count = len(base64_content)
                total_chars += char_count
                per_file_counts[csv_file.name] = {
                    "chars": char_count,
                    "rows": None,
                    "cols": None,
                }

    # Count additional format txt files
    txt_files = sorted(directory_path.glob("*_additional_format.txt"))
    for txt_file in txt_files:
        try:
            with open(txt_file, "r", encoding="utf-8") as f:
                content = f.read()
                char_count = len(content)
                total_chars += char_count
                num_lines = content.count("\n") + (
                    1 if content and not content.endswith("\n") else 0
                )

                per_file_counts[txt_file.name] = {
                    "chars": char_count,
                    "rows": num_lines,
                    "cols": None,
                }
        except UnicodeDecodeError:
            base64_content, _ = encode_file_to_base64(txt_file)
            char_count = len(base64_content)
            total_chars += char_count
            per_file_counts[txt_file.name] = {
                "chars": char_count,
                "rows": None,
                "cols": None,
            }

    return {
        "total": total_chars,
        "per_file": per_file_counts,
    }


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Test csv extraction.")
    parser.add_argument(
        "--filename",
        "-f",
        nargs="?",
        default="test.xlsx",
        help="Excel file to process (default: test.xlsx)",
    )

    parser.add_argument(
        "--output-dir",
        "-o",
        default=None,
        help="Output directory for CSV files (default: ./extracted_csvs)",
    )
    parser.add_argument(
        "--run-recalc",
        action="store_true",
        help="Recalculate formulas using LibreOffice before extraction",
    )

    args = parser.parse_args()
    if args.output_dir is None:
        from misc_utils import load_project_configs, relative_path_from_project_root

        configs, project_prefix = load_project_configs()
        args.output_dir = (
            configs.get(f"{project_prefix}_PATHS_SCRATCH_PATH", "./scratch")
            + f"/extracted_csvs_output/{Path(args.filename).stem}"
        )
        args.output_dir = relative_path_from_project_root(args.output_dir)
        print(
            f"Output directory not provided. Using {args.output_dir} based on SCRATCH_PATH from project configs."
        )

    process_all_worksheets(
        excel_file_path=args.filename,
        output_dir=args.output_dir,
        quiet=False,
        run_calculation=args.run_recalc,
    )
