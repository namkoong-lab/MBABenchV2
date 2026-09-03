"""Offline tests for the uncached-formula pre-flight (utils/formula_cache.py).

Run: python tests_offline/test_formula_cache.py   (from judge/)

The encoding assertions are round-tripped through the real encoder
(`create_enhanced_cell_variants`) with stub cells rather than hand-written
strings, so a change to the cell format breaks these tests instead of silently
breaking the census.
"""

import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from utils.misc_utils import load_project_configs

load_project_configs()

from utils import formula_cache
from utils.excel_utils import create_enhanced_cell_variants

failures = []


def ok(cond, msg):
    print(f"{'OK ' if cond else 'FAIL'} {msg}")
    if not cond:
        failures.append(msg)


# ---------------------------------------------------------------- stub cells
class _Color:
    rgb = None
    indexed = None


class _Font:
    name = None
    size = None
    bold = False
    italic = False
    underline = False
    color = None


class _Align:
    horizontal = None
    vertical = None


class _Cell:
    """Minimal stand-in for an openpyxl cell (formula side)."""

    def __init__(self, value, number_format="General"):
        self.value = value
        self.number_format = number_format
        self.font = _Font()
        self.fill = None
        self.alignment = _Align()
        self.border = None


class _DataCell:
    """Minimal stand-in for the data_only side."""

    def __init__(self, value):
        self.value = value


def encode(formula, cached_value):
    """Return the *_full.csv field the real encoder produces for this cell."""
    full, _data = create_enhanced_cell_variants(
        _Cell(formula), _DataCell(cached_value), 4, 2
    )
    return full


# ------------------------------------------------------- encoding round-trip
print("\n[1] The census classifies what the real encoder emits")

cached_field = encode("=SUM(B1:B3)", 1234)
uncached_field = encode("=SUM(B1:B3)", None)
literal_field = encode(1234, 1234)
empty_field = encode(None, None)

ok(
    formula_cache._census_field(cached_field) == "cached",
    f"calculated formula -> cached   ({cached_field!r})",
)
ok(
    formula_cache._census_field(uncached_field) == "uncached",
    f"uncalculated formula -> uncached ({uncached_field!r})",
)
ok(
    formula_cache._census_field(literal_field) is None,
    f"literal value is not a formula cell ({literal_field!r})",
)
ok(
    formula_cache._census_field(empty_field) is None,
    f"empty cell is not a formula cell ({empty_field!r})",
)

# A cached error value is a real result, not a missing one.
err_field = encode("=1/0", "#DIV/0!")
ok(
    formula_cache._census_field(err_field) == "cached",
    f"cached error value counts as cached ({err_field!r})",
)

# Display text containing '|' must not be mistaken for a blank display half.
piped = "[B4]a|b|FORMULA:=T(1)"
ok(
    formula_cache._census_field(piped) == "cached",
    "display text containing '|' still counts as cached",
)
ok(
    formula_cache._census_field("   |FORMULA:=T(1)") == "uncached",
    "whitespace-only display half counts as uncached",
)


# ------------------------------------------------------------------- census
print("\n[2] census_csv_dir counts per sheet and in total")


def write_dir(root, sheets):
    d = Path(root)
    d.mkdir(parents=True, exist_ok=True)
    for name, rows in sheets.items():
        (d / f"{name}_full.csv").write_text(
            "\n".join(",".join(f'"{c}"' for c in row) for row in rows),
            encoding="utf-8",
        )
    return str(d)


with tempfile.TemporaryDirectory() as tmp:
    good = write_dir(
        Path(tmp) / "good",
        {"Calcs": [[encode("=A1+1", 5), encode(3, 3)], [encode("=A2*2", 9), ""]]},
    )
    c = formula_cache.census_csv_dir(good)
    ok(c["formula_cells"] == 2, f"counted 2 formula cells (got {c['formula_cells']})")
    ok(c["uncached_formula_cells"] == 0, "none uncached")
    ok(c["ratio"] == 0.0, "ratio 0.0")

    bad = write_dir(
        Path(tmp) / "bad",
        {"Calcs": [[encode("=A1+1", None), encode("=A2*2", None)]]},
    )
    c = formula_cache.census_csv_dir(bad)
    ok(c["ratio"] == 1.0, f"fully uncalculated workbook -> ratio 1.0 (got {c['ratio']})")
    ok("Calcs_full.csv" in c["sheets"], "per-sheet detail recorded")

    # A workbook with no formulas at all must not trip anything.
    literal = write_dir(Path(tmp) / "lit", {"Data": [[encode(1, 1), encode(2, 2)]]})
    c = formula_cache.census_csv_dir(literal)
    ok(c["formula_cells"] == 0 and c["ratio"] == 0.0, "no formulas -> no signal")

    ok(
        formula_cache.census_csv_dir(None)["formula_cells"] == 0,
        "missing directory censuses to zero rather than raising",
    )

    # ------------------------------------------------------------- check_case
    print("\n[3] check_case refuses, and the escape hatch is recorded")

    prov = formula_cache.check_case(good, good)
    ok(prov["offenders"] == [], "clean case passes")
    ok(prov["checked"] is True and "max_ratio" in prov, "provenance shape")

    raised = None
    try:
        formula_cache.check_case(bad, good)
    except formula_cache.FormulaCacheError as e:
        raised = str(e)
    ok(raised is not None, "uncalculated attempt refuses the grading")
    ok(
        raised and "--run-calculation" in raised,
        "refusal names the fix (--run-calculation)",
    )

    raised = None
    try:
        formula_cache.check_case(good, bad)
    except formula_cache.FormulaCacheError as e:
        raised = str(e)
    ok(raised is not None, "uncalculated SOLUTION also refuses")
    ok(raised and "solution" in raised, "refusal names which workbook failed")

    os.environ[formula_cache.SKIP_ENV] = "1"
    try:
        prov = formula_cache.check_case(bad, good)
        ok(prov.get("skipped_via_env") is True, "escape hatch records skipped_via_env")
        ok(prov["offenders"] == ["attempt"], "offenders recorded even when skipped")
    finally:
        del os.environ[formula_cache.SKIP_ENV]

    # ---------------------------------------------------- workbook census
    print("\n[4] workbook census tells blank results from never-calculated cells")

    import zipfile

    import openpyxl

    # openpyxl writes <f>..</f><v /> with no type attribute: never calculated.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["A2"] = "=A1+1"
    ws["A3"] = '=IF(A1>5,"x","")'
    uncalc_path = Path(tmp) / "uncalc.xlsx"
    wb.save(uncalc_path)
    c = formula_cache.census_workbook(uncalc_path)
    ok(c["available"] is True, "workbook census reads an openpyxl file")
    ok(c["formula_cells"] == 2, f"counted 2 formula cells (got {c['formula_cells']})")
    ok(c["ratio"] == 1.0, f"openpyxl never-calculated -> ratio 1.0 (got {c['ratio']})")

    # Excel / LibreOffice write calculated empty-string results as
    # <c t="str"><f>..</f><v></v></c>, and numeric results with a value.
    # Rewrite the sheet XML the way those producers would have saved it.
    excel_path = Path(tmp) / "excel_like.xlsx"
    with zipfile.ZipFile(uncalc_path) as src, zipfile.ZipFile(excel_path, "w") as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                xml = data.decode("utf-8")
                xml = xml.replace(
                    '<c r="A2"><f>A1+1</f><v /></c>',
                    '<c r="A2" t="n"><f>A1+1</f><v>2</v></c>',
                )
                xml = xml.replace(
                    '<c r="A3"><f>IF(A1&gt;5,"x","")</f><v /></c>',
                    '<c r="A3" t="str"><f>IF(A1&gt;5,"x","")</f><v></v></c>',
                )
                data = xml.encode("utf-8")
            dst.writestr(item, data)
    c = formula_cache.census_workbook(excel_path)
    ok(c["formula_cells"] == 2, "rewritten workbook still has 2 formula cells")
    ok(
        c["uncached_formula_cells"] == 0 and c["empty_string_results"] == 1,
        f"typed-str empty value counts as a calculated empty string, not uncached "
        f"(uncached={c['uncached_formula_cells']}, empties={c['empty_string_results']})",
    )
    ok(c["ratio"] == 0.0, "calculated workbook with blank results -> ratio 0.0")

    # A cell whose <v> is empty but untyped is still uncached (guard against
    # the str rule being applied too broadly).
    ok(
        formula_cache._census_cell_xml(' r="B1"', "<f>1+1</f><v></v>") == "uncached",
        "untyped empty <v></v> is uncached",
    )
    ok(
        formula_cache._census_cell_xml(' r="B1" t="str"', "<f>T(1)</f><v></v>")
        == "empty_string",
        'typed str empty <v> is a calculated ""',
    )
    ok(
        formula_cache._census_cell_xml(' r="B1" t="e"', "<f>1/0</f><v>#DIV/0!</v>")
        == "cached",
        "cached error value is cached",
    )
    ok(
        formula_cache._census_cell_xml(' r="B1" t="n"', "<v>5</v>") is None,
        "literal cell is not a formula",
    )

    ok(
        formula_cache.census_workbook(Path(tmp) / "missing.xlsx")["available"] is False,
        "missing workbook -> unavailable (falls back to CSV), not clean",
    )

    # ------------------------------------------ check_case picks the basis
    print("\n[5] check_case decides on the workbook census when one is given")

    # CSV says 'bad' (blank display halves) but the workbook says calculated:
    # the 2026-09-03 TBondII false refusal. Workbook wins.
    prov = formula_cache.check_case(bad, good, attempt_xlsx=excel_path)
    ok(prov["offenders"] == [], "blank-result workbook is NOT refused")
    ok(prov["attempt"]["basis"] == "workbook", "attempt basis recorded as workbook")
    ok(prov["attempt"]["csv"]["ratio"] == 1.0, "CSV census still recorded alongside")

    # Workbook genuinely uncalculated: refused regardless of CSV.
    raised = None
    try:
        formula_cache.check_case(good, good, attempt_xlsx=uncalc_path)
    except formula_cache.FormulaCacheError as e:
        raised = str(e)
    ok(raised is not None and "workbook census" in raised, "uncalculated workbook refused on the workbook census")

    # No workbook path: CSV decides, as before.
    prov = formula_cache.check_case(good, good)
    ok(prov["attempt"]["basis"] == "csv", "no workbook -> CSV basis")
    raised = None
    try:
        formula_cache.check_case(bad, good, attempt_xlsx=Path(tmp) / "missing.xlsx")
    except formula_cache.FormulaCacheError as e:
        raised = str(e)
    ok(raised is not None and "csv census" in raised, "unreadable workbook -> CSV decides (refuses)")

print("\n" + ("ALL FORMULA-CACHE CHECKS PASSED" if not failures else f"{len(failures)} FAILURE(S)"))
sys.exit(1 if failures else 0)
