"""Offline tests for the harness answer checker (utils/answer_check.py).

Run from judge/:  python tests_offline/test_answer_check.py
Builds small workbooks with openpyxl — no DB, S3 or LLM. Formula answers are
written without cached values (openpyxl cannot write them), so the checker's
own LibreOffice recalc path runs, exactly as it does on openpyxl-produced
attempts in production (soffice must be installed, as the judge requires). Layouts
mirror the four canary attempts: a standard 'Questions' sheet, an 'Answers'
sheet with the 297-style header row (question in column C, "Model value"
and "Answer" columns), a decoy 'Answer Map' sheet, and a split two-sheet
golden like task 54.
"""
import json
import sys
import tempfile
from pathlib import Path

import openpyxl

JUDGE = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(JUDGE))

from utils.misc_utils import load_project_configs  # noqa: E402

load_project_configs()

from utils import answer_check as AC  # noqa: E402

FAILS = []


def check(cond, msg):
    if not cond:
        FAILS.append(msg)
        print("FAIL:", msg)
    else:
        print("OK ", msg)


TMP = Path(tempfile.mkdtemp(prefix="answer_check_test_"))

QUESTIONS = [
    ("What will be the net worth difference (Doctor vs. Trade) in the final model period?", -1890487.51, "[$]"),
    ("What will the YE amount for Cash on hand for the Doctor in the year 2028?", 1000.0, "[$]"),
    ("What is the gross margin in 2030?", 0.4213, "[%]"),
    ("What are total operating expenses in 2031?", -500.25, "[$]"),
    ("Is the Doctor path better after tax?", "Yes", ""),
    ("What is the cash balance in 2040?", "Outside model horizon", "[$]"),
]


def write_golden(path, header="Questions (please round your answers to two decimal places)",
                 pct_format=True):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Questions"
    ws["A1"], ws["B1"], ws["C1"] = header, "Answers", "Unit"
    for i, (q, a, u) in enumerate(QUESTIONS, start=2):
        ws.cell(row=i, column=1, value=q)
        c = ws.cell(row=i, column=2, value=a)
        if pct_format and u == "[%]":
            c.number_format = "0.00%"
        ws.cell(row=i, column=3, value=u)
    # a phantom row (single-space question) and a blank row, like DoorstepDecisions
    ws.cell(row=len(QUESTIONS) + 2, column=1, value=" ")
    wb.create_sheet("Model")["A1"] = "model"
    wb.save(path)


def _set(ws, ref, value, formula=None):
    """Write a value; when `formula` is given, write the formula in the
    formula layer and inject the cached value via a data-only twin later."""
    ws[ref] = formula if formula else value
    return ref


def write_attempt_standard(path, answers, sheet="Questions", hardcode=False,
                           answer_header="Answers", shuffle=False, extra_decoy=False):
    """Golden-like layout; `answers` = list aligned to QUESTIONS (None = blank)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    ws["A1"], ws["B1"], ws["C1"] = "Questions (please round your answers to two decimal places)", answer_header, "Unit"
    order = list(range(len(QUESTIONS)))
    if shuffle:
        order = order[::-1]
    for row_i, qi in enumerate(order, start=2):
        q, _, u = QUESTIONS[qi]
        ws.cell(row=row_i, column=1, value=q)
        a = answers[qi]
        if a is not None:
            cell = ws.cell(row=row_i, column=2)
            if hardcode or isinstance(a, str):
                cell.value = a
            else:
                cell.value = f"=Model!A{row_i}"  # formula layer
        ws.cell(row=row_i, column=3, value=u)
    m = wb.create_sheet("Model")
    for row_i, qi in enumerate(order, start=2):
        a = answers[qi]
        if a is not None and not isinstance(a, str):
            m.cell(row=row_i, column=1, value=a)
    if extra_decoy:
        d = wb.create_sheet("Answer Map")
        d["A1"] = "Map of answers"
        for row_i, (q, _, _) in enumerate(QUESTIONS, start=2):
            d.cell(row=row_i, column=1, value=q)
            d.cell(row=row_i, column=2, value=999)
    wb.save(path)
    _inject_cached_values(path)


def _inject_cached_values(path):
    """openpyxl cannot write cached formula results; emulate a calculated
    file by rewriting formula cells' cached <v> with the referenced Model
    value. Done by loading data_only=False, resolving =Model!A<n> refs
    ourselves, and saving a values twin as the cached layer via the XML."""
    import re
    import shutil
    import zipfile

    wb = openpyxl.load_workbook(path)
    model = wb["Model"] if "Model" in wb.sheetnames else None
    fixes = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("=Model!") and model is not None:
                    ref = c.value.split("!")[1]
                    fixes[(ws.title, c.coordinate)] = model[ref].value
    wb.close()
    if not fixes:
        return
    # Patch the sheet XML: <c r="B2"><f>Model!A2</f></c> -> add <v>value</v>
    tmp = str(path) + ".tmp"
    with zipfile.ZipFile(path) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        # map sheet titles to xml paths via workbook.xml order
        wbxml = zin.read("xl/workbook.xml").decode()
        names = re.findall(r'<sheet[^>]*name="([^"]+)"[^>]*r:id="rId(\d+)"', wbxml)
        title_to_xml = {}
        rels = zin.read("xl/_rels/workbook.xml.rels").decode()
        for name, rid in names:
            m = re.search(rf'Id="rId{rid}"[^>]*Target="([^"]+)"', rels) or re.search(
                rf'Target="([^"]+)"[^>]*Id="rId{rid}"', rels)
            if m:
                title_to_xml[name] = "xl/" + m.group(1).lstrip("/").removeprefix("xl/")
        for item in zin.infolist():
            data = zin.read(item.filename)
            for (title, coord), val in fixes.items():
                if title_to_xml.get(title) == item.filename:
                    text = data.decode()
                    text = re.sub(
                        rf'(<c r="{coord}"[^>]*>)(<f>[^<]*</f>)(</c>)',
                        rf"\1\2<v>{val}</v>\3",
                        text,
                    )
                    data = text.encode()
            zout.writestr(item, data)
    shutil.move(tmp, path)


def write_attempt_297_style(path):
    """'Answers' sheet: title in B2, notes, header row 7 with question in C,
    Path/Metric/Year, 'Model value [$]' in G and 'Answer [$] (2 dp)' in H."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Answers"
    ws["B2"] = "Answers — case questions (live links to the model)"
    ws["B3"] = "US dollars, nominal. 'Answer' = ROUND(model value, 2)."
    hdr = ["No.", "Question (as stated in the case)", "Path", "Metric (model label)", "Year",
           "Model value [$]", "Answer [$] (2 dp)", "Unit"]
    for j, h in enumerate(hdr, start=2):
        ws.cell(row=7, column=j, value=h)
    m = wb.create_sheet("Model")
    for i, (q, a, u) in enumerate(QUESTIONS):
        r = 8 + i
        ws.cell(row=r, column=2, value=i + 1)
        ws.cell(row=r, column=3, value=q)
        ws.cell(row=r, column=4, value="Doctor")
        ws.cell(row=r, column=5, value="Cash")
        ws.cell(row=r, column=6, value=2028)
        if isinstance(a, str):
            ws.cell(row=r, column=7, value=a)
            ws.cell(row=r, column=8, value=a)
        else:
            m.cell(row=r, column=1, value=a * 1.0000001)   # model value, unrounded
            # "2 dp" applies to the rendered figure: a %-formatted fraction
            # keeps 4 decimals (42.13% == 0.4213), money keeps 2.
            m.cell(row=r, column=2, value=round(a, 4) if abs(a) < 1 else round(a, 2))
            ws.cell(row=r, column=7, value=f"=Model!A{r}")
            ws.cell(row=r, column=8, value=f"=Model!B{r}")
        ws.cell(row=r, column=9, value=u)
    wb.save(path)
    _inject_cached_values(path)


def write_split_golden(path):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Questions Task 1"
    ws1["A1"], ws1["B1"], ws1["C1"] = "Questions (please round your answers to two decimal places)", "Answers", "Unit"
    for i, (q, a, u) in enumerate(QUESTIONS[:3], start=2):
        ws1.cell(row=i, column=1, value=q); ws1.cell(row=i, column=2, value=a); ws1.cell(row=i, column=3, value=u)
    ws2 = wb.create_sheet("Questions Task 2")
    ws2["A1"], ws2["B1"], ws2["C1"] = "Questions", "Answers", "Number for reference"
    for i, (q, a, u) in enumerate(QUESTIONS[3:], start=2):
        ws2.cell(row=i, column=1, value=q); ws2.cell(row=i, column=2, value=a); ws2.cell(row=i, column=3, value=-4e9)
    wb.save(path)


GOLD = TMP / "golden.xlsx"
write_golden(GOLD)
CORRECT = [-1890487.51, 1000.0, 0.4213, -500.25, "Yes", "n/a"]

# ---------------------------------------------------------------------------
# 1. Golden vs itself: everything matches, nothing hardcoded (golden has constants — hardcoded requires a golden formula)
# ---------------------------------------------------------------------------
r = AC.run_answer_check(GOLD, GOLD)
check(r["status"] == "ok" and r["n_questions"] == 6 and r["n_match"] == 6, "golden vs golden: 6/6 match")
check(r["n_hardcoded"] == 0, "constants vs a constant golden are not 'hardcoded'")
check(r["harness_verdicts"]["Accuracy/Final calculation accuracy"]["decision"] == "pass", "final accuracy pass")

# ---------------------------------------------------------------------------
# 2. Standard layout, all correct via formulas (penny/percent/sign/sentinel forms)
# ---------------------------------------------------------------------------
A1 = TMP / "a_standard.xlsx"
write_attempt_standard(A1, [-1890487.505, 1000.004, 42.13, 500.25, "yes", "N/A"])
r = AC.run_answer_check(A1, GOLD)
check(r["status"] == "ok", "standard layout parsed")
rules = r["rules_fired"]
check(r["n_match"] == 6, f"all six equivalent forms match (rules fired: {rules})")
check(rules.get("tolerance", 0) >= 2 and rules.get("percent_form") == 1 and rules.get("sign_outflow") == 1
      and rules.get("sentinel_synonym") == 1, "expected rules fired: tolerance, percent_form, sign_outflow, sentinel")
check(r["n_hardcoded"] == 0, "formula answers are not hardcoded")
fa = r["harness_verdicts"]["Accuracy/Final calculation accuracy"]
check(fa["engine"] == "harness" and fa["decision"] == "pass", "harness decides pass")
dc = r["harness_verdicts"]["Accuracy/Deliverable completeness"]
check(dc["engine"] == "llm" and "6/6 answers present" in dc["fallback_reason"], "completeness falls back to LLM when answers exist")

# ---------------------------------------------------------------------------
# 3. Wrong, blank, hardcoded, reordered rows, decoy sheet
# ---------------------------------------------------------------------------
A2 = TMP / "a_mixed.xlsx"
write_attempt_standard(A2, [1890487.51, 1000.0, 0.4213, None, "No", "n/a"], hardcode=True,
                       shuffle=True, extra_decoy=True)
r = AC.run_answer_check(A2, GOLD)
byq = {q["label"]: q for q in r["questions"]}
q0 = byq[QUESTIONS[0][0]]
check(q0["verdict"] == "mismatch" and "sign_flip_not_outflow" in q0["flags"], "flipped difference is wrong + flagged")
check(byq[QUESTIONS[3][0]]["verdict"] == "missing", "blank answer is missing")
check(byq[QUESTIONS[4][0]]["verdict"] == "mismatch", "No vs Yes mismatch")
check(byq[QUESTIONS[1][0]]["verdict"] == "match" and byq[QUESTIONS[1][0]]["hardcoded"] is False,
      "hardcoded requires a golden formula: constant golden -> not flagged")
check(all(q["attempt_cell"].startswith("Questions!") for q in r["questions"] if q["attempt_cell"]),
      "decoy 'Answer Map' loses to the 'Questions' sheet")
check(r["n_match"] == 3, "reordered rows paired by text: 3 matches")
fa = r["harness_verdicts"]["Accuracy/Final calculation accuracy"]
check(fa["decision"] == "fail" and len(fa["mistakes"]) == 3, "harness fails with one mistake per wrong/missing answer")
check(all("location" in m and "description" in m and "severity" in m for m in fa["mistakes"]), "mistakes carry location/description/severity")

# ---------------------------------------------------------------------------
# 4. Hardcoded detection when the golden computes with formulas
# ---------------------------------------------------------------------------
GOLD_F = TMP / "golden_formulas.xlsx"
wbg = openpyxl.load_workbook(GOLD)
wsg = wbg["Questions"]
for i in range(2, 2 + len(QUESTIONS)):
    v = wsg.cell(row=i, column=2).value
    if not isinstance(v, str):
        wbg["Model"].cell(row=i, column=1, value=v)
        wsg.cell(row=i, column=2, value=f"=Model!A{i}")
wbg.save(GOLD_F)
_inject_cached_values(GOLD_F)
A3 = TMP / "a_hardcoded.xlsx"
write_attempt_standard(A3, CORRECT, hardcode=True)
r = AC.run_answer_check(A3, GOLD_F)
check(r["status"] == "ok" and r["n_match"] == 6, "hardcoded attempt: values all match")
check(r["n_hardcoded"] == 4, f"4 numeric constants flagged hardcoded (got {r['n_hardcoded']}); text answers exempt")
fa = r["harness_verdicts"]["Accuracy/Final calculation accuracy"]
check(fa["decision"] == "fail" and sum(1 for m in fa["mistakes"] if "Hardcoded" in m["description"]) == 4,
      "hardcoded answers fail Final calculation accuracy when hardcoded_counts=True")
r2 = AC.run_answer_check(A3, GOLD_F, hardcoded_counts=False)
check(r2["harness_verdicts"]["Accuracy/Final calculation accuracy"]["decision"] == "pass",
      "hardcoded_counts=False: same attempt passes")

# ---------------------------------------------------------------------------
# 5. 297-style 'Answers' sheet: question in column C, Answer column H, header row 7
# ---------------------------------------------------------------------------
A4 = TMP / "a_297.xlsx"
write_attempt_297_style(A4)
r = AC.run_answer_check(A4, GOLD)
check(r["status"] == "ok", f"297-style layout parsed ({r.get('reason')})")
sheets = {s["name"]: s for s in r["attempt_sheets"]}
check(sheets.get("Answers", {}).get("answer_col") == "H", f"answer column is H, not 'Model value' G (got {sheets.get('Answers')})")
check(r["n_match"] == 6, f"297-style: 6/6 match (got {r['n_match']}; {[q['detail'] for q in r['questions'] if q['verdict'] != 'match']})")

# ---------------------------------------------------------------------------
# 6. Zero answers -> completeness FAIL decided by harness; whole-sheet missing -> answers_not_found
# ---------------------------------------------------------------------------
A5 = TMP / "a_empty.xlsx"
write_attempt_standard(A5, [None] * 6)
r = AC.run_answer_check(A5, GOLD)
dc = r["harness_verdicts"]["Accuracy/Deliverable completeness"]
check(dc["engine"] == "harness" and dc["decision"] == "fail", "0 answered -> harness fails completeness")
check(r["harness_verdicts"]["Accuracy/Final calculation accuracy"]["decision"] == "fail", "0 answered -> accuracy fail")

A6 = TMP / "a_nosheet.xlsx"
wb = openpyxl.Workbook(); wb.active["A1"] = "just a model"; wb.save(A6)
r = AC.run_answer_check(A6, GOLD)
check(r["status"] == "answers_not_found", "no question texts anywhere -> answers_not_found")
hv = r["harness_verdicts"]
check(all(v["engine"] == "llm" and v["fallback_reason"] for v in hv.values()), "fails closed to the LLM with a reason")

# ---------------------------------------------------------------------------
# 7. Split golden (task-54 shape) — both sheets read, Unit column absent on one
# ---------------------------------------------------------------------------
G54 = TMP / "golden_split.xlsx"
write_split_golden(G54)
r = AC.run_answer_check(G54, G54)
check(r["status"] == "ok" and r["n_questions"] == 6 and [s["name"] for s in r["golden_sheets"]] == ["Questions Task 1", "Questions Task 2"],
      "split golden: both Questions* sheets read")
check(r["golden_sheets"][1]["unit_col"] is None, "'Number for reference' is not treated as a Unit column")
check(r["n_match"] == 6, "split golden vs itself matches")

# ---------------------------------------------------------------------------
# 8. summary block shape + artifact
# ---------------------------------------------------------------------------
art = TMP / "answer_check.json"
r = AC.run_answer_check(A1, GOLD, output_json_path=art)
sb = AC.summary_block(r)
check(set(sb["harness"]) == {"Accuracy/Final calculation accuracy", "Accuracy/Deliverable completeness"}, "summary carries both harness checks")
check(art.exists() and json.loads(art.read_text())["status"] == "ok", "artifact written")
check("rules_version" in sb and sb["fraction_correct"] == 1.0, "summary carries rules_version + fraction_correct")

print()
if FAILS:
    print(f"{len(FAILS)} FAILURE(S)")
    sys.exit(1)
print("ALL ANSWER-CHECK CHECKS PASSED")
