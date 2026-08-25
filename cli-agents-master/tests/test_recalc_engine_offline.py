#!/usr/bin/env python3
"""
Offline tests for the LibreOffice recalc engine (soffice --convert-to) and the
server's fail-loud startup contract. No LibreOffice, DB, S3, or API keys
needed: a fake `soffice` stub emulates conversion (and its failure modes), so
these run anywhere — including CI. Real-LibreOffice behavior is covered by
test_libreoffice_calc.py.

Run with: pytest -o addopts='' tests/test_recalc_engine_offline.py
"""
import os
import stat
import sys
import tempfile
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))
from excel_mcp_server import libreoffice_calc as lc
from excel_mcp_server.libreoffice_calc import LibreOfficeCalcEngine, resolve_soffice
from excel_cli_agent.mcp_client import ExcelMCPClient

SERVER_PATH = str(Path(__file__).parent.parent / "excel_mcp_server" / "server.py")

# The stub emulates `soffice --convert-to`: mode "ok" rewrites formula cells
# with the value 3 (so data_only reads see a "computed" value — matching the
# engine's warmup workbook, whose =SUM(A1:A2) is also 3), "no-output" exits 0
# without writing (a real soffice failure mode), "fail" exits 1 with stderr,
# "hang" never returns.
_STUB = """#!{python}
import os, sys, time
from pathlib import Path

args = sys.argv[1:]
if "--version" in args:
    print("FakeOffice 1.0.0")
    sys.exit(0)
mode = os.environ.get("FAKE_SOFFICE_MODE", "ok")
if mode == "hang":
    time.sleep(600)
if mode == "fail":
    sys.stderr.write("fake soffice: conversion failed\\n")
    sys.exit(1)
outdir = Path(args[args.index("--outdir") + 1])
src = Path(args[-1])
if mode == "no-output":
    sys.exit(0)
import openpyxl
wb = openpyxl.load_workbook(src)
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                c.value = 3
wb.save(outdir / (src.stem + ".xlsx"))
"""


@pytest.fixture
def fake_soffice(tmp_path):
    stub = tmp_path / "fake_soffice"
    stub.write_text(_STUB.format(python=sys.executable))
    stub.chmod(stub.stat().st_mode | stat.S_IXUSR | stat.S_IXGRP | stat.S_IXOTH)
    return stub


def _make_engine(stub_path, tmp_path, timeout=30.0):
    """Engine with internals set directly — start() has its own tests."""
    engine = LibreOfficeCalcEngine(recalc_timeout=timeout)
    engine.soffice_path = str(stub_path)
    profile = tmp_path / "profile"
    profile.mkdir(exist_ok=True)
    engine._profile_dir = str(profile)
    engine._started = True
    return engine


def _make_workbook(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["A2"] = 2
    ws["A3"] = "=SUM(A1:A2)"
    wb.save(path)
    wb.close()


# ---------------------------------------------------------------------------
# resolve_soffice ladder
# ---------------------------------------------------------------------------

def test_resolve_env_wins(monkeypatch):
    monkeypatch.setenv("LIBREOFFICE_PATH", "/env/soffice")
    monkeypatch.setattr(lc, "_repo_libreoffice_path", lambda: "/cfg/soffice")
    assert resolve_soffice() == ("/env/soffice", "LIBREOFFICE_PATH env var")


def test_resolve_repo_config_beats_path(monkeypatch):
    monkeypatch.delenv("LIBREOFFICE_PATH", raising=False)
    monkeypatch.setattr(lc, "_repo_libreoffice_path", lambda: "/cfg/soffice")
    monkeypatch.setattr(lc.shutil, "which", lambda _: "/usr/bin/soffice")
    path, source = resolve_soffice()
    assert path == "/cfg/soffice"
    assert "config.yaml" in source


def test_resolve_path_lookup(monkeypatch):
    monkeypatch.delenv("LIBREOFFICE_PATH", raising=False)
    monkeypatch.setattr(lc, "_repo_libreoffice_path", lambda: None)
    monkeypatch.setattr(lc.shutil, "which", lambda _: "/usr/bin/soffice")
    assert resolve_soffice() == ("/usr/bin/soffice", "soffice on PATH")


def test_resolve_nothing(monkeypatch):
    monkeypatch.delenv("LIBREOFFICE_PATH", raising=False)
    monkeypatch.setattr(lc, "_repo_libreoffice_path", lambda: None)
    monkeypatch.setattr(lc.shutil, "which", lambda _: None)
    monkeypatch.setattr(lc, "_MACOS_SOFFICE", str(Path("/nonexistent/soffice")))
    assert resolve_soffice() is None


# ---------------------------------------------------------------------------
# recalculate() mechanics against the stub
# ---------------------------------------------------------------------------

def test_recalculate_replaces_file_in_place(fake_soffice, tmp_path, monkeypatch):
    monkeypatch.setenv("FAKE_SOFFICE_MODE", "ok")
    engine = _make_engine(fake_soffice, tmp_path)
    book = tmp_path / "solution.xlsx"
    _make_workbook(book)

    result = engine.recalculate(str(book))

    assert result["success"], result
    assert result["error"] is None
    assert result["duration_ms"] > 0
    wb = openpyxl.load_workbook(book, data_only=True)
    assert wb.active["A3"].value == 3  # "computed" by the stub
    wb.close()


def test_recalculate_not_started(fake_soffice, tmp_path):
    engine = LibreOfficeCalcEngine()
    result = engine.recalculate(str(tmp_path / "x.xlsx"))
    assert not result["success"]
    assert "not started" in result["error"]


def test_exit_zero_without_output_is_failure(fake_soffice, tmp_path, monkeypatch):
    monkeypatch.setenv("FAKE_SOFFICE_MODE", "no-output")
    engine = _make_engine(fake_soffice, tmp_path)
    book = tmp_path / "solution.xlsx"
    _make_workbook(book)

    result = engine.recalculate(str(book))

    assert not result["success"]
    assert "without producing output" in result["error"]
    # The original file must be untouched on failure.
    wb = openpyxl.load_workbook(book)
    assert wb.active["A3"].value == "=SUM(A1:A2)"
    wb.close()


def test_nonzero_exit_surfaces_stderr(fake_soffice, tmp_path, monkeypatch):
    monkeypatch.setenv("FAKE_SOFFICE_MODE", "fail")
    engine = _make_engine(fake_soffice, tmp_path)
    book = tmp_path / "solution.xlsx"
    _make_workbook(book)

    result = engine.recalculate(str(book))

    assert not result["success"]
    assert "conversion failed" in result["error"]


def test_timeout_kills_and_reports(fake_soffice, tmp_path, monkeypatch):
    monkeypatch.setenv("FAKE_SOFFICE_MODE", "hang")
    engine = _make_engine(fake_soffice, tmp_path, timeout=1.0)
    book = tmp_path / "solution.xlsx"
    _make_workbook(book)

    result = engine.recalculate(str(book))

    assert not result["success"]
    assert "timed out" in result["error"]


def test_outdirs_cleaned_up(fake_soffice, tmp_path, monkeypatch):
    """The engine-owned conversion outdir is removed on success AND failure."""
    created = []
    real_mkdtemp = tempfile.mkdtemp

    def tracking_mkdtemp(*args, **kwargs):
        if kwargs.get("prefix", "").startswith("lo_recalc_"):
            d = real_mkdtemp(dir=str(tmp_path), prefix=kwargs["prefix"])
            created.append(d)
            return d
        return real_mkdtemp(*args, **kwargs)

    monkeypatch.setattr(lc.tempfile, "mkdtemp", tracking_mkdtemp)
    engine = _make_engine(fake_soffice, tmp_path)
    book = tmp_path / "solution.xlsx"
    _make_workbook(book)

    monkeypatch.setenv("FAKE_SOFFICE_MODE", "ok")
    assert engine.recalculate(str(book))["success"]
    monkeypatch.setenv("FAKE_SOFFICE_MODE", "fail")
    assert not engine.recalculate(str(book))["success"]

    assert len(created) == 2
    assert all(not Path(d).exists() for d in created)


# ---------------------------------------------------------------------------
# start() / stop()
# ---------------------------------------------------------------------------

def test_start_raises_without_binary(monkeypatch):
    monkeypatch.setenv("LIBREOFFICE_PATH", "/nonexistent/soffice")
    engine = LibreOfficeCalcEngine()
    with pytest.raises(RuntimeError, match="not executable"):
        engine.start()
    assert not engine.is_running


def test_start_warmup_and_stop(fake_soffice, monkeypatch):
    monkeypatch.setenv("LIBREOFFICE_PATH", str(fake_soffice))
    monkeypatch.setenv("FAKE_SOFFICE_MODE", "ok")
    engine = LibreOfficeCalcEngine()
    engine.start()
    try:
        assert engine.is_running
        info = engine.info()
        assert info["engine"] == "libreoffice"
        assert info["soffice_path"] == str(fake_soffice)
        assert info["soffice_version"] == "FakeOffice 1.0.0"
        profile = engine._profile_dir
        assert profile and Path(profile).is_dir()
    finally:
        engine.stop()
    assert not engine.is_running
    assert not Path(profile).exists()


def test_start_fails_when_warmup_saves_no_values(fake_soffice, monkeypatch):
    """A binary that converts but embeds no cached values is a broken install."""
    monkeypatch.setenv("LIBREOFFICE_PATH", str(fake_soffice))
    monkeypatch.setenv("FAKE_SOFFICE_MODE", "no-output")
    engine = LibreOfficeCalcEngine()
    with pytest.raises(RuntimeError, match="warmup"):
        engine.start()


# ---------------------------------------------------------------------------
# Server startup contract (spawns the real server; still no LibreOffice needed)
# ---------------------------------------------------------------------------

def test_server_strict_startup_fails_without_libreoffice(tmp_path, monkeypatch):
    monkeypatch.setenv("LIBREOFFICE_PATH", "/nonexistent/soffice")
    client = ExcelMCPClient(SERVER_PATH, str(tmp_path))
    client.CONNECT_TIMEOUT = 20
    with pytest.raises(Exception, match="Failed to connect"):
        client.connect()


def test_server_no_libreoffice_flag_runs_with_fallback(tmp_path, monkeypatch):
    monkeypatch.setenv("LIBREOFFICE_PATH", "/nonexistent/soffice")
    client = ExcelMCPClient(SERVER_PATH, str(tmp_path), server_args=["--no-libreoffice"])
    try:
        client.connect()
        assert "get_recalc_engine_info" in client.available_tools
        result = client.call_tool("get_recalc_engine_info", {})
        assert result["success"]
        assert result["result"] == {"engine": "fallback"}
    finally:
        client.disconnect()


def test_server_allow_recalc_fallback_runs(tmp_path, monkeypatch):
    monkeypatch.setenv("LIBREOFFICE_PATH", "/nonexistent/soffice")
    client = ExcelMCPClient(SERVER_PATH, str(tmp_path),
                            server_args=["--allow-recalc-fallback"])
    try:
        client.connect()
        result = client.call_tool("get_recalc_engine_info", {})
        assert result["success"]
        assert result["result"] == {"engine": "fallback"}
    finally:
        client.disconnect()


# ---------------------------------------------------------------------------
# Batch preflight (_verify_recalc_engine)
# ---------------------------------------------------------------------------

def _make_runner(config):
    from excel_cli_agent.batch_runner import BatchRunner
    runner = BatchRunner.__new__(BatchRunner)
    runner.server_path = SERVER_PATH
    runner.config = config
    runner._recalc_engine_info = None
    return runner


def test_preflight_records_engine_info(fake_soffice, monkeypatch):
    monkeypatch.setenv("LIBREOFFICE_PATH", str(fake_soffice))
    monkeypatch.setenv("FAKE_SOFFICE_MODE", "ok")
    runner = _make_runner({})
    runner._verify_recalc_engine()
    assert runner._recalc_engine_info["engine"] == "libreoffice"
    assert runner._recalc_extra_configs() == {
        "recalc_engine": "libreoffice",
        "libreoffice_version": "FakeOffice 1.0.0",
    }


def test_preflight_aborts_batch_without_libreoffice(monkeypatch):
    monkeypatch.setenv("LIBREOFFICE_PATH", "/nonexistent/soffice")
    runner = _make_runner({})
    with pytest.raises(RuntimeError, match="allow_recalc_fallback"):
        runner._verify_recalc_engine()


def test_preflight_allows_fallback_when_configured(monkeypatch):
    monkeypatch.setenv("LIBREOFFICE_PATH", "/nonexistent/soffice")
    runner = _make_runner({"allow_recalc_fallback": True})
    runner._verify_recalc_engine()
    assert runner._recalc_engine_info == {"engine": "fallback"}
    assert runner._recalc_extra_configs() == {"recalc_engine": "fallback"}
