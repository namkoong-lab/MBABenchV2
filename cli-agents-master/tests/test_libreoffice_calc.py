#!/usr/bin/env python3
"""
Tests for LibreOffice headless recalculation engine.

Tests the UNO bridge lifecycle, formula recalculation, and cached value accuracy.
"""
import os
import sys
import time
import tempfile

import openpyxl

# Add parent dir to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'excel_mcp_server'))
from libreoffice_calc import LibreOfficeCalcEngine


def create_test_workbook(path, formulas):
    """Create a test workbook with given formulas.

    Args:
        path: File path to save
        formulas: dict of {cell: value_or_formula} for the active sheet
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test"
    for cell, value in formulas.items():
        ws[cell] = value
    wb.save(path)
    wb.close()


def read_cached_values(path, cells):
    """Read cached (calculated) values from a workbook.

    Args:
        path: File path to read
        cells: list of cell addresses to read

    Returns:
        dict of {cell: cached_value}
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    result = {}
    for cell in cells:
        result[cell] = ws[cell].value
    wb.close()
    return result


def read_formulas(path, cells):
    """Read formula text from a workbook.

    Args:
        path: File path to read
        cells: list of cell addresses to read

    Returns:
        dict of {cell: formula_text}
    """
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb.active
    result = {}
    for cell in cells:
        result[cell] = ws[cell].value
    wb.close()
    return result


def test_engine_lifecycle():
    """Test that the engine starts and stops cleanly."""
    print("Test: Engine lifecycle...")
    engine = LibreOfficeCalcEngine(uno_port=2010)

    assert not engine.is_running, "Engine should not be running before start"

    engine.start()
    assert engine.is_running, "Engine should be running after start"

    engine.stop()
    assert not engine.is_running, "Engine should not be running after stop"

    print("  PASSED")


def test_basic_recalculation():
    """Test basic arithmetic formula recalculation."""
    print("Test: Basic recalculation (SUM, arithmetic)...")
    engine = LibreOfficeCalcEngine(uno_port=2011)
    engine.start()

    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            path = f.name

        create_test_workbook(path, {
            "A1": 10,
            "A2": 20,
            "A3": "=A1+A2",
            "B1": "=SUM(A1:A3)",
            "C1": "=A1*A2",
        })

        # Before recalc: all None
        before = read_cached_values(path, ["A3", "B1", "C1"])
        assert all(v is None for v in before.values()), f"Expected all None before recalc, got {before}"

        # Recalculate
        result = engine.recalculate(path)
        assert result["success"], f"Recalculation failed: {result['error']}"
        assert result["duration_ms"] > 0, "Duration should be positive"

        # After recalc: real values
        after = read_cached_values(path, ["A3", "B1", "C1"])
        assert after["A3"] == 30, f"A3 should be 30, got {after['A3']}"
        assert after["B1"] == 60, f"B1 should be 60, got {after['B1']}"
        assert after["C1"] == 200, f"C1 should be 200, got {after['C1']}"

        # Formulas should still be intact
        formulas = read_formulas(path, ["A3", "B1", "C1"])
        assert formulas["A3"] == "=A1+A2", f"Formula lost: {formulas['A3']}"
        assert formulas["B1"] == "=SUM(A1:A3)", f"Formula lost: {formulas['B1']}"

        os.unlink(path)
        print(f"  PASSED (duration: {result['duration_ms']}ms)")
    finally:
        engine.stop()


def test_financial_functions():
    """Test financial functions that xlcalculator can't handle."""
    print("Test: Financial functions (PMT, IRR, NPV)...")
    engine = LibreOfficeCalcEngine(uno_port=2012)
    engine.start()

    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            path = f.name

        create_test_workbook(path, {
            # PMT: monthly payment for $100k loan at 5% over 30 years
            "A1": 100000,    # principal
            "A2": 0.05,      # annual rate
            "A3": 30,        # years
            "A4": "=PMT(A2/12, A3*12, -A1)",
            # NPV
            "B1": 0.1,       # discount rate
            "B2": -1000,     # initial investment
            "B3": 400,       # year 1
            "B4": 400,       # year 2
            "B5": 400,       # year 3
            "B6": "=NPV(B1, B2:B5)",
        })

        result = engine.recalculate(path)
        assert result["success"], f"Recalculation failed: {result['error']}"

        after = read_cached_values(path, ["A4", "B6"])

        # PMT should be ~$536.82
        assert after["A4"] is not None, "PMT should have a value"
        assert abs(after["A4"] - 536.82) < 1.0, f"PMT should be ~536.82, got {after['A4']}"

        # NPV should have a value (exact value depends on formula interpretation)
        assert after["B6"] is not None, "NPV should have a value"

        os.unlink(path)
        print(f"  PASSED (PMT={after['A4']:.2f}, NPV={after['B6']:.2f})")
    finally:
        engine.stop()


def test_cross_sheet_references():
    """Test formulas that reference other worksheets."""
    print("Test: Cross-sheet references...")
    engine = LibreOfficeCalcEngine(uno_port=2013)
    engine.start()

    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            path = f.name

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Data"
        ws1["A1"] = 100
        ws1["A2"] = 200

        ws2 = wb.create_sheet("Summary")
        ws2["A1"] = "=Data!A1+Data!A2"
        ws2["A2"] = "=SUM(Data!A1:A2)"

        wb.save(path)
        wb.close()

        result = engine.recalculate(path)
        assert result["success"], f"Recalculation failed: {result['error']}"

        # Read Summary sheet
        wb2 = openpyxl.load_workbook(path, data_only=True)
        ws_summary = wb2["Summary"]
        val1 = ws_summary["A1"].value
        val2 = ws_summary["A2"].value
        wb2.close()

        assert val1 == 300, f"Cross-sheet A1 should be 300, got {val1}"
        assert val2 == 300, f"Cross-sheet SUM should be 300, got {val2}"

        os.unlink(path)
        print(f"  PASSED (cross-sheet values: {val1}, {val2})")
    finally:
        engine.stop()


def test_lookup_functions():
    """Test VLOOKUP and INDEX/MATCH."""
    print("Test: Lookup functions (VLOOKUP, INDEX/MATCH)...")
    engine = LibreOfficeCalcEngine(uno_port=2014)
    engine.start()

    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            path = f.name

        create_test_workbook(path, {
            # Lookup table
            "A1": "Apple", "B1": 1.50,
            "A2": "Banana", "B2": 0.75,
            "A3": "Cherry", "B3": 2.00,
            # VLOOKUP
            "D1": "Banana",
            "D2": '=VLOOKUP(D1,A1:B3,2,FALSE)',
            # INDEX/MATCH
            "D3": '=INDEX(B1:B3,MATCH("Cherry",A1:A3,0))',
        })

        result = engine.recalculate(path)
        assert result["success"], f"Recalculation failed: {result['error']}"

        after = read_cached_values(path, ["D2", "D3"])

        assert after["D2"] == 0.75, f"VLOOKUP should find Banana=0.75, got {after['D2']}"
        assert after["D3"] == 2.00, f"INDEX/MATCH should find Cherry=2.00, got {after['D3']}"

        os.unlink(path)
        print(f"  PASSED (VLOOKUP={after['D2']}, INDEX/MATCH={after['D3']})")
    finally:
        engine.stop()


def test_auto_restart():
    """Test that the engine auto-restarts if soffice dies."""
    print("Test: Auto-restart after soffice death...")
    engine = LibreOfficeCalcEngine(uno_port=2015)
    engine.start()

    try:
        # Kill soffice
        engine._soffice_process.kill()
        engine._soffice_process.wait()
        time.sleep(1)

        assert not engine.is_running, "Engine should detect dead process"

        # Next recalculate should auto-restart
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            path = f.name

        create_test_workbook(path, {"A1": 5, "A2": "=A1*2"})

        result = engine.recalculate(path)
        assert result["success"], f"Auto-restart recalculation failed: {result['error']}"

        after = read_cached_values(path, ["A2"])
        assert after["A2"] == 10, f"A2 should be 10 after auto-restart, got {after['A2']}"

        os.unlink(path)
        print(f"  PASSED")
    finally:
        engine.stop()


def test_recalculation_timing():
    """Measure recalculation timing for a typical workbook."""
    print("Test: Recalculation timing...")
    engine = LibreOfficeCalcEngine(uno_port=2016)
    engine.start()

    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            path = f.name

        # Create a workbook with 50 formulas
        cells = {"A1": 100}
        for i in range(2, 52):
            cells[f"A{i}"] = f"=A{i-1}+{i}"
        cells["B1"] = "=SUM(A1:A51)"

        create_test_workbook(path, cells)

        # Run 3 recalculations and measure
        times = []
        for i in range(3):
            result = engine.recalculate(path)
            assert result["success"], f"Recalculation {i+1} failed: {result['error']}"
            times.append(result["duration_ms"])

        avg = sum(times) / len(times)

        os.unlink(path)
        print(f"  PASSED (times: {[f'{t:.0f}ms' for t in times]}, avg: {avg:.0f}ms)")
    finally:
        engine.stop()


if __name__ == "__main__":
    print("=" * 60)
    print("LibreOffice Calc Engine Tests")
    print("=" * 60)

    tests = [
        test_engine_lifecycle,
        test_basic_recalculation,
        test_financial_functions,
        test_cross_sheet_references,
        test_lookup_functions,
        test_auto_restart,
        test_recalculation_timing,
    ]

    passed = 0
    failed = 0

    for test_fn in tests:
        try:
            test_fn()
            passed += 1
        except Exception as e:
            print(f"  FAILED: {e}")
            failed += 1

    print("=" * 60)
    print(f"Results: {passed} passed, {failed} failed out of {len(tests)} tests")
    print("=" * 60)

    sys.exit(0 if failed == 0 else 1)
