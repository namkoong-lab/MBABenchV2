#!/usr/bin/env python3
"""
Test to verify circular reference detection is working correctly
"""
import sys
import tempfile
import openpyxl
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from excel_cli_agent.task_executor import ExcelTaskExecutor

def test_circular_ref_detection():
    """Test that circular reference detection works correctly"""
    # Create test file with actual circular reference vs valid cross-sheet reference
    wb = openpyxl.Workbook()

    # Sheet 1 - Q1 with valid cross-sheet reference (should NOT be flagged)
    q1 = wb.active
    q1.title = "Q1"
    q1['B2'] = "=Assumptions!B2*1.05"  # Valid cross-sheet reference to B2 on different sheet

    # Sheet 2 - Assumptions
    assumptions = wb.create_sheet("Assumptions")
    assumptions['B2'] = 1000000

    # Sheet 3 - Q2 with actual circular reference (SHOULD be flagged)
    q2 = wb.create_sheet("Q2")
    q2['B2'] = "=B2*2"  # Actual circular reference - B2 referring to itself locally

    # Sheet 4 - Q3 with explicit circular reference (SHOULD be flagged)
    q3 = wb.create_sheet("Q3")
    q3['B2'] = "=Q3!B2*2"  # Explicit circular reference

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp_path = tmp.name

    try:
        wb.save(tmp_path)
        wb.close()

        # Test detection
        executor = ExcelTaskExecutor(None, "dummy", fresh_context_mode=True)
        output = executor._format_excel_as_grid(tmp_path, is_solution=True)

        print("Grid output:")
        print(output)
        print("\n" + "="*60)

        # Check what was flagged
        if "POTENTIAL CIRCULAR REFS:" in output:
            circular_refs_line = [line for line in output.split('\n') if 'POTENTIAL CIRCULAR REFS:' in line][0]
            print(f"Flagged references: {circular_refs_line}")

            # Q1!B2 should NOT be flagged (cross-sheet reference)
            if "Q1!B2" in circular_refs_line:
                print("❌ FALSE POSITIVE: Q1!B2 flagged incorrectly (=Assumptions!B2*1.05)")
            else:
                print("✅ CORRECT: Q1!B2 not flagged (valid cross-sheet reference)")

            # Q2!B2 SHOULD be flagged (local self-reference)
            if "Q2!B2" in circular_refs_line:
                print("✅ CORRECT: Q2!B2 flagged (=B2*2 is circular)")
            else:
                print("❌ MISSED: Q2!B2 should be flagged (=B2*2 is circular)")

            # Q3!B2 SHOULD be flagged (explicit self-reference)
            if "Q3!B2" in circular_refs_line:
                print("✅ CORRECT: Q3!B2 flagged (=Q3!B2*2 is circular)")
            else:
                print("❌ MISSED: Q3!B2 should be flagged (=Q3!B2*2 is circular)")
        else:
            print("No circular references detected")

    finally:
        Path(tmp_path).unlink()

if __name__ == "__main__":
    test_circular_ref_detection()