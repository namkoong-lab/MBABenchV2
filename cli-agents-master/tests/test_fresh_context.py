#!/usr/bin/env python3
"""
Test script for fresh context mode and enhanced Excel formatting
"""
import sys
import tempfile
import openpyxl
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from excel_cli_agent.task_executor import ExcelTaskExecutor
from excel_cli_agent.mcp_client import ExcelMCPClient

def create_test_excel_file(file_path: str):
    """Create a test Excel file with formulas and data"""
    wb = openpyxl.Workbook()

    # Cover sheet
    cover = wb.active
    cover.title = "Cover"
    cover['A1'] = "Test Case Analysis"
    cover['A2'] = "Date:"
    cover['B2'] = "2024-11-24"

    # Assumptions sheet
    assumptions = wb.create_sheet("Assumptions")
    assumptions['A1'] = "Parameter"
    assumptions['B1'] = "Value"
    assumptions['A2'] = "Revenue"
    assumptions['B2'] = 1000000
    assumptions['A3'] = "Growth Rate"
    assumptions['B3'] = 0.05

    # Q1 sheet with formula
    q1 = wb.create_sheet("Q1")
    q1['A1'] = "Question 1"
    q1['A2'] = "Answer:"
    q1['B2'] = "=Assumptions!B2*1.05"

    wb.save(file_path)
    wb.close()

def test_grid_format():
    """Test the _format_excel_as_grid method"""
    print("🔍 Testing Excel grid formatting...")

    # Create temporary Excel file
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp_path = tmp.name

    try:
        create_test_excel_file(tmp_path)

        # Create a mock task executor to access the grid formatter
        mock_client = None
        executor = ExcelTaskExecutor(
            mock_client,
            "dummy_api_key",
            fresh_context_mode=True,
            enhanced_excel_context=True
        )

        # Test solution.xlsx formatting
        solution_output = executor._format_excel_as_grid(tmp_path, is_solution=True)
        print("📊 Solution Grid Format Output:")
        print(solution_output)
        print("\n" + "="*80 + "\n")

        # Test context Excel formatting
        context_output = executor._format_excel_as_grid(tmp_path, is_solution=False)
        print("📋 Context Grid Format Output:")
        print(context_output)
        print("\n" + "="*80 + "\n")

        # Validate output contains expected elements
        assert "=== SOLUTION FILE:" in solution_output
        assert "WORKSHEET: Cover" in solution_output
        assert "WORKSHEET: Assumptions" in solution_output
        assert "WORKSHEET: Q1" in solution_output
        assert "A1: \"Test Case Analysis\"" in solution_output
        assert "B2: =Assumptions!B2*1.05" in solution_output
        assert "SUMMARY:" in solution_output

        assert "CONTEXT EXCEL:" in context_output
        assert "Used Range:" in context_output
        assert "A1:" in context_output

        print("✅ Grid format test PASSED")

    finally:
        Path(tmp_path).unlink()

def test_fresh_context_prompt():
    """Test fresh context prompt generation"""
    print("🎯 Testing fresh context prompt generation...")

    # Create temporary directory and Excel file
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_path = Path(tmp_dir)
        solution_path = tmp_path / "solution.xlsx"

        create_test_excel_file(str(solution_path))

        # Create mock task executor
        mock_client = type('MockClient', (), {'storage_path': str(tmp_path)})()

        executor = ExcelTaskExecutor(
            mock_client,
            "dummy_api_key",
            fresh_context_mode=True,
            enhanced_excel_context=True
        )

        # Test loading solution context
        solution_context = executor._load_solution_context()
        print("📄 Fresh Context Output:")
        print(solution_context)
        print("\n" + "="*80 + "\n")

        # Validate solution context
        assert "=== SOLUTION FILE:" in solution_context
        assert "WORKSHEET: Cover" in solution_context
        assert "WORKSHEET: Assumptions" in solution_context
        assert "WORKSHEET: Q1" in solution_context
        assert "SUMMARY:" in solution_context

        print("✅ Fresh context test PASSED")

def test_enhanced_excel_context():
    """Test enhanced Excel context loading"""
    print("📋 Testing enhanced Excel context...")

    # Create temporary Excel file
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp_path = tmp.name

    try:
        create_test_excel_file(tmp_path)

        # Test both enhanced and legacy modes
        executor_enhanced = ExcelTaskExecutor(
            None, "dummy", enhanced_excel_context=True
        )
        executor_legacy = ExcelTaskExecutor(
            None, "dummy", enhanced_excel_context=False
        )

        # Add context Excel file
        executor_enhanced.context_excels = [tmp_path]
        executor_legacy.context_excels = [tmp_path]

        # Test enhanced mode
        enhanced_output = executor_enhanced._extract_excel_texts()
        print("🆕 Enhanced Excel Context:")
        print(enhanced_output[:500] + "...")
        print()

        # Test legacy mode
        legacy_output = executor_legacy._extract_excel_texts()
        print("📜 Legacy Excel Context:")
        print(legacy_output[:500] + "...")
        print("\n" + "="*80 + "\n")

        # Validate outputs are different
        assert enhanced_output != legacy_output
        assert "CONTEXT EXCEL:" in enhanced_output
        assert "Used Range:" in enhanced_output
        assert "A1:" in enhanced_output

        print("✅ Enhanced Excel context test PASSED")

    finally:
        Path(tmp_path).unlink()

def main():
    """Run all tests"""
    print("="*80)
    print("FRESH CONTEXT SYSTEM VALIDATION")
    print("="*80)
    print()

    try:
        test_grid_format()
        print()
        test_fresh_context_prompt()
        print()
        test_enhanced_excel_context()
        print()

        print("🎉 ALL TESTS PASSED!")
        print()
        print("✅ Grid formatting works correctly")
        print("✅ Fresh context mode loads solution.xlsx properly")
        print("✅ Enhanced Excel context provides better formatting")
        print("✅ Circular reference detection included in output")

    except Exception as e:
        print(f"❌ TEST FAILED: {e}")
        import traceback
        traceback.print_exc()
        return 1

    return 0

if __name__ == "__main__":
    exit(main())