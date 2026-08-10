"""Offline tests for the context-budget assembler (no API, no MCP server).

Covers: byte-identical fast path, sheet-summary fallback, read-result carry,
overflow detection, context-window resolution, and end-to-end _assemble_context
behavior on both sides of the budget.

Run:  python tests/test_context_budget.py   (or pytest)
"""
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook

from excel_cli_agent import models_config
from excel_cli_agent.task_executor import (
    ExcelTaskExecutor,
    ExecutionStep,
    TaskExecution,
    TaskStatus,
)


def make_executor(tmp_path, **attrs):
    """Bare executor for exercising context methods without API/MCP setup."""
    ex = object.__new__(ExcelTaskExecutor)
    ex.context_pdfs = []
    ex.context_excels = []
    ex.fresh_context_mode = True
    ex.enhanced_excel_context = True
    ex.recent_history_count = 3
    ex.max_completion_tokens = 8000
    ex.model = "test-model"
    ex.context_window = 100_000
    ex._chars_per_token = 4.0
    ex._overflow_shrink = 1.0

    class FakeClient:
        storage_path = str(tmp_path)

    ex.excel_client = FakeClient()
    for k, v in attrs.items():
        setattr(ex, k, v)
    return ex


def make_task(steps=None):
    return TaskExecution(
        task_id="t1",
        user_prompt="Build the model.",
        status=TaskStatus.IN_PROGRESS,
        steps=steps or [],
        start_time=time.time(),
        total_iterations=1,
        max_iterations=30,
    )


def write_workbook(path, big_rows=0):
    wb = Workbook()
    ws = wb.active
    ws.title = "Model"
    ws["A1"] = "Revenue"
    ws["B1"] = 100
    ws["B2"] = "=B1*2"
    if big_rows:
        data = wb.create_sheet("Data")
        for r in range(1, big_rows + 1):
            data.cell(row=r, column=1, value=r)
            data.cell(row=r, column=2, value=r * 1.5)
            data.cell(row=r, column=3, value=f"label_{r}")
    wb.save(path)
    wb.close()


def test_fast_path_byte_identical(tmp_path):
    """max_chars=None and a huge budget must render identically."""
    xlsx = tmp_path / "solution.xlsx"
    write_workbook(xlsx, big_rows=50)
    ex = make_executor(tmp_path)
    full = ex._format_excel_as_grid(str(xlsx), is_solution=True)
    budgeted = ex._format_excel_as_grid(str(xlsx), is_solution=True, max_chars=10**9)
    assert full == budgeted
    # Known formatting invariants of the historical renderer
    assert "=== WORKSHEET: Model ===" in full
    assert "Used Range: A1:B2" in full
    assert 'A1: "Revenue" [text]' in full
    assert "B2: =B1*2" in full
    assert "SUMMARY:" in full and "formulas total" in full
    assert "SHEET SUMMARIZED" not in full


def test_summary_fallback(tmp_path):
    """Over budget, the largest sheet is summarized; small sheets stay full."""
    xlsx = tmp_path / "solution.xlsx"
    write_workbook(xlsx, big_rows=800)
    ex = make_executor(tmp_path)
    full = ex._format_excel_as_grid(str(xlsx), is_solution=True)
    budget = len(full) // 4
    reduced = ex._format_excel_as_grid(str(xlsx), is_solution=True, max_chars=budget)
    assert len(reduced) <= budget
    assert "=== WORKSHEET: Data ===" in reduced
    assert "SHEET SUMMARIZED" in reduced
    assert "'Data'!A2:C800" in reduced        # formula-addressable pointer
    assert "get_cell_range" in reduced        # read-on-demand pointer
    assert "rows hidden" in reduced
    # Small sheet untouched, tail summary intact
    assert 'A1: "Revenue" [text]' in reduced
    assert "SUMMARY:" in reduced


def test_read_result_carry(tmp_path):
    """Reduced mode carries READ results in full; normal mode truncates at 150."""
    # Non-periodic so prefix-substring counts are unambiguous (~3600 chars)
    long_result = "".join(f"R{i}C1={i * 7}; " for i in range(400))
    steps = [
        ExecutionStep(
            step_number=1, description="read", tool_name="get_cell_range",
            tool_args={"filename": "s.xlsx"}, result={"result": long_result},
        ),
        ExecutionStep(
            step_number=2, description="write", tool_name="set_cell_formula",
            tool_args={"cell": "A1"}, result={"result": long_result},
        ),
    ]
    ex = make_executor(tmp_path)
    task = make_task(steps)

    normal = ex._get_fresh_context_prompt(task, solution_override="")
    assert long_result[:1000] not in normal  # both truncated to 150

    carried = ex._get_fresh_context_prompt(
        task, solution_override="", carry_read_results=True
    )
    assert long_result[:1000] in carried               # read tool: full carry
    assert carried.count(long_result[:1000]) == 1      # write tool: still 150


def test_overflow_detection(tmp_path):
    ex = make_executor(tmp_path)
    positives = [
        "Error code: 400 - prompt is too long: 250000 tokens > 200000 maximum",
        "This model's maximum context length is 128000 tokens",
        "context_length_exceeded",
        "input length and `max_tokens` exceed context limit",
    ]
    for msg in positives:
        assert ex._is_context_overflow_error(Exception(msg)), msg
    assert not ex._is_context_overflow_error(Exception("rate limit exceeded"))
    assert not ex._is_context_overflow_error(Exception("connection reset"))


def test_context_window_resolution():
    """Static fallback and default when the live fetch is unavailable."""
    old = (models_config._live_pricing, models_config._live_pricing_attempted)
    try:
        models_config._live_pricing = None
        models_config._live_pricing_attempted = True
        assert models_config.resolve_context_window("claude-fable-5") == 200_000
        assert (
            models_config.resolve_context_window("unknown-model-xyz")
            == models_config.DEFAULT_CONTEXT_WINDOW
        )
    finally:
        models_config._live_pricing, models_config._live_pricing_attempted = old


def test_assemble_fast_path_unchanged(tmp_path):
    """Under budget the assembled message matches the historical assembly."""
    write_workbook(tmp_path / "solution.xlsx", big_rows=20)
    ex = make_executor(tmp_path)
    task = make_task()
    assembled = ex._assemble_context(task, system_prompt="SYS")
    manual = ex._get_context_prompt(task)  # historical path: no extras
    assert assembled == manual
    assert task.context_reduced is False


def test_assemble_reduced_path(tmp_path):
    """Over budget: summaries engage, flag set, output fits the budget."""
    write_workbook(tmp_path / "solution.xlsx", big_rows=3000)
    ex = make_executor(tmp_path, context_window=12_000)  # tiny window
    task = make_task()
    assembled = ex._assemble_context(task, system_prompt="SYS")
    assert task.context_reduced is True
    assert "SHEET SUMMARIZED" in assembled
    budget = ex._context_budget_chars(len("SYS"))
    # Head + budgeted solution grid must respect the overall budget with
    # modest slack (head is measured, not clamped).
    assert len(assembled) <= budget * 1.1


def main() -> int:
    import inspect
    import tempfile

    failures = 0
    for name, fn in sorted(globals().items()):
        if not name.startswith("test_"):
            continue
        with tempfile.TemporaryDirectory() as td:
            kwargs = (
                {"tmp_path": Path(td)}
                if "tmp_path" in inspect.signature(fn).parameters
                else {}
            )
            try:
                fn(**kwargs)
                print(f"OK  {name}")
            except AssertionError as e:
                failures += 1
                print(f"FAIL {name}: {e}")
    print("ALL CONTEXT BUDGET TESTS PASSED" if not failures else f"{failures} FAILURES")
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())
