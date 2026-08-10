"""
Iterative Task Execution Engine
Provides iterative reasoning and Excel tool execution capabilities
"""
import json
import re
import signal
import time
import os
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple
from dataclasses import dataclass
from enum import Enum
from pathlib import Path

# Use sync OpenAI client to avoid asyncio socket issues with CLOSE-WAIT
from openai import OpenAI
import httpx

# Try to import Langfuse for observability (optional)
try:
    from langfuse import Langfuse
    LANGFUSE_ENABLED = True
except ImportError:
    LANGFUSE_ENABLED = False
    Langfuse = None

# Try to import Anthropic for direct Claude API access (optional)
try:
    import anthropic
    ANTHROPIC_ENABLED = True
except ImportError:
    ANTHROPIC_ENABLED = False
    anthropic = None

from .mcp_client import ExcelMCPClient
from .models_config import MODEL_PRICING, calculate_cost


class TaskStatus(Enum):
    PENDING = "pending"
    IN_PROGRESS = "in_progress"
    COMPLETED = "completed"
    FAILED = "failed"
    NEEDS_CLARIFICATION = "needs_clarification"


@dataclass
class ExecutionStep:
    step_number: int
    description: str
    tool_name: str
    tool_args: Dict[str, Any]
    result: Optional[Dict[str, Any]] = None
    execution_time: Optional[float] = None
    error: Optional[str] = None
    # Detailed reasoning artifacts
    reasoning_json: Optional[Dict[str, Any]] = None
    raw_model_response: Optional[str] = None


@dataclass
class TaskExecution:
    task_id: str
    user_prompt: str
    status: TaskStatus
    steps: List[ExecutionStep]
    start_time: float
    end_time: Optional[float] = None
    total_iterations: int = 0
    max_iterations: int = 20
    final_result: Optional[str] = None
    error: Optional[str] = None
    total_cost_usd: float = 0.0


class StreamTimeoutError(Exception):
    """Raised when streaming API call exceeds hard timeout via signal.alarm."""
    pass


class ExcelTaskExecutor:
    def __init__(self, excel_client: ExcelMCPClient, api_key: str, model: str = "gpt-5-nano-2025-08-07", custom_reasoning: bool = False, fresh_context_mode: bool = False, enhanced_excel_context: bool = True, recent_history_count: int = 5, max_completion_tokens: int = 65535, reasoning_effort: str = None, api_timeout_seconds: int = None, use_anthropic_direct: bool = False, anthropic_api_key: str = None, thinking_budget_tokens: int = None, use_openai_direct: bool = False, system_prompt_path: str = None, base_url: str = None):
        self.excel_client = excel_client

        # --- Unified base_url routing ---
        # Resolve base_url: explicit arg > env var > legacy OpenRouter sniffing > default
        self.base_url = base_url or os.getenv("BASE_URL")

        # Auto-detect provider from base_url
        self.use_anthropic_direct = False
        self.use_openai_direct = False  # True when NOT using OpenRouter (affects reasoning_effort format)
        self.anthropic_client = None
        self.thinking_budget_tokens = thinking_budget_tokens

        if self.base_url and "anthropic" in self.base_url.lower():
            # Anthropic-compatible endpoint
            self.use_anthropic_direct = True
            # The api_key argument may be a generic key resolved upstream (cli.py
            # prefers OPENAI_API_KEY when several are set in .env). Only use it
            # here if it is actually an Anthropic key; otherwise fall back to
            # ANTHROPIC_API_KEY, then to the argument as a last resort.
            _looks_anthropic = (api_key or "").startswith("sk-ant-")
            self.api_key = (
                (api_key if _looks_anthropic else None)
                or os.getenv("ANTHROPIC_API_KEY")
                or api_key
            )
            if not self.api_key:
                raise ValueError("API key required. Set ANTHROPIC_API_KEY in .env or pass --api-key")
            if not ANTHROPIC_ENABLED:
                raise ImportError("Anthropic SDK not installed. Run: pip install anthropic")
            self.anthropic_client = anthropic.Anthropic(
                api_key=self.api_key,
                base_url=self.base_url,
                # The SDK default read timeout is 600s. A max-effort model
                # (claude-fable-5) can spend longer than that on a single heavy
                # turn, and the default fires "The read operation timed out"
                # mid-generation, abandoning the whole task ($0, wasted). Give
                # the streaming read a generous window (matches the max-effort
                # per-call budget); connect stays short.
                timeout=httpx.Timeout(3600.0, connect=30.0),
            )
        elif self.base_url:
            # OpenAI-compatible endpoint (vLLM, SGLang, OpenRouter, etc.)
            # For OpenRouter URLs, prefer the OpenRouter key
            if "openrouter" in self.base_url.lower():
                self.api_key = os.getenv("OPENROUTER_API_KEY") or api_key or os.getenv("OPENAI_API_KEY") or "no-key"
            else:
                self.api_key = api_key or os.getenv("OPENAI_API_KEY") or os.getenv("OPENROUTER_API_KEY") or "no-key"
                self.use_openai_direct = True
        else:
            # No base_url: legacy behavior — sniff OpenRouter key, fall back to OpenAI
            openrouter_key = os.getenv("OPENROUTER_API_KEY")
            if use_openai_direct or not openrouter_key:
                # OpenAI direct
                self.api_key = api_key or os.getenv("OPENAI_API_KEY")
                self.use_openai_direct = True
            else:
                # OpenRouter
                self.api_key = openrouter_key
                self.base_url = os.getenv("OPENROUTER_BASE_URL", "https://openrouter.ai/api/v1")

            # Legacy anthropic direct flag (for backward compat with existing configs)
            if use_anthropic_direct and ANTHROPIC_ENABLED:
                self.use_anthropic_direct = True
                self.api_key = anthropic_api_key or os.getenv("ANTHROPIC_API_KEY") or api_key
                if not self.api_key:
                    raise ValueError("API key required for Anthropic direct access")
                self.anthropic_client = anthropic.Anthropic(
                    api_key=self.api_key,
                    timeout=httpx.Timeout(3600.0, connect=30.0),  # see note above
                )
            elif not self.api_key:
                self.api_key = api_key

        # Determine timeout based on reasoning effort or explicit config
        # High reasoning efforts need more time (xhigh: 300s, high: 240s, etc.)
        if api_timeout_seconds:
            timeout_secs = api_timeout_seconds
        elif reasoning_effort == "max":
            timeout_secs = 3600  # Anthropic effort-based models (e.g. claude-fable-5)
        elif reasoning_effort == "xhigh":
            timeout_secs = 300  # 5 minutes for xhigh reasoning
        elif reasoning_effort == "high":
            timeout_secs = 240  # 4 minutes for high reasoning
        else:
            timeout_secs = 180  # 3 minutes default

        # Use httpx.Timeout for granular control over read/connect timeouts.
        # read must cover the longest SILENT gap in a stream: high-reasoning
        # models (e.g. gpt-5.6-sol at xhigh) can think for minutes between
        # streamed chunks, and the old fixed read=60.0 killed healthy calls
        # mid-reasoning ("Request timed out", observed 2026-07-23). Tie it to
        # the overall timeout instead; the SIGALRM hard timeout still bounds
        # the whole call.
        self.api_timeout = httpx.Timeout(
            float(timeout_secs), read=float(timeout_secs), write=60.0, connect=30.0
        )

        # Hard timeout for signal.alarm (must be less than cloud NAT timeout ~600s)
        self.hard_timeout_seconds = timeout_secs

        # Create initial client
        self._create_openai_client()

        self.model = model
        self.custom_reasoning = custom_reasoning
        self.current_execution: Optional[TaskExecution] = None
        self.execution_history: List[TaskExecution] = []

        # New context management flags
        self.fresh_context_mode = fresh_context_mode  # Load solution.xlsx each iteration (no history)
        self.enhanced_excel_context = enhanced_excel_context  # Use grid format for Excel files
        self.recent_history_count = recent_history_count  # Number of recent tool calls to include in fresh context
        self.max_completion_tokens = max_completion_tokens  # Max tokens for model output (higher for thinking models)
        self.reasoning_effort = reasoning_effort  # For OpenRouter: None, "xhigh", "high", "medium", "low", "minimal", "none"

        # Check Langfuse configuration
        self.langfuse_enabled = LANGFUSE_ENABLED and all([
            os.getenv("LANGFUSE_SECRET_KEY"),
            os.getenv("LANGFUSE_PUBLIC_KEY"),
            os.getenv("LANGFUSE_HOST")
        ])

        # Initialize Langfuse client for explicit trace management
        self.langfuse_client = None
        self._current_trace = None
        self._current_generation = None
        if self.langfuse_enabled:
            self.langfuse_client = Langfuse()

        # Prepare logs directory under the storage path
        try:
            base = Path(self.excel_client.storage_path)
        except Exception:
            base = Path(".")
        self.logs_dir = base / "agent_logs"
        self.logs_dir.mkdir(parents=True, exist_ok=True)
        # OpenAI request logging (CSV backup)
        self.openai_log_file = self.logs_dir / "openai_requests.csv"
        self._init_openai_log()
        # Context PDFs
        self.context_pdfs: List[str] = []
        # Context Excel files
        self.context_excels: List[str] = []
        # Iteration control
        self.default_max_iterations: int = 30
        # Snapshot mode: save solution.xlsx + AI context text after each iteration
        self.snapshot_iterations: bool = False
        # Verbose mode
        self.verbose: bool = False
        # System prompt path override (for versioned prompts)
        self.system_prompt_path = system_prompt_path

    def _create_openai_client(self):
        """Create OpenAI client with current configuration."""
        kwargs = {"api_key": self.api_key, "timeout": self.api_timeout}
        if self.base_url:
            kwargs["base_url"] = self.base_url
        if self.base_url:
            print(f"🔑 Using API: {self.base_url}")
        self.openai_client = OpenAI(**kwargs)

    def _recreate_openai_client(self):
        """Recreate OpenAI client to clear stale/dead connections."""
        print("🔄 Recreating OpenAI client with fresh connections...")
        self._create_openai_client()

    def set_max_iterations(self, n: int) -> None:
        """Set the default maximum iterations for new tasks."""
        self.default_max_iterations = max(1, int(n))

    def set_verbose(self, enabled: bool) -> None:
        """Set verbose mode to print full AI responses."""
        self.verbose = enabled

    def add_context_pdfs(self, pdf_paths: List[str]) -> Dict[str, Any]:
        """Add PDF files to context for task execution (only from storage directory)"""
        added = []
        not_found = []

        for pdf_path in pdf_paths:
            # Only look in storage path
            file_path = Path(self.excel_client.storage_path) / pdf_path

            if file_path.exists() and file_path.suffix.lower() == '.pdf':
                abs_path = str(file_path.absolute())
                if abs_path not in self.context_pdfs:
                    self.context_pdfs.append(abs_path)
                    added.append(abs_path)
            else:
                not_found.append(pdf_path)

        return {"added": added, "not_found": not_found, "total_context_pdfs": len(self.context_pdfs)}

    def clear_context_pdfs(self):
        """Clear all PDF context"""
        count = len(self.context_pdfs)
        self.context_pdfs = []
        return {"cleared": count}

    def add_context_excels(self, excel_paths: List[str]) -> Dict[str, Any]:
        """Add Excel files to context for task execution (only from storage directory)"""
        added = []
        not_found = []

        for excel_path in excel_paths:
            # Only look in storage path
            file_path = Path(self.excel_client.storage_path) / excel_path

            if file_path.exists() and file_path.suffix.lower() in ['.xlsx', '.xlsm']:
                abs_path = str(file_path.absolute())
                if abs_path not in self.context_excels:
                    self.context_excels.append(abs_path)
                    added.append(abs_path)
            else:
                not_found.append(excel_path)

        return {"added": added, "not_found": not_found, "total_context_excels": len(self.context_excels)}

    def clear_context_excels(self):
        """Clear all Excel context"""
        count = len(self.context_excels)
        self.context_excels = []
        return {"cleared": count}

    def _extract_pdf_texts(self) -> str:
        """Extract raw text from all context PDFs"""
        if not self.context_pdfs:
            return ""

        from pypdf import PdfReader

        all_text = []
        for pdf_path in self.context_pdfs:
            reader = PdfReader(pdf_path)
            pdf_name = Path(pdf_path).name
            text_parts = [f"\n{'='*60}\nPDF: {pdf_name}\n{'='*60}\n"]

            for page_num, page in enumerate(reader.pages, 1):
                page_text = page.extract_text()
                if page_text.strip():
                    text_parts.append(f"\n--- Page {page_num} ---\n{page_text}\n")

            all_text.append("".join(text_parts))

        return "\n".join(all_text)

    def _extract_excel_texts(self) -> str:
        """Extract data from all context Excel files using enhanced grid format"""
        if not self.context_excels:
            return ""

        if self.enhanced_excel_context:
            # Use enhanced grid format for better context understanding
            all_text = []
            for excel_path in self.context_excels:
                grid_text = self._format_excel_as_grid(excel_path, is_solution=False)
                all_text.append(grid_text)
            return "\n".join(all_text)
        else:
            # Fallback to legacy CSV-like format
            return self._extract_excel_texts_legacy()

    def _extract_excel_texts_legacy(self) -> str:
        """Legacy CSV-like Excel text extraction (for backward compatibility)"""
        if not self.context_excels:
            return ""

        from openpyxl import load_workbook

        all_text = []
        for excel_path in self.context_excels:
            try:
                # Load workbook with formulas and calculated values
                wb_formulas = load_workbook(excel_path, data_only=False)
                wb_data = load_workbook(excel_path, data_only=True)
                excel_name = Path(excel_path).name
                text_parts = [f"\n{'='*60}\nEXCEL: {excel_name}\n{'='*60}\n"]

                for sheet in wb_formulas.worksheets:
                    sheet_name = sheet.title
                    text_parts.append(f"\n--- Sheet: {sheet_name} ---\n")

                    # Get the used range
                    if sheet.max_row is None or sheet.max_column is None:
                        text_parts.append("(Empty sheet)\n")
                        continue

                    # Get corresponding data sheet
                    data_sheet = wb_data[sheet_name]

                    # Read cells and format as CSV with formulas
                    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row,
                                               min_col=1, max_col=sheet.max_column):
                        row_values = []
                        for cell in row:
                            # Get cell value (with formula)
                            value = cell.value

                            # Check if it's a formula
                            if isinstance(value, str) and value.startswith('='):
                                # Get calculated value from data workbook
                                calculated_value = data_sheet.cell(row=cell.row, column=cell.column).value
                                if calculated_value is not None:
                                    row_values.append(f"{calculated_value} [{value}]")
                                else:
                                    row_values.append(f"[{value}]")
                            elif value is None:
                                row_values.append("")
                            else:
                                row_values.append(str(value))

                        # Only add non-empty rows
                        if any(v for v in row_values):
                            text_parts.append(",".join(row_values) + "\n")

                all_text.append("".join(text_parts))
                wb_formulas.close()
                wb_data.close()

            except Exception as e:
                all_text.append(f"\n{'='*60}\nEXCEL: {Path(excel_path).name}\n{'='*60}\nError reading file: {str(e)}\n")

        return "\n".join(all_text)

    def _format_excel_as_grid(self, excel_path: str, is_solution: bool = False) -> str:
        """Format Excel file as grid with column letters + row numbers

        Used for both:
        - solution.xlsx (fresh context mode)
        - context Excel files (enhanced context)

        Shows all non-empty cells with full formula text and computed values.

        Args:
            excel_path: Path to Excel file
            is_solution: True if this is solution.xlsx, False for context files

        Returns:
            Formatted grid text with column letters and row numbers
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter

            # Load workbook with formulas and calculated values
            wb_formulas = load_workbook(excel_path, data_only=False)
            wb_data = load_workbook(excel_path, data_only=True)
            excel_name = Path(excel_path).name

            # Header for solution vs context files
            if is_solution:
                text_parts = [f"\n=== SOLUTION FILE: {excel_name} ===\n"]
            else:
                text_parts = [f"\n{'='*60}\nCONTEXT EXCEL: {excel_name}\n{'='*60}\n"]

            total_formulas = 0
            circular_refs_detected = []
            empty_q_sheets = []

            for sheet in wb_formulas.worksheets:
                sheet_name = sheet.title
                data_sheet = wb_data[sheet_name]

                # Get used range
                if sheet.max_row is None or sheet.max_column is None:
                    text_parts.append(f"\n=== WORKSHEET: {sheet_name} ===\n(Empty sheet)\n")
                    # Check if this is a Q sheet that's empty
                    if sheet_name.startswith('Q') and sheet_name[1:].isdigit():
                        empty_q_sheets.append(sheet_name)
                    continue

                # Use actual used range — show all non-empty cells
                actual_max_row = sheet.max_row
                actual_max_col = sheet.max_column

                used_range = f"{get_column_letter(1)}{1}:{get_column_letter(actual_max_col)}{actual_max_row}"
                text_parts.append(f"\n=== WORKSHEET: {sheet_name} ===\n")
                text_parts.append(f"Used Range: {used_range}\n")

                # Create grid display with column letters + row numbers
                for row in range(1, actual_max_row + 1):
                    row_cells = []
                    for col in range(1, actual_max_col + 1):
                        col_letter = get_column_letter(col)
                        cell_addr = f"{col_letter}{row}"

                        # Get formula and calculated value
                        formula_cell = sheet.cell(row=row, column=col)
                        data_cell = data_sheet.cell(row=row, column=col)

                        formula_value = formula_cell.value
                        calculated_value = data_cell.value

                        if formula_value is None:
                            continue  # Skip empty cells entirely

                        cell_type = ""
                        if isinstance(formula_value, str) and formula_value.startswith('='):
                            # It's a formula
                            total_formulas += 1
                            # Check for potential circular reference
                            self_ref_patterns = [
                                f"'{sheet_name}'!{cell_addr}",  # 'Sheet'!B2
                                f"{sheet_name}!{cell_addr}",    # Sheet!B2
                            ]
                            if any(pattern in formula_value for pattern in self_ref_patterns):
                                circular_refs_detected.append(f"{sheet_name}!{cell_addr}")
                            elif "!" not in formula_value:
                                import re
                                cell_refs = re.findall(r'\b[A-Z]{1,3}\d+\b', formula_value)
                                if cell_addr in cell_refs:
                                    circular_refs_detected.append(f"{sheet_name}!{cell_addr}")
                            cell_display = formula_value
                            # Show computed value if available
                            if calculated_value is not None:
                                cell_type = f"[={calculated_value}]"
                            else:
                                cell_type = "[formula]"
                        elif isinstance(formula_value, (int, float)):
                            cell_display = str(formula_value)
                            cell_type = "[number]"
                        else:
                            # String or other value
                            cell_display = f'"{str(formula_value)}"'
                            cell_type = "[text]"

                        row_cells.append(f"{cell_addr}: {cell_display} {cell_type}")

                    # Only show rows that have non-empty cells
                    if row_cells:
                        text_parts.append(f"  {' | '.join(row_cells)}\n")

                # Check for Q sheets without formulas
                if sheet_name.startswith('Q') and sheet_name[1:].isdigit():
                    sheet_formulas = 0
                    for row in sheet.iter_rows():
                        for cell in row:
                            if isinstance(cell.value, str) and cell.value.startswith('='):
                                sheet_formulas += 1
                    if sheet_formulas == 0:
                        empty_q_sheets.append(sheet_name)

            # Add summary for solution files
            if is_solution:
                text_parts.append(f"\nSUMMARY:\n")
                text_parts.append(f"- {len(wb_formulas.sheetnames)} worksheets: {', '.join(wb_formulas.sheetnames)}\n")
                text_parts.append(f"- {total_formulas} formulas total\n")

                if circular_refs_detected:
                    text_parts.append(f"- ⚠️ POTENTIAL CIRCULAR REFS: {', '.join(circular_refs_detected)}\n")
                else:
                    text_parts.append(f"- ✅ No circular references detected\n")

                if empty_q_sheets:
                    text_parts.append(f"- ⚠️ EMPTY Q SHEETS: {', '.join(empty_q_sheets)}\n")
                else:
                    text_parts.append(f"- ✅ All Q sheets have content\n")

            wb_formulas.close()
            wb_data.close()

            return "".join(text_parts)

        except Exception as e:
            excel_name = Path(excel_path).name
            return f"\n{'='*60}\nEXCEL: {excel_name}\n{'='*60}\nError reading file: {str(e)}\n"

    def _load_solution_context(self) -> str:
        """Load solution.xlsx into context with full grid view for fresh context mode"""
        solution_path = Path(self.excel_client.storage_path) / "solution.xlsx"
        if not solution_path.exists():
            return "\n=== SOLUTION FILE ===\nNo solution.xlsx found yet. Create it to begin building your Excel model.\n"

        return self._format_excel_as_grid(str(solution_path), is_solution=True)

    def _save_iteration_snapshot(self, iteration: int, task: TaskExecution = None):
        """Save solution.xlsx copy and the EXACT prompt the AI saw for this iteration."""
        import shutil
        try:
            solution_path = Path(self.excel_client.storage_path) / "solution.xlsx"

            snapshots_dir = self.logs_dir / "snapshots"
            snapshots_dir.mkdir(parents=True, exist_ok=True)

            prefix = f"iter_{iteration:03d}"

            # Copy solution.xlsx (post-action state for comparison)
            if solution_path.exists():
                shutil.copy2(str(solution_path), str(snapshots_dir / f"{prefix}_solution.xlsx"))

            # Use the cached prompt from _call_reasoning_engine - this is exactly
            # what the AI saw BEFORE it made its decision for this iteration
            context_text = getattr(self, '_last_user_message', None)
            if context_text is None:
                # Fallback: rebuild (won't be exact but better than nothing)
                context_text = self._load_solution_context() if not task else self._get_context_prompt(task)

            # Save system prompt on first iteration
            if iteration == 1:
                try:
                    system_prompt = self._get_system_prompt()
                    (snapshots_dir / "system_prompt.md").write_text(system_prompt, encoding="utf-8")
                except Exception:
                    pass

            (snapshots_dir / f"{prefix}_context.md").write_text(context_text, encoding="utf-8")

            print(f"📸 Snapshot saved: {prefix}")
        except Exception as e:
            print(f"⚠️ Snapshot failed: {e}")

    def _init_openai_log(self):
        """Initialize CSV log file for OpenAI requests"""
        import csv
        if not self.openai_log_file.exists():
            with open(self.openai_log_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow([
                    'timestamp', 'task_id', 'iteration', 'model',
                    'messages', 'tools', 'tool_choice',
                    'response_content', 'tool_calls', 'finish_reason',
                    'prompt_tokens', 'completion_tokens', 'total_tokens', 'cost_usd'
                ])

    def _log_openai_request(self, request_data: dict, response_data: dict, task_id: str = "", iteration: int = 0):
        """Log OpenAI request and response to CSV"""
        import csv
        import time

        with open(self.openai_log_file, 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Extract data
            timestamp = time.strftime('%Y-%m-%d %H:%M:%S')
            model = request_data.get('model', '')
            messages = json.dumps(request_data.get('messages', []))
            tools = json.dumps(request_data.get('tools', []))
            tool_choice = str(request_data.get('tool_choice', ''))

            # Response data
            response_content = ''
            tool_calls = ''
            finish_reason = ''
            prompt_tokens = 0
            completion_tokens = 0
            total_tokens = 0

            if response_data:
                choice = response_data.choices[0] if response_data.choices else None
                if choice:
                    if choice.message.content:
                        response_content = choice.message.content
                    if choice.message.tool_calls:
                        tool_calls = json.dumps([{
                            'id': tc.id,
                            'function': tc.function.name,
                            'arguments': tc.function.arguments
                        } for tc in choice.message.tool_calls])
                    finish_reason = choice.finish_reason or ''

                if hasattr(response_data, 'usage') and response_data.usage:
                    prompt_tokens = response_data.usage.prompt_tokens
                    completion_tokens = response_data.usage.completion_tokens
                    total_tokens = response_data.usage.total_tokens

            # Calculate cost
            cost_usd = calculate_cost(model, prompt_tokens, completion_tokens)

            # Accumulate cost on the current execution
            if self.current_execution:
                self.current_execution.total_cost_usd += cost_usd

            writer.writerow([
                timestamp, task_id, iteration, model,
                messages, tools, tool_choice,
                response_content, tool_calls, finish_reason,
                prompt_tokens, completion_tokens, total_tokens, f"{cost_usd:.6f}"
            ])

    def _call_api_with_hard_timeout(self, request_data: dict, timeout_seconds: int = None) -> Tuple[str, dict]:
        """Make streaming API call with hard signal-based timeout that actually works.

        This is the ONLY reliable way to timeout blocking I/O in sync Python on Unix.
        httpx timeouts don't work when server sends occasional bytes (low-and-slow).

        Args:
            request_data: The request data dict for chat.completions.create
            timeout_seconds: Hard timeout in seconds (default: self.hard_timeout_seconds)

        Returns:
            Tuple of (response_text, usage_info)

        Raises:
            StreamTimeoutError: If the hard timeout is exceeded
        """
        if timeout_seconds is None:
            timeout_seconds = self.hard_timeout_seconds

        def timeout_handler(signum, frame):
            raise StreamTimeoutError(f"API call exceeded {timeout_seconds}s hard timeout - connection likely dead")

        # Set up alarm (Unix only, main thread only)
        old_handler = signal.signal(signal.SIGALRM, timeout_handler)
        signal.alarm(timeout_seconds)

        try:
            stream = self.openai_client.chat.completions.create(**request_data)
            response_text, usage_info = self._collect_stream_response(stream, timeout_seconds)
            return response_text, usage_info
        except StreamTimeoutError:
            print(f"⚠️ Hard timeout after {timeout_seconds}s - recreating client")
            self._recreate_openai_client()
            raise
        finally:
            signal.alarm(0)  # Cancel alarm
            signal.signal(signal.SIGALRM, old_handler)

    def _call_anthropic_api(self, system_prompt: str, user_message: str, task: 'TaskExecution') -> Tuple[str, dict]:
        """Call Anthropic API directly with extended thinking support using streaming.

        Streaming is required for extended thinking operations that may take >10 minutes.

        Args:
            system_prompt: System prompt content
            user_message: User message content
            task: Current task execution for logging

        Returns:
            Tuple of (response_text, usage_info)
        """
        if not self.anthropic_client:
            raise ValueError("Anthropic client not initialized")

        # Build the request
        # Anthropic uses max_tokens for output limit
        # For extended thinking, we need thinking.budget_tokens
        request_kwargs = {
            "model": self.model,  # e.g., "claude-opus-4-5-20251101"
            "max_tokens": self.max_completion_tokens,
            "system": system_prompt,
            "messages": [{"role": "user", "content": user_message}]
        }

        # Add extended thinking if budget is set
        if self.thinking_budget_tokens:
            request_kwargs["thinking"] = {
                "type": "enabled",
                "budget_tokens": self.thinking_budget_tokens
            }
            print(f"🧠 Anthropic extended thinking: budget={self.thinking_budget_tokens} tokens")
        elif self.reasoning_effort:
            # Effort-based thinking models (claude-fable-5, claude-opus-4-8):
            # budget_tokens is rejected with HTTP 400; reasoning depth is
            # requested via output_config ({"effort": "low|medium|high|xhigh|max"}).
            # Thinking must be enabled EXPLICITLY: Fable's is always on, but on
            # Opus 4.8/4.7 omitting `thinking` runs WITHOUT thinking, so effort
            # alone would spend tokens with no reasoning. `{"type": "adaptive"}`
            # is accepted by every effort-based Anthropic model, so set it here
            # for all of them. Budget-based models (Opus 4.6 and older) keep
            # setting thinking_budget_tokens and take the branch above unchanged.
            request_kwargs["thinking"] = {"type": "adaptive"}
            request_kwargs["output_config"] = {"effort": self.reasoning_effort}
            print(
                f"🧠 Anthropic adaptive thinking, effort={self.reasoning_effort} "
                "(output_config)"
            )

        print(f"📡 Calling Anthropic API (streaming): {self.model}")
        start_time = time.time()

        try:
            # Use streaming for extended thinking (required for operations >10 min)
            response_text = ""
            thinking_text = ""
            input_tokens = 0
            output_tokens = 0

            with self.anthropic_client.messages.stream(**request_kwargs) as stream:
                for event in stream:
                    # Handle different event types
                    if hasattr(event, 'type'):
                        if event.type == 'content_block_start':
                            pass  # Block starting
                        elif event.type == 'content_block_delta':
                            delta = event.delta
                            if hasattr(delta, 'text'):
                                response_text += delta.text
                            elif hasattr(delta, 'thinking'):
                                thinking_text += delta.thinking
                        elif event.type == 'message_delta':
                            # Final message with usage
                            if hasattr(event, 'usage'):
                                output_tokens = event.usage.output_tokens
                        elif event.type == 'message_start':
                            if hasattr(event, 'message') and hasattr(event.message, 'usage'):
                                input_tokens = event.message.usage.input_tokens

            elapsed = time.time() - start_time
            print(f"✅ Anthropic response received in {elapsed:.1f}s")

            if thinking_text:
                print(f"💭 Extended thinking: {len(thinking_text)} chars")

            # Build usage info
            usage_info = {
                "prompt_tokens": input_tokens,
                "completion_tokens": output_tokens,
                "total_tokens": input_tokens + output_tokens
            }

            # Log the request (adapted for Anthropic format)
            self._log_anthropic_request(request_kwargs, response_text, usage_info, task.task_id, task.total_iterations)

            return response_text, usage_info

        except Exception as e:
            print(f"❌ Anthropic API error: {e}")
            raise

    def _log_anthropic_request(self, request: dict, response_text: str, usage_info: dict, task_id: str, iteration: int):
        """Log Anthropic API request to CSV for cost tracking."""
        import csv
        try:
            prompt_tokens = usage_info.get("prompt_tokens", 0)
            completion_tokens = usage_info.get("completion_tokens", 0)
            total_tokens = usage_info.get("total_tokens", 0)

            # Calculate cost (Opus 4.5 pricing: $5/MTok input, $25/MTok output)
            cost_usd = (prompt_tokens / 1_000_000 * 5.0) + (completion_tokens / 1_000_000 * 25.0)

            # Accumulate cost on the current execution
            if self.current_execution:
                self.current_execution.total_cost_usd += cost_usd

            with open(self.openai_log_file, "a", newline='') as f:
                writer = csv.writer(f)
                writer.writerow([
                    datetime.now().isoformat(),
                    request.get("model", "unknown"),
                    prompt_tokens, completion_tokens, total_tokens, f"{cost_usd:.6f}"
                ])
        except Exception as e:
            print(f"⚠️ Failed to log Anthropic request: {e}")

    def _collect_stream_response(self, stream, timeout_seconds: int = 300) -> Tuple[str, dict]:
        """Collect streaming response chunks and return full text + usage info.

        Streaming prevents Cloudflare's 100-second idle timeout by receiving
        data continuously as the model generates tokens.

        Args:
            stream: The streaming response from OpenAI
            timeout_seconds: Maximum time to wait for the entire stream (default 5 min)
        """
        import time as time_module

        response_text = ""
        usage_info = {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}
        start_time = time_module.time()
        last_content_time = start_time
        empty_chunk_count = 0
        max_empty_chunks = 100  # Detect dead connection after 100 empty chunks

        try:
            for chunk in stream:
                # Check for overall timeout
                elapsed = time_module.time() - start_time
                if elapsed > timeout_seconds:
                    print(f"⚠️ Stream timeout after {elapsed:.1f}s, returning partial response")
                    break

                # Extract content from delta
                if chunk.choices and len(chunk.choices) > 0:
                    delta = chunk.choices[0].delta
                    # Check both content and reasoning (for thinking models like Qwen3-Thinking, OLMo-Think)
                    has_content = False
                    if delta:
                        if delta.content:
                            response_text += delta.content
                            has_content = True
                        # For thinking models, reasoning tokens come through delta.reasoning
                        if hasattr(delta, 'reasoning') and delta.reasoning:
                            # Don't add reasoning to response_text (it's internal thinking)
                            # But count it as activity to prevent idle timeout
                            has_content = True

                    if has_content:
                        last_content_time = time_module.time()
                        empty_chunk_count = 0  # Reset empty chunk counter
                    else:
                        empty_chunk_count += 1
                else:
                    empty_chunk_count += 1

                # Detect dead connection (many consecutive empty chunks)
                if empty_chunk_count > max_empty_chunks:
                    idle_time = time_module.time() - last_content_time
                    print(f"⚠️ Detected {empty_chunk_count} empty chunks, idle {idle_time:.1f}s, breaking stream")
                    break

                # Extract usage info from the final chunk (if stream_options.include_usage=True)
                if hasattr(chunk, 'usage') and chunk.usage:
                    usage_info = {
                        "prompt_tokens": chunk.usage.prompt_tokens or 0,
                        "completion_tokens": chunk.usage.completion_tokens or 0,
                        "total_tokens": chunk.usage.total_tokens or 0
                    }
        except Exception as e:
            print(f"⚠️ Stream error: {e}, returning partial response ({len(response_text)} chars)")

        return response_text, usage_info

    def _log_streaming_request(self, request_data: dict, response_text: str, usage_info: dict, task_id: str = "", iteration: int = 0):
        """Log streaming OpenAI request and response to CSV"""
        import csv
        import time

        with open(self.openai_log_file, 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Extract data
            timestamp = time.strftime('%Y-%m-%d %H:%M:%S')
            model = request_data.get('model', '')
            messages = json.dumps(request_data.get('messages', []))
            tools = json.dumps(request_data.get('tools', []))
            tool_choice = str(request_data.get('tool_choice', ''))

            # Response data from streaming
            response_content = response_text
            tool_calls = ''
            finish_reason = 'stop'  # Streaming typically ends with stop
            prompt_tokens = usage_info.get('prompt_tokens', 0)
            completion_tokens = usage_info.get('completion_tokens', 0)
            total_tokens = usage_info.get('total_tokens', 0)

            # Calculate cost
            cost_usd = calculate_cost(model, prompt_tokens, completion_tokens)

            # Accumulate cost on the current execution
            if self.current_execution:
                self.current_execution.total_cost_usd += cost_usd

            writer.writerow([
                timestamp, task_id, iteration, model,
                messages, tools, tool_choice,
                response_content, tool_calls, finish_reason,
                prompt_tokens, completion_tokens, total_tokens, f"{cost_usd:.6f}"
            ])

    def _get_system_prompt(self) -> str:
        """Get the system prompt for the AI reasoning engine"""
        if self.system_prompt_path:
            system_prompt_file = Path(self.system_prompt_path)
        else:
            system_prompt_file = Path(__file__).parent / "prompts" / "system_prompt_v0.txt"

        try:
            prompt = system_prompt_file.read_text(encoding='utf-8')

            # Add JSON format instructions based on custom_reasoning flag
            if not self.custom_reasoning:
                # Compact format without reasoning field (saves tokens)
                prompt += "\n\n## JSON RESPONSE FORMAT (COMPACT)\n"
                prompt += "Respond with JSON in this exact format (NO 'reasoning' field to save tokens):\n"
                prompt += '{"is_complete": boolean, "actions": [{"tool": "tool_name", "parameters": {...}}]}\n'
                prompt += "The 'reasoning' field is OMITTED to reduce token usage and avoid context limits.\n"
            else:
                # Verbose format with reasoning field
                prompt += "\n\n## JSON RESPONSE FORMAT (WITH REASONING)\n"
                prompt += "Respond with JSON including reasoning field:\n"
                prompt += '{"reasoning": "your explanation...", "is_complete": boolean, "actions": [{"tool": "tool_name", "parameters": {...}}]}\n'

            return prompt
        except Exception as e:
            # Fallback to default if file not found
            print(f"⚠️ Warning: Could not load system prompt from {system_prompt_file}: {e}")
            return "You are an intelligent assistant with Excel automation capabilities."

    def _get_context_prompt(self, task: TaskExecution) -> str:
        """Generate context prompt - fresh mode vs traditional history mode"""
        if self.fresh_context_mode:
            return self._get_fresh_context_prompt(task)
        else:
            return self._get_traditional_context_prompt(task)

    def _get_fresh_context_prompt(self, task: TaskExecution) -> str:
        """Generate fresh context prompt with current file state and recent tool call history"""
        context = f"""CURRENT TASK: {task.user_prompt}
ITERATION: {task.total_iterations}/{task.max_iterations}

{self._load_solution_context()}
"""

        # Add recent tool call history (last N calls)
        if self.recent_history_count > 0 and task.steps:
            recent_steps = task.steps[-self.recent_history_count:]
            context += f"\n=== RECENT ACTIONS (last {len(recent_steps)}) ===\n"
            for step in recent_steps:
                status_icon = "✅" if not step.error else "❌"
                context += f"\n{status_icon} Step {step.step_number}: {step.tool_name}\n"
                # Show tool arguments (compact)
                args_str = json.dumps(step.tool_args, separators=(',', ':'))
                if len(args_str) > 150:
                    args_str = args_str[:147] + "..."
                context += f"   Args: {args_str}\n"
                # Show result summary
                if step.error:
                    context += f"   Error: {step.error[:100]}\n"
                elif step.result:
                    result_str = str(step.result.get('result', ''))
                    if len(result_str) > 150:
                        result_str = result_str[:147] + "..."
                    context += f"   Result: {result_str}\n"
            context += "\n"

        # Note: PDF files are attached to this request as multimodal content
        if self.context_pdfs:
            context += f"\n\n=== ATTACHED PDFs ({len(self.context_pdfs)}) ===\n"
            for pdf_path in self.context_pdfs:
                context += f"- {Path(pdf_path).name}\n"
            context += """
🔥 IMPORTANT - PDF TEXT IS EMBEDDED BELOW! 🔥
The extracted text from these PDFs will appear below in sections marked like this:

    ============================================================
    PDF: filename.pdf
    ============================================================

    --- Page 1 ---
    [PDF text content here - read it directly!]

📖 DO NOT use get_cell_range or any Excel tools to access PDFs!
✅ Just scroll down and find the "PDF: filename" sections - that's your PDF content!
============================================================
"""

        # Enhanced Excel context files (if any)
        if self.context_excels:
            context += f"\n\n=== ATTACHED EXCEL CONTEXT ({len(self.context_excels)}) ===\n"
            for excel_path in self.context_excels:
                context += f"- {Path(excel_path).name}\n"
            # Note: Enhanced Excel context will be added at the end via _extract_excel_texts()

        context += f"\n🎯 FOCUS: Use the current file state above to determine your next action."
        context += f"\n💡 TIP: You can see exactly what's in each cell with column letters and row numbers."
        context += f"\n⚠️  CRITICAL: Never reference the same cell you're putting a formula in (circular reference)."

        return context

    def _get_traditional_context_prompt(self, task: TaskExecution) -> str:
        """Generate traditional context prompt with execution history"""
        context = f"""
        CURRENT TASK: {task.user_prompt}
        TASK_ID: {task.task_id}
        """

        # Note: PDF files are attached to this request as multimodal content
        if self.context_pdfs:
            context += f"\n\n=== ATTACHED PDFs ({len(self.context_pdfs)}) ===\n"
            for pdf_path in self.context_pdfs:
                context += f"- {Path(pdf_path).name}\n"
            context += """
🔥 IMPORTANT - PDF TEXT IS EMBEDDED BELOW! 🔥
The extracted text from these PDFs will appear below in sections marked like this:

    ============================================================
    PDF: filename.pdf
    ============================================================

    --- Page 1 ---
    [PDF text content here - read it directly!]

📖 DO NOT use get_cell_range or any Excel tools to access PDFs!
✅ Just scroll down and find the "PDF: filename" sections - that's your PDF content!
============================================================
"""

        # Note: Excel files are attached to this request as context
        if self.context_excels:
            context += f"\n\n=== ATTACHED EXCEL FILES ({len(self.context_excels)}) ===\n"
            for excel_path in self.context_excels:
                context += f"- {Path(excel_path).name}\n"
            context += "=== The Excel file content is attached to this message for your analysis ===\n"

        context += """
EXECUTION HISTORY:
"""
        for step in task.steps:
            context += f"\nStep {step.step_number}: {step.description}"
            context += f"\n  Tool: {step.tool_name}({step.tool_args})"
            if step.result:
                context += f"\n  Result: {json.dumps(step.result.get('result', 'No result'), indent=2)[:500]}..."
            if step.error:
                context += f"\n  Error: {step.error}"

        context += f"\n\nCURRENT STATUS: {task.status.value}"
        context += f"\nITERATIONS: {task.total_iterations}/{task.max_iterations}"

        return context



    def _detect_loop(self, task: TaskExecution, lookback: int = 3) -> bool:
        """Detect if the agent is stuck in a loop repeating failed actions"""
        if len(task.steps) < lookback:
            return False

        recent_steps = task.steps[-lookback:]

        # Check if same tool with same args repeated multiple times
        tool_calls = [
            (s.tool_name, json.dumps(s.tool_args, sort_keys=True))
            for s in recent_steps
        ]

        # Check if all recent steps are the same action
        if len(set(tool_calls)) == 1:
            # Check if they all failed or had no useful result
            all_failed = all(
                s.error or
                (s.result and (
                    s.result.get('result') == [] or
                    s.result.get('result') == {} or
                    not s.result.get('success')
                ))
                for s in recent_steps
            )
            return all_failed

        return False

    def _call_reasoning_engine(self, task: TaskExecution) -> Tuple[Dict[str, Any], str]:
        """Call OpenAI to determine next action.

        Returns (parsed_json, raw_text) for auditing.
        """
        try:
            system_prompt = self._get_system_prompt()
            context_prompt = self._get_context_prompt(task)

            # Extract PDF and Excel text if they are in context
            additional_context = ""
            if self.context_pdfs:
                pdf_text = self._extract_pdf_texts()
                additional_context += "\n\n" + pdf_text
            if self.context_excels:
                excel_text = self._extract_excel_texts()
                additional_context += "\n\n" + excel_text

            if additional_context:
                full_context = context_prompt + additional_context
                user_message_content = full_context
            else:
                user_message_content = context_prompt

            # Store the exact prompt for snapshot capture
            self._last_user_message = user_message_content

            # Use Anthropic direct API if configured (for Claude models with extended thinking)
            if self.use_anthropic_direct:
                response_text, usage_info = self._call_anthropic_api(
                    system_prompt, user_message_content, task
                )
                # Skip to JSON parsing (after the OpenAI try block)
            else:
                user_message = {"role": "user", "content": user_message_content}
                # OpenAI wrapper will automatically create generations with Langfuse
                # We just enhance the metadata for better tracking
                # Use STREAMING to prevent Cloudflare 100s timeout on long-running requests
                try:
                    request_data = {
                        "model": self.model,
                        "messages": [
                            {"role": "system", "content": system_prompt},
                            user_message
                        ],
                        "max_completion_tokens": self.max_completion_tokens,
                        "stream": True,  # Enable streaming to prevent Cloudflare timeout
                        "stream_options": {"include_usage": True}  # Get token usage in stream
                    }

                    # Add reasoning effort for reasoning models (GPT-5 series, o3, etc.)
                    if self.reasoning_effort:
                        if self.use_openai_direct:
                            # OpenAI direct API: reasoning_effort is a top-level parameter
                            request_data["reasoning_effort"] = self.reasoning_effort
                            print(f"🧠 Reasoning effort (OpenAI Direct): {self.reasoning_effort}")
                        else:
                            # OpenRouter: nested under extra_body
                            request_data["extra_body"] = {
                                "reasoning": {"effort": self.reasoning_effort}
                            }
                            print(f"🧠 Reasoning effort (OpenRouter): {self.reasoning_effort}")

                    # Note: We rely on _extract_json() and _parse_jsonl_response() to handle
                    # any preamble text (e.g., "Wait: ...") that GPT-5 with reasoning may output.
                    # The response_format: {"type": "json_object"} was removed as it caused issues.

                    # Add Langfuse metadata if enabled (for OpenRouter only)
                    # Note: metadata must go through extra_body, not as a top-level SDK param
                    if self.langfuse_enabled and not self.use_openai_direct:
                        metadata = {
                            "task_id": str(task.task_id),
                            "iteration": str(task.total_iterations),
                            "user_prompt": task.user_prompt[:200],  # Truncate for readability
                            "status": str(task.status.value),
                            "step_number": str(len(task.steps) + 1),
                            "context_pdfs": str(len(self.context_pdfs)),
                            "context_excels": str(len(self.context_excels)),
                            "has_errors": str(any(s.error for s in task.steps))
                        }
                        if "extra_body" in request_data:
                            request_data["extra_body"]["metadata"] = metadata
                        else:
                            request_data["extra_body"] = {"metadata": metadata}

                    # Stream the response with hard timeout and retry logic
                    # This prevents hanging on dead VM connections
                    MAX_RETRIES = 3
                    last_error = None

                    for attempt in range(MAX_RETRIES):
                        try:
                            response_text, usage_info = self._call_api_with_hard_timeout(request_data)
                            # Log the streamed response
                            self._log_streaming_request(request_data, response_text, usage_info, task.task_id, task.total_iterations)
                            break  # Success, exit retry loop
                        except StreamTimeoutError as timeout_err:
                            last_error = timeout_err
                            print(f"⚠️ Attempt {attempt+1}/{MAX_RETRIES} failed: {timeout_err}")
                            if attempt < MAX_RETRIES - 1:
                                backoff_time = 5 * (attempt + 1)
                                print(f"⏳ Waiting {backoff_time}s before retry...")
                                time.sleep(backoff_time)
                            else:
                                print(f"❌ All {MAX_RETRIES} attempts failed, raising error")
                                raise
                        except httpx.TimeoutException as httpx_err:
                            last_error = httpx_err
                            print(f"⚠️ httpx timeout on attempt {attempt+1}/{MAX_RETRIES}: {httpx_err}")
                            self._recreate_openai_client()
                            if attempt < MAX_RETRIES - 1:
                                backoff_time = 5 * (attempt + 1)
                                print(f"⏳ Waiting {backoff_time}s before retry...")
                                time.sleep(backoff_time)
                            else:
                                print(f"❌ All {MAX_RETRIES} attempts failed, raising error")
                                raise

                except Exception as api_err:
                    # Fallback if streaming isn't supported by the model
                    err_text = str(api_err)
                    if "stream" in err_text.lower() or "unsupported_parameter" in err_text:
                        request_data = {
                            "model": self.model,
                            "messages": [
                                {"role": "system", "content": system_prompt},
                                user_message
                            ],
                            "max_completion_tokens": self.max_completion_tokens
                        }

                        # Add Langfuse metadata if enabled
                        if self.langfuse_enabled:
                            request_data["extra_body"] = {"metadata": {
                                "task_id": task.task_id,
                                "iteration": task.total_iterations,
                                "user_prompt": task.user_prompt[:200],
                                "status": task.status.value,
                                "step_number": len(task.steps) + 1,
                                "context_pdfs": len(self.context_pdfs),
                                "has_errors": any(s.error for s in task.steps),
                                "fallback_mode": True
                            }}

                        response = self.openai_client.chat.completions.create(**request_data)
                        self._log_openai_request(request_data, response, task.task_id, task.total_iterations)
                        response_text = response.choices[0].message.content or ""
                    else:
                        raise
            
            # Parse JSON response with light cleanup for fenced blocks
            def _extract_json(txt: str) -> Dict[str, Any]:
                t = (txt or "").strip()
                if t.startswith("```"):
                    # strip leading fence (optionally with language) and trailing fence
                    lines = t.splitlines()
                    # drop the first fence line
                    if lines:
                        lines = lines[1:]
                    # drop last fence line if present
                    if lines and lines[-1].strip().startswith("```"):
                        lines = lines[:-1]
                    t = "\n".join(lines).strip()
                def _loads_candidate(candidate: str) -> Dict[str, Any]:
                    # First, raw attempt
                    return json.loads(candidate)

                def _coerce_json(candidate: str) -> Dict[str, Any]:
                    s = candidate
                    # Collapse whitespace
                    s = s.replace('\r', '\n')
                    # Remove JSON comments (// style)
                    s = re.sub(r'//.*$', '', s, flags=re.MULTILINE)
                    # Extract innermost object if extra text around
                    start = s.find('{')
                    end = s.rfind('}')
                    if start != -1 and end != -1 and end > start:
                        s = s[start:end+1]
                    # Replace True/False/None with JSON booleans/null
                    s = re.sub(r"\bTrue\b", "true", s)
                    s = re.sub(r"\bFalse\b", "false", s)
                    s = re.sub(r"\bNone\b", "null", s)
                    # Quote bare keys: { key: value } -> { "key": value }
                    s = re.sub(r'([\{,]\s*)([A-Za-z_][A-Za-z0-9_\- ]*)(\s*:)','\\1"\\2"\\3', s)
                    # Convert single-quoted strings to double-quoted (basic)
                    s = re.sub(r":\s*'([^'\\]*(?:\\.[^'\\]*)*)'", lambda m: ": \"" + m.group(1).replace("\\\"","\"").replace("\"","\"").replace("\\'","'") + "\"", s)
                    s = re.sub(r"\[\s*'([^\]]*?)'\s*\]", lambda m: "[" + m.group(1).replace("'", '"') + "]", s)
                    # Remove trailing commas before } or ]
                    s = re.sub(r",\s*([}\]])", r"\1", s)
                    return json.loads(s)

                try:
                    return _loads_candidate(t)
                except Exception:
                    # try to find first {...} block and load
                    start = t.find('{')
                    end = t.rfind('}')
                    if start != -1 and end != -1 and end > start:
                        sub = t[start:end+1]
                        try:
                            return _loads_candidate(sub)
                        except Exception:
                            return _coerce_json(sub)
                    # last resort coercion
                    return _coerce_json(t)

            def _split_concatenated_json(text: str) -> list:
                """Split concatenated top-level JSON objects like {...}{...}{...}.

                Tracks brace nesting depth so inner '}{' (inside nested objects/arrays)
                are not treated as object boundaries.  String literals and escaped
                quotes are handled correctly.
                """
                objects = []
                depth = 0
                start = 0
                in_string = False
                escape_next = False

                for i, ch in enumerate(text):
                    if escape_next:
                        escape_next = False
                        continue
                    if ch == '\\' and in_string:
                        escape_next = True
                        continue
                    if ch == '"' and not escape_next:
                        in_string = not in_string
                        continue
                    if in_string:
                        continue
                    if ch == '{':
                        if depth == 0:
                            start = i
                        depth += 1
                    elif ch == '}':
                        depth -= 1
                        if depth == 0:
                            objects.append(text[start:i+1])

                return objects

            def _parse_jsonl_response(jsonl_text: str) -> Dict[str, Any]:
                """Parse JSONL format response (multiple JSON objects on separate lines or concatenated).

                This handles GPT-5.1's tendency to output multiple JSON objects separated by newlines,
                and GPT-5.2's tendency to concatenate JSON objects without separators (e.g., {...}{...}).
                Combines them into a single structured response compatible with our action system.
                """
                # First, try to split on }{ for concatenated JSON objects (GPT-5.2 style)
                text = jsonl_text.strip()
                if '}{' in text:
                    # Brace-depth-aware split — only splits at top-level boundaries
                    lines = _split_concatenated_json(text)
                    print(f"📋 DEBUG: Split concatenated JSON into {len(lines)} objects")
                else:
                    # Standard JSONL - split on newlines
                    lines = text.split('\n')

                all_actions = []
                is_complete = False
                reasoning = ""

                for line_num, line in enumerate(lines, 1):
                    line = line.strip()
                    if not line:
                        continue

                    try:
                        obj = json.loads(line)

                        # Handle different types of JSON objects
                        if isinstance(obj, dict):
                            # Extract reasoning from first object
                            if "reasoning" in obj and not reasoning:
                                reasoning = obj["reasoning"]

                            # Extract completion status
                            if "is_complete" in obj:
                                is_complete = obj["is_complete"]

                            # Extract actions from objects with actions array
                            if "actions" in obj and isinstance(obj["actions"], list):
                                for action in obj["actions"]:
                                    if isinstance(action, dict):
                                        # Normalize action schema
                                        if "tool" in action and "parameters" in action:
                                            all_actions.append(action)
                                        elif "tool_name" in action and "arguments" in action:
                                            all_actions.append({
                                                "tool": action["tool_name"],
                                                "parameters": action["arguments"]
                                            })
                                        else:
                                            all_actions.append(action)

                            # Single action objects (just tool + parameters)
                            elif "tool" in obj and "parameters" in obj:
                                all_actions.append(obj)
                            elif "tool_name" in obj and "arguments" in obj:
                                # Handle GPT-5.1 schema variant
                                all_actions.append({
                                    "tool": obj["tool_name"],
                                    "parameters": obj["arguments"]
                                })

                    except json.JSONDecodeError as e:
                        print(f"📋 DEBUG: Skipping malformed JSON on line {line_num}: {e}")
                        continue

                # Construct unified response
                result = {
                    "reasoning": reasoning or "JSONL response processed successfully",
                    "is_complete": is_complete,
                    "actions": all_actions
                }

                print(f"📋 DEBUG: JSONL parsing successful - merged {len(all_actions)} actions from {len(lines)} lines")
                return result

            # Hybrid parsing: try standard JSON first, fallback to JSONL for GPT-5.1
            try:
                # First try standard JSON parsing (GPT-4o, proper behavior)
                return _extract_json(response_text), response_text
            except Exception as e:
                # If standard JSON fails, try JSONL parsing (GPT-5.1 workaround)
                try:
                    print(f"📋 DEBUG: Standard JSON failed, attempting JSONL parsing...")
                    return _parse_jsonl_response(response_text), response_text
                except Exception as jsonl_error:
                    print(f"📋 DEBUG: JSONL parsing also failed: {str(jsonl_error)}")
                    # Both parsers failed, continue with original error handling
                    print(f"📋 DEBUG: JSON parsing failed with error: {str(e)}")
                    print(f"📋 DEBUG: Response length: {len(response_text)} chars")
                    print(f"📋 DEBUG: FULL RESPONSE:")
                    print(response_text)
                    print(f"📋 DEBUG: END OF RESPONSE")
                    return ({
                        "reasoning": f"Failed to parse AI response: {response_text}",
                        "action": {"tool": "request_clarification", "parameters": {"question": "Failed to parse response"}},
                        "is_complete": False,
                        "error": f"JSON parsing error: {str(e)}"
                    }, response_text)

        except Exception as e:
            return ({
                "reasoning": f"Error calling reasoning engine: {str(e)}",
                "action": {"tool_name": "request_clarification", "arguments": {"question": str(e)}},
                "is_complete": False,
                "error": str(e)
            }, str(e))

    def _execute_action(self, action: Dict[str, Any]) -> Dict[str, Any]:
        """Execute an MCP tool action"""
        # Support both schema formats: tool/parameters (new) and tool_name/arguments (legacy)
        tool_name = action.get("tool") or action.get("tool_name")
        arguments = action.get("parameters") or action.get("arguments", {})

        # Handle special tool requests
        if tool_name == "request_clarification":
            question = arguments.get('question', 'Additional details required') if isinstance(arguments, dict) else 'Additional details required'
            return {
                "success": False,
                "tool": tool_name,
                "result": f"Clarification needed: {question}",
                "error": f"Clarification needed: {question}",
                "needs_user_input": True,
            }

        # If filing an MCP issue, attach task_id by default
        if tool_name == "report_mcp_issue":
            if isinstance(arguments, dict) and "task_id" not in arguments and self.current_execution:
                arguments = {**arguments, "task_id": self.current_execution.task_id}

        # CRITICAL: Detect attempts to use Excel tools on PDF files (Fix #2)
        # Excel tools only work on .xlsx files, NOT .pdf files
        excel_tools_with_filename = [
            "get_cell_range", "get_formula", "list_worksheets", "create_worksheet",
            "delete_worksheet", "edit_cells", "set_cell_formula",
            "get_used_range", "scan_worksheet_structure", "search_worksheet",
            "summarize_workbook_context", "describe_worksheet", "get_file_metadata",
            "copy_file", "delete_file"
        ]

        if tool_name in excel_tools_with_filename and isinstance(arguments, dict):
            filename = arguments.get("filename", "")
            if filename and filename.lower().endswith(".pdf"):
                error_msg = (
                    f"❌ ERROR: You tried to use Excel tool '{tool_name}' on PDF file '{filename}'.\n\n"
                    f"📋 PDFs are NOT Excel files!\n\n"
                    f"✅ The PDF text is ALREADY IN YOUR PROMPT - just read it directly!\n"
                    f"   Look for the section marked: 'PDF: {filename}'\n\n"
                    f"🚫 DO NOT use get_cell_range, list_worksheets, or ANY Excel tools on PDFs.\n"
                    f"📖 The PDF text content is embedded in your system prompt - scroll up and find it!\n\n"
                    f"Excel tools work ONLY on .xlsx files."
                )
                return {
                    "success": False,
                    "tool": tool_name,
                    "error": error_msg,
                    "needs_user_input": False
                }

        # Create Langfuse child span for tool execution (v3 API)
        span_obj = None
        if self.langfuse_enabled and self._current_trace:
            try:
                span_obj = self._current_trace.start_span(
                    name=f"tool_{tool_name}",
                    input={"tool_name": tool_name, "arguments": arguments}
                )
                # Add metadata
                span_obj.update(
                    metadata={
                        "step_number": len(self.current_execution.steps) + 1 if self.current_execution else 0
                    }
                )
            except Exception as e:
                print(f"⚠️ Warning: Could not create Langfuse span: {e}")

        # Execute MCP tool normally
        result = self.excel_client.call_tool(tool_name, arguments)

        # Update and end span with results (v3 API)
        if span_obj:
            try:
                span_obj.update(
                    output=result,
                    metadata={
                        "success": result.get("success", False),
                        "error": result.get("error"),
                        "has_result": "result" in result
                    }
                )
                span_obj.end()
            except Exception:
                pass

        return result

    def _get_prompt_files(self, task: TaskExecution, task_folder: Path) -> List[str]:
        """Collect all prompt files for database schema (absolute paths)"""
        prompt_files = []

        # 1. System prompt file
        system_prompt_path = Path(__file__).parent / "prompts" / "system_prompt.txt"
        if system_prompt_path.exists():
            prompt_files.append(str(system_prompt_path.absolute()))

        # 2. MCP config files (if they exist)
        mcp_config_candidates = [
            Path(__file__).parent / "mcp_config.json",
            Path(__file__).parent.parent / "mcp_config.json",
        ]
        for config_path in mcp_config_candidates:
            if config_path.exists():
                prompt_files.append(str(config_path.absolute()))

        # 3. Save task prompt to file and add to list
        task_prompt_path = task_folder / "task_prompt.txt"
        try:
            task_prompt_path.write_text(task.user_prompt, encoding="utf-8")
            prompt_files.append(str(task_prompt_path.absolute()))
        except Exception:
            pass  # Don't fail if we can't save the task prompt

        return prompt_files

    def _serialize_task(self, task: TaskExecution, task_folder: Optional[Path] = None) -> Dict[str, Any]:
        """Serialize a task with all details for persistence (including database schema fields)"""
        # Calculate time-related fields
        start_time_iso = datetime.fromtimestamp(task.start_time).isoformat() + "Z" if task.start_time else None
        end_time_iso = datetime.fromtimestamp(task.end_time).isoformat() + "Z" if task.end_time else None
        time_taken_mins = None
        if task.start_time and task.end_time:
            time_taken_mins = (task.end_time - task.start_time) / 60.0

        # Get prompt files if task folder is provided
        prompt_files = []
        if task_folder:
            prompt_files = self._get_prompt_files(task, task_folder)

        # Get attempt files (newly created files)
        attempt_files = self.excel_client.get_created_files()

        # Fallback: check if solution.xlsx exists but wasn't tracked
        # This handles cases where task hit max iterations but still created a solution
        solution_path = Path(self.excel_client.storage_path) / "solution.xlsx"
        if solution_path.exists():
            solution_str = str(solution_path)
            if solution_str not in attempt_files:
                attempt_files.append(solution_str)

        # Format agent model name as openpyxl_{model}
        agent_model_name = f"openpyxl_{self.model}"

        # Build database-compatible structure
        return {
            # Database schema fields (Task Attempts Table)
            "task_id": task.task_id,
            "agent_model_name": agent_model_name,
            "agent_model_type": "api",
            "prompt_files": prompt_files,
            "start_end_times": [start_time_iso, end_time_iso] if start_time_iso and end_time_iso else [],
            "attempt_files": attempt_files,
            "time_taken_mins": time_taken_mins,
            "cost": None,  # Placeholder for future cost tracking

            # Legacy/additional fields (keep for backward compatibility)
            "user_prompt": task.user_prompt,
            "status": task.status.value,
            "start_time": task.start_time,
            "end_time": task.end_time,
            "iterations": task.total_iterations,
            "max_iterations": task.max_iterations,
            "final_result": task.final_result,
            "error": task.error,
            "steps": [
                {
                    "step_number": s.step_number,
                    "description": s.description,
                    "tool_name": s.tool_name,
                    "tool_args": s.tool_args,
                    "result": s.result,
                    "execution_time": s.execution_time,
                    "error": s.error,
                    "reasoning_json": s.reasoning_json,
                    "raw_model_response": s.raw_model_response,
                }
                for s in task.steps
            ],
        }

    def _write_markdown_summary(self, task: TaskExecution, path: Path) -> None:
        """Create a human-readable transcript for quick review"""
        lines: List[str] = []
        lines.append(f"# Task {task.task_id}")
        lines.append("")
        lines.append(f"Task: {task.user_prompt}")
        lines.append(f"Status: {task.status.value}")
        if task.final_result:
            lines.append(f"Final Result: {task.final_result}")
        if task.error:
            lines.append(f"Error: {task.error}")
        lines.append("")
        lines.append("## Steps")
        for s in task.steps:
            lines.append(f"### Step {s.step_number}: {s.tool_name}")
            lines.append("")
            if s.description:
                lines.append(f"Reasoning: {s.description}")
            if s.reasoning_json is not None:
                lines.append("Reasoning JSON:")
                try:
                    lines.append("```")
                    lines.append(json.dumps(s.reasoning_json, indent=2))
                    lines.append("```")
                except Exception:
                    pass
            lines.append("Tool Call:")
            try:
                lines.append("```")
                lines.append(json.dumps({"tool": s.tool_name, "args": s.tool_args}, indent=2))
                lines.append("```")
            except Exception:
                lines.append(f"{s.tool_name}({s.tool_args})")
            if s.result is not None:
                lines.append("Result:")
                try:
                    lines.append("```")
                    lines.append(json.dumps(s.result, indent=2))
                    lines.append("```")
                except Exception:
                    lines.append(str(s.result))
            if s.error:
                lines.append(f"Error: {s.error}")
            lines.append("")
        path.write_text("\n".join(lines), encoding="utf-8")

    def _persist_task(self, task: TaskExecution) -> None:
        """Persist the task JSON + markdown transcript under storage/agent_logs"""
        try:
            folder = self.logs_dir / task.task_id
            folder.mkdir(parents=True, exist_ok=True)
            # Pass task_folder to collect prompt files and save task prompt
            (folder / "task.json").write_text(
                json.dumps(self._serialize_task(task, task_folder=folder), indent=2), encoding="utf-8"
            )
            self._write_markdown_summary(task, folder / "transcript.md")
        except Exception:
            # Never allow logging to break the run
            pass

    def execute_task(self, user_prompt: str, task_id: Optional[str] = None, resume: Optional[TaskExecution] = None) -> TaskExecution:
        """Execute a task with iterative reasoning.

        If "resume" is provided, continues execution using the existing TaskExecution
        (preserving steps/history) and ignores "task_id". Otherwise, starts a new task.
        """
        if resume is not None:
            task = resume
            task.status = TaskStatus.IN_PROGRESS
            # keep existing task_id, start_time, steps, etc.
        else:
            # Starting a new task - clear created files tracking
            self.excel_client.clear_created_files()

            if task_id is None:
                task_id = f"task_{int(time.time())}"
            task = TaskExecution(
                task_id=task_id,
                user_prompt=user_prompt,
                status=TaskStatus.IN_PROGRESS,
                steps=[],
                start_time=time.time(),
                max_iterations=self.default_max_iterations
            )

        self.current_execution = task

        # Create Langfuse root span for this task execution (v3 API)
        if self.langfuse_enabled and self.langfuse_client:
            try:
                self._current_trace = self.langfuse_client.start_span(
                    name="excel_task_execution",
                    input={"user_prompt": user_prompt, "task_id": task.task_id}
                )
                # Set trace-level attributes
                self._current_trace.update_trace(
                    user_id=task.task_id,
                    session_id=task.task_id,
                    tags=["excel_automation", self.model, task.status.value],
                    metadata={
                        "model": self.model,
                        "max_iterations": task.max_iterations,
                        "storage_path": self.excel_client.storage_path,
                        "context_pdfs": [Path(p).name for p in self.context_pdfs],
                        "pdf_count": len(self.context_pdfs),
                        "verbose_mode": self.verbose,
                        "is_resumed": resume is not None
                    }
                )
            except Exception as e:
                print(f"⚠️ Warning: Could not create Langfuse span: {e}")
                self._current_trace = None

        print(f"🚀 Starting task execution: {user_prompt}")
        print(f"📊 Task ID: {task_id}")
        print(f"⚙️ Max iterations: {task.max_iterations}")
        if self.langfuse_enabled:
            print(f"🪢 Langfuse tracking: enabled")
        else:
            print(f"📝 Logging: CSV only (Langfuse not configured)")

        try:
            # Ensure Excel client is connected
            if not self.excel_client.is_connected():
                self.excel_client.connect()

            # Iterative execution loop
            while task.total_iterations < task.max_iterations and task.status == TaskStatus.IN_PROGRESS:
                task.total_iterations += 1

                print(f"\n🤖 Iteration {task.total_iterations}/{task.max_iterations}")

                # Get next action from reasoning engine
                reasoning_result, raw_text = self._call_reasoning_engine(task)

                # Print full AI response if verbose mode is enabled
                if self.verbose:
                    print(f"\n{'='*80}")
                    print(f"🔍 FULL AI RESPONSE (Iteration {task.total_iterations}):")
                    print(f"{'='*80}")
                    print(raw_text)
                    print(f"{'='*80}\n")

                print(f"💭 Reasoning: {reasoning_result.get('reasoning', 'No reasoning provided')}")

                # Check if task is complete (NO VALIDATION - agent decides when complete)
                # BUT: if there are also actions to execute, run them first before marking complete
                # This handles JSONL/concatenated responses where is_complete=true appears
                # alongside pending actions (GPT-5.2 pattern)
                has_pending_actions = bool(
                    reasoning_result.get("action") or reasoning_result.get("actions")
                )
                if reasoning_result.get("is_complete", False) and not has_pending_actions:
                    # Record completion step
                    completion_step = ExecutionStep(
                        step_number=len(task.steps) + 1,
                        description=reasoning_result.get("reasoning", "Task marked complete")[:100],
                        tool_name="complete",
                        tool_args={},
                        result={
                            "success": True,
                            "result": reasoning_result.get("completion_summary")
                        },
                        reasoning_json=reasoning_result,
                        raw_model_response=raw_text,
                    )
                    task.steps.append(completion_step)
                    self._persist_task(task)
                    task.status = TaskStatus.COMPLETED
                    task.final_result = reasoning_result.get("completion_summary", "Task completed")
                    print(f"✅ Task completed: {task.final_result}")
                    # Save final snapshot before exiting loop
                    if self.snapshot_iterations:
                        self._save_iteration_snapshot(task.total_iterations, task)
                    break

                # Execute action(s) - support both single action and multiple actions
                single_action = reasoning_result.get("action")
                multiple_actions = reasoning_result.get("actions")

                # Normalize GPT-5.1 schema while maintaining GPT-4o compatibility
                def add_schema_aliases(action):
                    """Add alias fields to support both GPT-4o and GPT-5.1 schemas"""
                    if action is None:
                        return None
                    # Add GPT-4o aliases if GPT-5.1 format is detected
                    if "tool_name" in action and "tool" not in action:
                        action["tool"] = action["tool_name"]
                    if "arguments" in action and "parameters" not in action:
                        action["parameters"] = action["arguments"]
                    return action

                if single_action:
                    single_action = add_schema_aliases(single_action)
                if multiple_actions:
                    multiple_actions = [add_schema_aliases(action) for action in multiple_actions]

                if not single_action and not multiple_actions:
                    # Agent didn't provide a tool call - refeed context and continue
                    # Completion ONLY happens via explicit is_complete=true flag
                    print(f"⚠️  No tool call provided - refeeding context to agent")
                    # Save snapshot even when no actions executed
                    if self.snapshot_iterations:
                        self._save_iteration_snapshot(task.total_iterations, task)
                    continue

                # Determine which format was used
                if multiple_actions:
                    # Multiple actions - execute sequentially with fail-fast
                    print(f"🔧 Executing {len(multiple_actions)} actions sequentially...")
                    actions_to_execute = multiple_actions
                else:
                    # Single action - wrap in list for uniform processing
                    actions_to_execute = [single_action]

                # Execute actions sequentially
                all_action_results = []
                failed_action_index = None

                for action_idx, action in enumerate(actions_to_execute):
                    resolved_tool = action.get("tool") or action.get("tool_name") or "unknown"
                    resolved_args = action.get("parameters") or action.get("arguments") or {}
                    step = ExecutionStep(
                        step_number=len(task.steps) + 1,
                        description=f"Action {action_idx + 1}/{len(actions_to_execute)}: {resolved_tool}",
                        tool_name=resolved_tool,
                        tool_args=resolved_args,
                        reasoning_json=reasoning_result if action_idx == 0 else None,
                        raw_model_response=raw_text if action_idx == 0 else None
                    )

                    print(f"🔧 [{action_idx + 1}/{len(actions_to_execute)}] Executing: {step.tool_name}({step.tool_args})")

                    # Execute the action with error handling
                    step_start = time.time()
                    try:
                        step.result = self._execute_action(action)
                        step.execution_time = time.time() - step_start
                    except ValueError as e:
                        # Tool doesn't exist - feed error back to AI for self-correction
                        step.execution_time = time.time() - step_start
                        step.result = {
                            "success": False,
                            "error": str(e),
                            "tool": step.tool_name,
                            "arguments": step.tool_args
                        }
                        step.error = str(e)
                        print(f"❌ Invalid tool requested: {step.error}")
                    except Exception as e:
                        # Other errors - also handle gracefully
                        step.execution_time = time.time() - step_start
                        step.result = {
                            "success": False,
                            "error": f"Execution failed: {str(e)}",
                            "tool": step.tool_name,
                            "arguments": step.tool_args
                        }
                        step.error = str(e)
                        print(f"❌ Tool execution error: {step.error}")

                    # Check for errors
                    if not step.result.get("success", True):
                        step.error = step.result.get("error", "Unknown error")
                        print(f"❌ Error: {step.error}")

                        # If we have user input request, handle it
                        if step.result.get("needs_user_input"):
                            task.status = TaskStatus.NEEDS_CLARIFICATION
                            failed_action_index = action_idx
                            task.steps.append(step)
                            all_action_results.append({
                                "action_number": action_idx + 1,
                                "tool_name": step.tool_name,
                                "success": False,
                                "error": step.error
                            })
                            break
                        else:
                            # Regular failure - mark and stop execution
                            failed_action_index = action_idx
                            task.steps.append(step)
                            all_action_results.append({
                                "action_number": action_idx + 1,
                                "tool_name": step.tool_name,
                                "success": False,
                                "error": step.error
                            })
                            break
                    else:
                        # Check if the result indicates a failed operation (common with MCP tools)
                        result_str = str(step.result.get('result', 'No result'))

                        # Parse JSON result to check for success=false pattern
                        is_actual_failure = False
                        try:
                            import json
                            if result_str.strip().startswith('{') and result_str.strip().endswith('}'):
                                parsed_result = json.loads(result_str)
                                if isinstance(parsed_result, dict) and parsed_result.get('success') is False:
                                    is_actual_failure = True
                        except:
                            pass

                        if is_actual_failure:
                            print(f"❌ Failed: {result_str[:100]}...")
                            step.error = f"Operation failed: {result_str[:200]}"
                            all_action_results.append({
                                "action_number": action_idx + 1,
                                "tool_name": step.tool_name,
                                "success": False,
                                "error": step.error
                            })
                            failed_action_index = action_idx
                            break
                        else:
                            print(f"✅ Success: {result_str[:100]}...")
                            task.steps.append(step)
                            all_action_results.append({
                                "action_number": action_idx + 1,
                                "tool_name": step.tool_name,
                                "success": True,
                                "result": step.result.get("result", "")
                            })

                    # Persist after each step
                    self._persist_task(task)

                # If any action failed, mark remaining as skipped
                if failed_action_index is not None and len(actions_to_execute) > 1:
                    for skip_idx in range(failed_action_index + 1, len(actions_to_execute)):
                        skipped_action = actions_to_execute[skip_idx]
                        all_action_results.append({
                            "action_number": skip_idx + 1,
                            "tool_name": skipped_action.get("tool_name", "unknown"),
                            "status": "skipped",
                            "reason": f"Action {failed_action_index + 1} failed"
                        })
                        print(f"⏭️  [{skip_idx + 1}/{len(actions_to_execute)}] Skipped: {skipped_action.get('tool_name')} (previous action failed)")

                # If we had multiple actions, add summary to context for next iteration
                if len(actions_to_execute) > 1:
                    success_count = sum(1 for r in all_action_results if r.get("success"))
                    failed_count = sum(1 for r in all_action_results if r.get("success") is False)
                    skipped_count = sum(1 for r in all_action_results if r.get("status") == "skipped")

                    print(f"📊 Multi-action summary: {success_count} succeeded, {failed_count} failed, {skipped_count} skipped")

                    # Store results in task history for next iteration
                    # This gets added to conversation history automatically

                # Check for loops - agent repeating same failed action
                if self._detect_loop(task, lookback=3):
                    print(f"⚠️ Loop detected: Agent repeating '{step.tool_name}' without progress")
                    task.status = TaskStatus.NEEDS_CLARIFICATION
                    task.error = f"Agent stuck in loop: repeated '{step.tool_name}' {3} times without progress"
                    break

                # Save iteration snapshot if enabled
                if self.snapshot_iterations:
                    self._save_iteration_snapshot(task.total_iterations, task)

                # Deferred completion: if is_complete was true but we had actions to run first
                if reasoning_result.get("is_complete", False) and has_pending_actions:
                    task.status = TaskStatus.COMPLETED
                    task.final_result = reasoning_result.get("completion_summary", "Task completed")
                    print(f"✅ Task completed (after executing actions): {task.final_result}")
                    self._persist_task(task)
                    break

                # Small delay to prevent overwhelming the system
                time.sleep(0.1)

            # Handle completion
            if task.status == TaskStatus.IN_PROGRESS and task.total_iterations >= task.max_iterations:
                task.status = TaskStatus.FAILED
                task.error = f"Max iterations ({task.max_iterations}) reached"
                # Try to produce a concise summary of work done
                try:
                    summary = self._summarize_task_progress(task)
                    if summary:
                        task.final_result = summary
                except Exception:
                    pass
                print(f"⏰ Task stopped: reached maximum iterations")

            task.end_time = time.time()
            self._persist_task(task)

            # Add Langfuse scoring for task outcomes (v3 API)
            if self.langfuse_enabled and self._current_trace:
                try:
                    # Score 1: Task success
                    success_value = 1.0 if task.status == TaskStatus.COMPLETED else 0.0
                    self._current_trace.score_trace(
                        name="task_success",
                        value=success_value,
                        comment=task.final_result or task.error or "Task completed"
                    )

                    # Score 2: Iteration efficiency
                    if task.total_iterations > 0 and task.max_iterations > 0:
                        efficiency = 1.0 - (task.total_iterations / task.max_iterations)
                        self._current_trace.score_trace(
                            name="iteration_efficiency",
                            value=max(0.0, efficiency),  # Ensure non-negative
                            comment=f"Used {task.total_iterations}/{task.max_iterations} iterations"
                        )

                    # Score 3: Error count
                    error_count = sum(1 for s in task.steps if s.error)
                    if task.total_iterations > 0:
                        error_rate = 1.0 - (error_count / task.total_iterations)
                        self._current_trace.score_trace(
                            name="execution_quality",
                            value=max(0.0, error_rate),
                            comment=f"{error_count} errors in {task.total_iterations} iterations"
                        )

                except Exception as e:
                    print(f"⚠️ Warning: Could not add Langfuse scores: {e}")

            # Finalize Langfuse root span (v3 API)
            if self.langfuse_enabled and self._current_trace:
                try:
                    # Update span with final output and metadata
                    self._current_trace.update(
                        output={
                            "status": task.status.value,
                            "final_result": task.final_result or task.error,
                            "total_steps": len(task.steps)
                        },
                        metadata={
                            "duration_seconds": task.end_time - task.start_time if task.end_time else 0,
                            "total_iterations": task.total_iterations
                        }
                    )
                    # End the span
                    self._current_trace.end()
                except Exception as e:
                    print(f"⚠️ Warning: Could not finalize Langfuse span: {e}")

            # Record in history only if this is a brand new entry or last entry is different task instance
            if not self.execution_history or self.execution_history[-1] is not task:
                self.execution_history.append(task)

        except Exception as e:
            task.status = TaskStatus.FAILED
            task.error = str(e)
            task.end_time = time.time()
            print(f"💥 Task failed with exception: {str(e)}")

            # Add Langfuse scoring for failed task (v3 API)
            if self.langfuse_enabled and self._current_trace:
                try:
                    self._current_trace.score_trace(
                        name="task_success",
                        value=0.0,
                        comment=f"Task failed with exception: {str(e)}"
                    )
                except Exception:
                    pass

            self.execution_history.append(task)

        # Flush Langfuse client to ensure all data is sent
        if self.langfuse_enabled and self.langfuse_client:
            try:
                self.langfuse_client.flush()
            except Exception as e:
                print(f"⚠️ Warning: Could not flush Langfuse client: {e}")

        # Clean up trace reference
        self._current_trace = None

        return task


    def _summarize_task_progress(self, task: TaskExecution) -> Optional[str]:
        """Ask the LLM for a concise summary of the work done and current state."""
        try:
            # OpenAI wrapper will automatically create generation for this summary call
            messages = [
                {"role": "system", "content": (
                    "You summarize Excel automation progress. Be concise, actionable, and specific. "
                    "Do NOT propose further actions; only summarize what was accomplished, current sheet(s) created/edited, key ranges touched, and any remaining gaps."
                )},
                {"role": "user", "content": self._get_context_prompt(task) + "\n\nReturn only plain text."},
            ]
            request_data = {
                "model": self.model,
                "messages": messages,
                "max_completion_tokens": 400
            }

            # Add Langfuse metadata if enabled
            if self.langfuse_enabled:
                request_data["extra_body"] = {"metadata": {
                    "task_id": task.task_id,
                    "iteration": task.total_iterations,
                    "purpose": "task_summary",
                    "status": task.status.value
                }}

            resp = self.openai_client.chat.completions.create(**request_data)
            self._log_openai_request(request_data, resp, task_id=task.task_id, iteration=task.total_iterations)

            return resp.choices[0].message.content
        except Exception:
            return None

    def get_execution_summary(self, task: TaskExecution, show_steps: bool = True) -> str:
        """Generate a human-readable execution summary"""
        duration = (task.end_time or time.time()) - task.start_time

        summary = f"""
        === TASK EXECUTION SUMMARY ===
        Task: "{task.user_prompt}"
        Status: {task.status.value.upper()}
        Duration: {duration:.2f} seconds
        Iterations: {task.total_iterations}
        Steps completed: {len(task.steps)}
        """

        if task.final_result:
            summary += f"\nResult: {task.final_result}"

        # Collect all errors with iteration IDs
        errors_by_iteration = []
        for step in task.steps:
            if step.error:
                errors_by_iteration.append((step.step_number, step.error))

        if errors_by_iteration:
            summary += f"\n\nErrors encountered ({len(errors_by_iteration)}):"
            for step_num, error in errors_by_iteration:
                summary += f"\n  - Iteration {step_num}: {error}"

        if task.error and task.error not in [e[1] for e in errors_by_iteration]:
            # Show task-level error if it's not already in step errors
            summary += f"\n\nTask Error: {task.error}"

        if show_steps:
            summary += "\n\nEXECUTION STEPS:\n"
            for step in task.steps:
                status_icon = "✅" if not step.error else "❌"
                summary += f"{status_icon} Step {step.step_number}: {step.description}\n"
                summary += f"   Tool: {step.tool_name}\n"
                if step.error:
                    summary += f"   Error: {step.error}\n"
                elif step.result and step.result.get('success'):
                    result_preview = str(step.result.get('result', ''))[:100]
                    summary += f"   Result: {result_preview}...\n"
                summary += "\n"

        return summary



# Example usage and testing
def test_task_execution():
    """Test the task execution engine"""
    from .mcp_client import ExcelMCPClient

    # Initialize clients
    server_path = "excel_mcp_server/server.py"
    excel_client = ExcelMCPClient(server_path, "./test_excel_files")

    # You would need to provide your OpenAI API key here
    api_key = "your-openai-api-key"  # Replace with actual key

    executor = ExcelTaskExecutor(excel_client, api_key)

    # Test task
    test_prompt = """
    Create a sales analysis spreadsheet called 'sales_analysis.xlsx' with the following:
    1. A 'Sales Data' worksheet with columns: Product, Units Sold, Unit Price, Total Revenue
    2. Add sample data for 5 products
    3. Include formulas to calculate Total Revenue (Units Sold * Unit Price)
    4. Create a 'Summary' worksheet with total revenue across all products
    5. Verify all calculations are correct
    """

    print("Starting task execution test...")
    task = executor.execute_task(test_prompt)

    print("\n" + executor.get_execution_summary(task))


if __name__ == "__main__":
    test_task_execution()
