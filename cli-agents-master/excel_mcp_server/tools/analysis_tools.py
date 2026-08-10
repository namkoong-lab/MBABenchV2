"""Analysis tools: summarize_workbook_context, describe_worksheet, scan_worksheet_structure."""
import json
import os
from typing import Any, Dict, List, Optional

import openpyxl

from ..core.shared_state import mcp, _get_openai_client
from ..core.workbook_io import _load_workbook
from ..helpers.structure_analysis import (
    _actual_used_range,
    _contiguous_blocks,
    _iter_sheet_cells,
    _find_task_like_text,
    _find_text_sheet_references,
    _tokenize_words,
    _collect_sheet_keywords,
    _extract_references_from_formula,
    _detect_header_row,
    _summarize_columns,
)


@mcp.tool()
async def summarize_workbook_context(
    filename: str,
    row_cap: int = 2000,
    col_cap: int = 120,
    max_cells: int = 150_000,
) -> str:
    """Summarize workbook context: per-sheet used ranges, blocks, headers, task-like lines, and references.

    This is a non-LLM structural pass suitable for large files. It avoids dumping raw cell values.

    Args:
        filename: Excel file name ('.xlsx' added if missing)
        row_cap: Rows to inspect for task-like text per sheet
        col_cap: Columns to inspect for task-like text per sheet
        max_cells: If used range exceeds this, only metadata + limited scans are performed

    Returns:
        JSON string with workbook, sheets, blocks, tasks, and references.
    """
    try:
        wb = _load_workbook(filename)
        out: Dict[str, Any] = {"filename": filename, "sheets": [], "notes": {"row_cap": row_cap, "col_cap": col_cap, "max_cells": max_cells}}
        sheet_names = [ws.title for ws in wb.worksheets]
        for ws in wb.worksheets:
            name = ws.title
            bounds = _actual_used_range(ws)
            if not bounds:
                out["sheets"].append({"name": name, "empty": True})
                continue
            min_r, min_c, max_r, max_c = bounds
            rows = max_r - min_r + 1
            cols = max_c - min_c + 1
            total_cells = rows * cols
            used_range = f"{openpyxl.utils.get_column_letter(min_c)}{min_r}:{openpyxl.utils.get_column_letter(max_c)}{max_r}"

            blocks = _contiguous_blocks(ws, bounds)
            blocks_out: List[Dict[str, Any]] = []
            for b in blocks[:50]:
                header_row = _detect_header_row(ws, b)
                cols_summary = []
                if header_row is not None:
                    cols_summary = _summarize_columns(ws, header_row, b["left"], b["right"], b["bottom"])[:20]
                blocks_out.append({
                    "range": b["range"],
                    "top": b["top"], "left": b["left"], "bottom": b["bottom"], "right": b["right"],
                    "header_row": header_row,
                    "columns": [{"name": (c.get("name") if isinstance(c.get("name"), (str, int, float)) else str(c.get("name"))), "column_index": c.get("column_index"), "type": c.get("inferred_type")} for c in cols_summary],
                })

            tasks = _find_task_like_text(ws, bounds, row_cap=row_cap, col_cap=col_cap)

            refs: List[Dict[str, Any]] = []
            formulas_count = 0
            ref_scan_limited = False
            if total_cells <= max_cells:
                for r, c, cell in _iter_sheet_cells(ws, bounds):
                    try:
                        val = cell.value
                    except Exception:
                        val = None
                    if isinstance(val, str) and val.startswith('='):
                        formulas_count += 1
                        for ref in _extract_references_from_formula(val, name):
                            refs.append({"from": f"{openpyxl.utils.get_column_letter(c)}{r}", **ref, "kind": "formula"})
            else:
                ref_scan_limited = True
                sample_rows = list(range(min_r, min(min_r + 200, max_r + 1)))
                for r in sample_rows:
                    for c in range(min_c, min(max_c, min_c + col_cap - 1) + 1):
                        val = ws.cell(row=r, column=c).value
                        if isinstance(val, str) and val.startswith('='):
                            formulas_count += 1
                            for ref in _extract_references_from_formula(val, name):
                                refs.append({"from": f"{openpyxl.utils.get_column_letter(c)}{r}", **ref, "kind": "formula"})

            text_refs = _find_text_sheet_references(ws, bounds, sheet_names, name, row_cap=row_cap, col_cap=col_cap)
            for tr in text_refs:
                refs.append({
                    "from": tr.get("from"),
                    "ref_sheet": tr.get("to_sheet"),
                    "ref_range": None,
                    "kind": "text",
                    "context": tr.get("context"),
                })

            seen = set()
            uniq_refs: List[Dict[str, Any]] = []
            for r in refs:
                key = (r.get("from"), r.get("ref_sheet"), r.get("ref_range"))
                if key in seen:
                    continue
                seen.add(key)
                uniq_refs.append(r)

            try:
                sheet_keywords = _collect_sheet_keywords(ws, bounds, blocks_out)
                other_info: Dict[str, Dict[str, Any]] = {}
                deny = {"tasks and questions", "tasks", "questions"}
                for other_name in sheet_names:
                    if other_name == name:
                        continue
                    if other_name.lower() in deny:
                        continue
                    try:
                        ows = wb[other_name]
                        obounds = _actual_used_range(ows)
                        if not obounds:
                            continue
                        oblocks = _contiguous_blocks(ows, obounds)
                        oblocks_out: List[Dict[str, Any]] = []
                        for ob in oblocks[:10]:
                            hrow = _detect_header_row(ows, ob)
                            cols_summary = []
                            if hrow is not None:
                                cols_summary = _summarize_columns(ows, hrow, ob["left"], ob["right"], ob["bottom"])[:10]
                            oblocks_out.append({
                                "range": f"{openpyxl.utils.get_column_letter(ob['left'])}{ob['top']}:{openpyxl.utils.get_column_letter(ob['right'])}{ob['bottom']}",
                                "columns": cols_summary,
                            })
                        other_info[other_name] = {
                            "keywords": _collect_sheet_keywords(ows, obounds, oblocks_out),
                            "blocks": oblocks_out,
                        }
                    except Exception:
                        continue
                for t in tasks:
                    t_tokens = set(_tokenize_words(t.get("text", "")))
                    candidates = []
                    for oname, oinfo in other_info.items():
                        okws = set(oinfo.get("keywords", []))
                        overlap = t_tokens.intersection(okws)
                        if overlap:
                            cand = {"sheet": oname, "score": len(overlap), "matched": sorted(list(overlap))[:10]}
                            best_block = None
                            best_hits = 0
                            for b in oinfo.get("blocks", [])[:10]:
                                headers = [str(c.get("name")) for c in (b.get("columns") or []) if c.get("name") is not None]
                                h_toks = set()
                                for h in headers:
                                    h_toks.update(_tokenize_words(h))
                                hits = len(t_tokens.intersection(h_toks))
                                if hits > best_hits:
                                    best_hits = hits
                                    best_block = (b.get("range"), [h for h in headers if h and any(tok in _tokenize_words(h) for tok in t_tokens)])
                            if best_block:
                                cand["range"] = best_block[0]
                                cand["header_matches"] = best_block[1][:6]
                            candidates.append(cand)
                    try:
                        self_blocks = []
                        for b in blocks_out[:20]:
                            headers = [str(c.get("name")) for c in (b.get("columns") or []) if c.get("name") is not None]
                            h_toks = set()
                            for h in headers:
                                h_toks.update(_tokenize_words(h))
                            hits = len(t_tokens.intersection(h_toks))
                            if hits:
                                cand = {
                                    "sheet": name,
                                    "score": hits,
                                    "range": b.get("range"),
                                    "header_matches": [h for h in headers if h and any(tok in _tokenize_words(h) for tok in t_tokens)][:6],
                                }
                                self_blocks.append(cand)
                        self_blocks.sort(key=lambda x: -x["score"])
                        if self_blocks:
                            candidates.append(self_blocks[0])
                    except Exception:
                        pass
                    candidates.sort(key=lambda x: -x["score"])
                    if candidates:
                        t["candidates"] = candidates[:5]
            except Exception:
                pass

            sheet_out = {
                "name": name,
                "bounds": {"min_row": min_r, "min_col": min_c, "max_row": max_r, "max_col": max_c},
                "used_range": used_range,
                "total_cells": total_cells,
                "blocks": blocks_out,
                "tasks": tasks,
                "references": uniq_refs[:2000],
                "formulas_count": formulas_count,
                "ref_scan_limited": ref_scan_limited,
            }
            out["sheets"].append(sheet_out)

        cross = []
        try:
            for s in out["sheets"]:
                for r in s.get("references", []):
                    if r.get("ref_sheet") and r.get("ref_sheet") != s["name"]:
                        cross.append({
                            "from_sheet": s["name"],
                            "from_cell": r.get("from"),
                            "to_sheet": r.get("ref_sheet"),
                            "to_range": r.get("ref_range"),
                            "kind": r.get("kind"),
                            "context": r.get("context"),
                        })
        except Exception:
            pass
        out["cross_sheet_references"] = cross[:5000]

        return json.dumps(out, indent=2, default=str)
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


@mcp.tool()
async def describe_worksheet(
    filename: str,
    worksheet_name: str,
    focus: Optional[str] = None,
    row_cap: int = 2000,
    col_cap: int = 100,
    model: Optional[str] = None,
) -> str:
    """Use an LLM to describe the contents and structure of a worksheet with precise ranges.

    Args:
        filename: Excel file name
        worksheet_name: Worksheet to analyze
        focus: Optional extra query to emphasize particular information needs
        row_cap: Max rows to scan (for snapshot)
        col_cap: Max columns to scan (for snapshot)
        model: Optional OpenAI model override (defaults to env or client default)

    Returns:
        JSON string with keys: summary, findings, entities, text_spans, tables, ranges_index, notes
    """
    try:
        wb = _load_workbook(filename)
        if worksheet_name not in wb.sheetnames:
            return json.dumps({"success": False, "error": f"Worksheet '{worksheet_name}' not found"})
        ws = wb[worksheet_name]

        bounds = _actual_used_range(ws, max_rows=row_cap, max_cols=col_cap)
        if not bounds:
            return json.dumps({
                "success": True,
                "worksheet": worksheet_name,
                "summary": "Worksheet is empty within scan caps.",
                "findings": [],
                "tables": [],
                "entities": [],
                "text_spans": [],
                "ranges_index": {},
                "notes": {"row_cap": row_cap, "col_cap": col_cap},
            }, indent=2)

        min_r, min_c, max_r, max_c = bounds

        snapshot_rows = []
        for r in range(min_r, min(max_r, min_r + 200) + 1):
            row_vals = []
            empty = True
            for c in range(min_c, min(max_c, min_c + col_cap - 1) + 1):
                v = ws.cell(row=r, column=c).value
                if v not in (None, ""):
                    empty = False
                sv = v
                if isinstance(v, str):
                    sv = v.strip()
                    if len(sv) > 80:
                        sv = sv[:77] + "..."
                row_vals.append(sv)
            if not empty:
                snapshot_rows.append({"row": r, "values": row_vals})

        client = _get_openai_client()
        if client is None:
            return json.dumps({
                "success": True,
                "worksheet": worksheet_name,
                "summary": "LLM unavailable (no OPENAI_API_KEY). Returning structural snapshot only.",
                "bounds": {"min_row": min_r, "min_col": min_c, "max_row": max_r, "max_col": max_c},
                "snapshot": snapshot_rows[:200],
                "notes": {"row_cap": row_cap, "col_cap": col_cap},
            }, indent=2)

        sys_prompt = (
            "You analyze Excel worksheets and return STRICT JSON. Identify tables, named sections, "
            "text blocks, headers, and meaningful entities. Always include exact cell ranges. "
            "Do not omit other findings when focusing on a query; prioritize them but keep others."
        )
        user_payload = {
            "worksheet": worksheet_name,
            "bounds": {"min_row": min_r, "min_col": min_c, "max_row": max_r, "max_col": max_c},
            "focus": focus or "",
            "snapshot_rows": snapshot_rows,
            "required_schema": {
                "summary": "string",
                "findings": ["short bullet points"],
                "tables": [
                    {
                        "name": "string",
                        "range": "A1:D20",
                        "header_row": 1,
                        "columns": [{"name": "string", "column_index": 1}]
                    }
                ],
                "entities": [
                    {"type": "label|value|section|note|other", "text": "string", "range": "A10"}
                ],
                "text_spans": [
                    {"kind": "paragraph|note|title", "range": "A2:B5", "preview": "string"}
                ],
                "ranges_index": {"logical_name": "A1:C14"},
                "notes": {"caps": {"row_cap": row_cap, "col_cap": col_cap}},
            },
            "constraints": [
                "Return valid JSON only (no markdown)",
                "Prefer compact ranges over verbose dumps",
                "Use absolute ranges like 'A1:D20' and single cells like 'B3'",
                "Include non-focus findings briefly",
            ],
        }

        try:
            resp = client.chat.completions.create(
                model=model or os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
                messages=[
                    {"role": "system", "content": sys_prompt},
                    {"role": "user", "content": json.dumps(user_payload)},
                ],
                max_completion_tokens=800,
                response_format={"type": "json_object"},
            )
        except Exception:
            resp = client.chat.completions.create(
                model=model or os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
                messages=[
                    {"role": "system", "content": sys_prompt},
                    {"role": "user", "content": json.dumps(user_payload)},
                ],
                max_completion_tokens=800,
            )

        content = resp.choices[0].message.content or "{}"
        try:
            parsed = json.loads(content)
        except Exception:
            return json.dumps({
                "success": True,
                "worksheet": worksheet_name,
                "raw": content,
                "note": "Non-JSON response; client should sanitize.",
            }, indent=2)

        parsed["success"] = True
        parsed["worksheet"] = worksheet_name
        return json.dumps(parsed, indent=2)
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


@mcp.tool()
async def scan_worksheet_structure(filename: str, worksheet_name: str, row_cap: int = 2000, col_cap: int = 100) -> str:
    """Scan a worksheet and return a compact structural summary without all raw cells.

    Finds contiguous non-empty blocks, detects header rows, and summarizes
    column types and sample values. Useful for planning targeted reads.

    Args:
        filename: Excel file name ('.xlsx' added if missing)
        worksheet_name: Worksheet to scan
        row_cap: Maximum rows to scan (default: 2000)
        col_cap: Maximum columns to scan (default: 100)

    Returns:
        JSON string with {filename, worksheet, empty, used_range, bounds,
        block_count, blocks} where each block has range, header info, and columns
    """
    try:
        wb = _load_workbook(filename)
        ws = wb[worksheet_name]
        bounds = _actual_used_range(ws, max_rows=row_cap, max_cols=col_cap)
        if not bounds:
            return json.dumps({
                "filename": filename,
                "worksheet": worksheet_name,
                "empty": True,
                "used_range": None,
                "blocks": []
            }, indent=2)
        r1, c1, r2, c2 = bounds
        used_range = f"{openpyxl.utils.get_column_letter(c1)}{r1}:{openpyxl.utils.get_column_letter(c2)}{r2}"
        blocks = _contiguous_blocks(ws, bounds)
        out_blocks: List[Dict[str, Any]] = []
        for b in blocks:
            header_row = _detect_header_row(ws, b)
            entry: Dict[str, Any] = {
                "range": b["range"],
                "top": b["top"],
                "left": b["left"],
                "bottom": b["bottom"],
                "right": b["right"],
                "has_header": header_row is not None,
                "header_row": header_row,
            }
            if header_row is not None:
                entry["columns"] = _summarize_columns(ws, header_row, b["left"], b["right"], b["bottom"])
            out_blocks.append(entry)
        summary = {
            "filename": filename,
            "worksheet": worksheet_name,
            "empty": False,
            "used_range": used_range,
            "bounds": {"top": r1, "left": c1, "bottom": r2, "right": c2},
            "block_count": len(out_blocks),
            "blocks": out_blocks,
        }
        return json.dumps(summary, indent=2, default=str)
    except Exception as e:
        return f"Error scanning worksheet: {str(e)}"
