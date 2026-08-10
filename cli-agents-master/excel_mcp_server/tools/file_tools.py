"""File management tools: create, list, metadata, copy."""
import json
from typing import List, Optional

import openpyxl

from ..core.shared_state import mcp
from ..core import shared_state as _state
from ..core.workbook_io import _get_file_path, _save_workbook_sync, _load_workbook


@mcp.tool()
async def create_file(filename: str, worksheets: Optional[List[str]] = None) -> str:
    """Create a new Excel file with optional worksheets.

    WARNING: This will OVERWRITE an existing file with the same name!
    Check with list_files() first to avoid data loss.

    Args:
        filename: Name of the Excel file to create (will add .xlsx if missing)
        worksheets: List of worksheet names to create (defaults to ["Sheet1"])

    Returns:
        Success message with created worksheets, or error if file exists
    """
    try:
        file_path = _get_file_path(filename)

        if file_path.exists():
            file_size = file_path.stat().st_size
            return (
                f"ERROR: File '{filename}' already exists!\n\n"
                f"create_file would DESTROY the existing file and all its data.\n\n"
                f"To modify the file, use:\n"
                f"   - set_cell_formula() to add formulas\n"
                f"   - edit_cells() to add values\n"
                f"   - create_worksheet() to add new sheets\n\n"
                f"If you really need to replace the file, first delete it with delete_file().\n"
                f"Current file size: {file_size} bytes"
            )

        if worksheets is None:
            worksheets = ["Sheet1"]

        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        for ws_name in worksheets:
            wb.create_sheet(ws_name)

        _save_workbook_sync(wb, file_path)

        return f"Excel file '{filename}' created successfully with worksheets: {worksheets}"
    except Exception as e:
        return f"Error creating file: {str(e)}"


@mcp.tool()
async def list_files() -> str:
    """List all Excel files in storage directory.

    Returns:
        JSON string with list of file information (filename, size, modified time)
    """
    try:
        excel_files = []
        for file_path in _state.STORAGE_PATH.glob("*.xlsx"):
            size = file_path.stat().st_size
            modified = file_path.stat().st_mtime
            excel_files.append({
                "filename": file_path.name,
                "size_bytes": size,
                "modified_timestamp": modified
            })

        return json.dumps(excel_files, indent=2)
    except Exception as e:
        return f"Error listing files: {str(e)}"


@mcp.tool()
async def get_file_metadata(filename: str) -> str:
    """Get metadata for an Excel file.

    Args:
        filename: Name of the Excel file

    Returns:
        JSON string with file metadata (worksheets, count, active sheet)
    """
    try:
        wb = _load_workbook(filename)
        metadata = {
            "filename": filename,
            "worksheets": wb.sheetnames,
            "worksheet_count": len(wb.sheetnames),
            "active_sheet": wb.active.title if wb.active else None
        }

        return json.dumps(metadata, indent=2)
    except Exception as e:
        return f"Error getting metadata: {str(e)}"


@mcp.tool()
async def copy_file(source_filename: str, destination_filename: str) -> str:
    """Copy an Excel file from one location to another.

    Args:
        source_filename: Name of the source Excel file (can be absolute path or relative to storage)
        destination_filename: Name of the destination Excel file (relative to storage)

    Returns:
        Success message or error description
    """
    try:
        from pathlib import Path
        source_path = Path(source_filename)
        if not source_path.is_absolute():
            source_path = _get_file_path(source_filename)

        if not source_path.exists():
            return f"Error: Source file '{source_filename}' not found"

        dest_path = _get_file_path(destination_filename)
        dest_path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.load_workbook(source_path)
        _save_workbook_sync(wb, dest_path)

        return f"Successfully copied '{source_filename}' to '{destination_filename}'"
    except Exception as e:
        return f"Error copying file: {str(e)}"
