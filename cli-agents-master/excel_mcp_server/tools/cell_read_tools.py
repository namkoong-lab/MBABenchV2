"""Cell read tools: get_cell_range, get_formula, get_used_range, search_worksheet."""
import json
import re
from typing import Any, Dict, List, Optional

import openpyxl

from ..core.shared_state import mcp
from ..core.workbook_io import _load_workbook, _load_workbook_view
from ..helpers.formula_evaluation import _eval_formula
from ..helpers.structure_analysis import _actual_used_range


@mcp.tool()
async def get_cell_range(filename: str, worksheet_name: str, range_address: str, mode: str = "default") -> str:
    """Get values from a cell range with optional formula/value pairing.

    Args:
        filename: Name of the Excel file
        worksheet_name: Name of the worksheet
        range_address: Cell range address (e.g., 'A1:C5' or 'A1' for single cell)
        mode: 'default' returns paired raw formulas + calculated values,
              'raw' returns only raw cell values/formulas without evaluation

    Returns:
        JSON string with range data: {range, values} in raw mode,
        or {range, values, view_values} in default mode
    """
    try:
        wb_raw = _load_workbook(filename)
        ws_raw = wb_raw[worksheet_name]
        use_raw_only = (str(mode).lower() == "raw")
        if not use_raw_only:
            wb_view = _load_workbook_view(filename)
            ws_view = wb_view[worksheet_name]

        target = ws_raw[range_address]

        def to_raw(cell_or_range):
            if hasattr(cell_or_range, 'value'):
                return [[cell_or_range.value]]
            out = []
            for row in ws_raw[range_address]:
                out.append([c.value for c in row])
            return out

        if use_raw_only:
            raw_values = to_raw(target)
            result = {"range": range_address, "values": raw_values}
        else:
            def to_pairs(cell_or_range):
                pairs = []
                if hasattr(cell_or_range, 'value'):
                    raw_val = cell_or_range.value
                    view_val = ws_view[cell_or_range.coordinate].value
                    if view_val is None and isinstance(raw_val, str) and raw_val.startswith('='):
                        try:
                            view_val = _eval_formula(raw_val, wb_raw, ws_raw, {})
                        except Exception:
                            view_val = None
                    pairs = [[(raw_val, view_val)]]
                else:
                    for row in ws_raw[range_address]:
                        row_pairs = []
                        for c in row:
                            raw_val = c.value
                            v = ws_view[c.coordinate].value
                            if v is None and isinstance(raw_val, str) and raw_val.startswith('='):
                                try:
                                    v = _eval_formula(raw_val, wb_raw, ws_raw, {})
                                except Exception:
                                    v = None
                            row_pairs.append((raw_val, v))
                        pairs.append(row_pairs)
                return pairs

            pairs = to_pairs(target)
            raw_values = [[rv for (rv, vv) in row] for row in pairs]
            view_values = [[vv for (rv, vv) in row] for row in pairs]
            result = {"range": range_address, "values": raw_values, "view_values": view_values}
        return json.dumps(result, indent=2, default=str)
    except Exception as e:
        return f"Error getting cell range: {str(e)}"


@mcp.tool()
async def get_formula(filename: str, worksheet_name: str, cell_address: str) -> str:
    """Get formula from a specific cell.

    Args:
        filename: Name of the Excel file
        worksheet_name: Name of the worksheet
        cell_address: Cell address (e.g., 'A1')

    Returns:
        JSON string with cell formula and value information
    """
    try:
        wb = _load_workbook(filename)
        ws = wb[worksheet_name]
        cell = ws[cell_address]

        formula = None
        try:
            if hasattr(cell, 'formula') and cell.formula:
                formula = cell.formula
            elif isinstance(cell.value, str) and cell.value.startswith('='):
                formula = cell.value
        except:
            pass

        result = {
            "cell": cell_address,
            "formula": formula,
            "value": cell.value,
            "data_type": str(cell.data_type) if hasattr(cell, 'data_type') else "unknown"
        }

        return json.dumps(result, indent=2, default=str)
    except Exception as e:
        return f"Error getting formula: {str(e)}"


@mcp.tool()
async def get_used_range(filename: str, worksheet_name: str, max_rows: Optional[int] = None, max_cols: Optional[int] = None) -> str:
    """Return the actual used range (non-empty bounds) of a worksheet.

    Args:
        filename: Excel file name
        worksheet_name: Worksheet name
        max_rows: Optional cap on rows to scan
        max_cols: Optional cap on columns to scan

    Returns:
        JSON with keys: used_range (e.g., 'A1:D25'), bounds {top,left,bottom,right}
    """
    try:
        wb = _load_workbook(filename)
        ws = wb[worksheet_name]
        bounds = _actual_used_range(ws, max_rows=max_rows, max_cols=max_cols)
        if not bounds:
            return json.dumps({"used_range": None, "bounds": None, "empty": True})
        r1, c1, r2, c2 = bounds
        used_range = f"{openpyxl.utils.get_column_letter(c1)}{r1}:{openpyxl.utils.get_column_letter(c2)}{r2}"
        return json.dumps({
            "used_range": used_range,
            "bounds": {"top": r1, "left": c1, "bottom": r2, "right": c2},
            "empty": False,
        }, indent=2)
    except Exception as e:
        return f"Error computing used range: {str(e)}"


@mcp.tool()
async def search_worksheet(filename: str, worksheet_name: str, query: str, match_type: str = "substring", case_sensitive: bool = False, max_results: int = 200) -> str:
    """Search for text in a worksheet and return matching cells.

    Args:
        filename: Excel file name
        worksheet_name: Worksheet name
        query: Text or regex pattern (if match_type='regex')
        match_type: 'substring' | 'equals' | 'regex'
        case_sensitive: Whether to match case sensitively
        max_results: Maximum number of matches to return

    Returns:
        JSON with matches: [{cell, row, column, value}] (truncated to max_results)
    """
    try:
        wb = _load_workbook(filename)
        ws = wb[worksheet_name]
        flags = 0 if case_sensitive else re.IGNORECASE
        pattern = None
        if match_type == 'regex':
            try:
                pattern = re.compile(query, flags)
            except re.error as e:
                return json.dumps({"error": f"Invalid regex: {e}"})

        matches: List[Dict[str, Any]] = []
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str):
                    continue
                text = v if case_sensitive else v.lower()
                q = query if case_sensitive else query.lower()
                ok = False
                if match_type == 'equals':
                    ok = text == q
                elif match_type == 'substring':
                    ok = q in text
                elif match_type == 'regex':
                    ok = bool(pattern.search(v)) if pattern else False
                if ok:
                    matches.append({
                        "cell": cell.coordinate,
                        "row": cell.row,
                        "column": cell.column,
                        "value": v,
                    })
                    if len(matches) >= max_results:
                        break
            if len(matches) >= max_results:
                break
        return json.dumps({"query": query, "count": len(matches), "matches": matches}, indent=2)
    except Exception as e:
        return f"Error searching worksheet: {str(e)}"
