"""Formatting tools: format_cells, freeze_panes."""
import json
from typing import Any, Dict, Optional

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ..core.libreoffice_bridge import _save_with_recalc
from ..core.shared_state import mcp
from ..core.workbook_io import _load_workbook


@mcp.tool()
async def freeze_panes(filename: str, worksheet_name: str, cell: str) -> str:
    """Freeze rows and columns at a cell position so they stay visible when scrolling.

    Freezing at "B2" keeps row 1 (headers) and column A (labels) pinned.

    Args:
        filename: Name of the Excel file
        worksheet_name: Name of the worksheet
        cell: Cell position to freeze at (e.g., "B2" freezes row 1 and column A)

    Returns:
        Success or error message
    """
    try:
        wb = _load_workbook(filename)
        ws = wb[worksheet_name]
        ws.freeze_panes = cell
        recalc_engine_info = _save_with_recalc(wb, filename)
        return json.dumps({"success": True, "frozen_at": cell, "worksheet": worksheet_name,
                           "recalc_engine": recalc_engine_info})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


@mcp.tool()
async def format_cells(
    filename: str,
    worksheet_name: str,
    range_address: str,
    font: Optional[Dict[str, Any]] = None,
    fill: Optional[Dict[str, Any]] = None,
    border: Optional[Dict[str, Any]] = None,
    alignment: Optional[Dict[str, Any]] = None,
    number_format: Optional[str] = None,
) -> str:
    """Apply formatting (font, fill, border, alignment, number_format) to a cell range.

    Apply ONLY after ALL calculations are verified and answer sheets linked.

    Args:
        filename: Name of the Excel file
        worksheet_name: Name of the worksheet
        range_address: Cell range (e.g., "A1:D1" for header row, "B2:B20" for data column)
        font: Font properties -- {"color": "0000FF", "bold": true, "italic": false, "size": 11, "name": "Calibri"}
        fill: Fill properties -- {"color": "002060"} (solid fill with hex color)
        border: Border properties -- {"style": "thin"} applies to all sides
        alignment: Alignment -- {"horizontal": "center", "vertical": "center", "wrap_text": true}
        number_format: Excel number format string -- "#,##0", "0.00%", "$#,##0_);($#,##0)", "0.00"

    Returns:
        Success message with count of formatted cells
    """
    try:
        wb = _load_workbook(filename)
        ws = wb[worksheet_name]

        font_obj = None
        if font:
            font_obj = Font(
                color=font.get("color"),
                bold=font.get("bold", False),
                italic=font.get("italic", False),
                size=font.get("size"),
                name=font.get("name"),
            )

        fill_obj = None
        if fill:
            fill_obj = PatternFill(
                start_color=fill.get("color", "FFFFFF"),
                end_color=fill.get("color", "FFFFFF"),
                fill_type="solid",
            )

        border_obj = None
        if border:
            side = Side(style=border.get("style", "thin"))
            border_obj = Border(top=side, bottom=side, left=side, right=side)

        align_obj = None
        if alignment:
            align_obj = Alignment(
                horizontal=alignment.get("horizontal"),
                vertical=alignment.get("vertical"),
                wrap_text=alignment.get("wrap_text", False),
            )

        cell_count = 0
        target = ws[range_address]
        from openpyxl.cell.cell import Cell as _Cell
        if isinstance(target, _Cell):
            rows = ((target,),)
        elif isinstance(target, tuple) and target and not isinstance(target[0], tuple):
            rows = (target,)
        else:
            rows = target
        for row in rows:
            for c in row:
                if font_obj:
                    c.font = font_obj
                if fill_obj:
                    c.fill = fill_obj
                if border_obj:
                    c.border = border_obj
                if align_obj:
                    c.alignment = align_obj
                if number_format:
                    c.number_format = number_format
                cell_count += 1

        recalc_engine_info = _save_with_recalc(wb, filename)
        return json.dumps({"success": True, "cells_formatted": cell_count, "range": range_address,
                           "worksheet": worksheet_name, "recalc_engine": recalc_engine_info})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})
