"""
LibreOffice Headless Integration for Excel MCP Server

Recalculates workbook formulas with LibreOffice Calc's engine, which supports
ALL Excel functions including PMT, IRR, VLOOKUP, INDEX/MATCH, etc.

Each recalculation is one self-contained subprocess:

    soffice --headless --calc --convert-to xlsx --outdir <tmp> <file>

An xlsx->xlsx conversion re-saves the workbook through Calc, performing a full
recalculation and embedding cached values (the same filter the old UNO path's
store() used, and the same invocation the judge's --run-calculation uses).
The converted file then replaces the original in place, so
openpyxl.load_workbook(data_only=True) sees computed values.

This replaced the UNO socket bridge (persistent soffice listener + a second
UNO-capable python) in 2026-08: `import uno` only exists in an interpreter
shipped with LibreOffice, and on macOS Apple's launch constraint on that
interpreter made the architecture unusable — runs silently degraded to the
_eval_formula fallback. See plan/libreoffice_recalc.md (decision) and
plan/libreoffice_recalc_b1_impl.md (this design).
"""

import logging
import os
import shutil
import subprocess
import sys
import tempfile
import threading
import time
from pathlib import Path
from typing import Optional, Tuple

logger = logging.getLogger(__name__)

# macOS default install location; `soffice` is not on PATH there by default.
_MACOS_SOFFICE = "/Applications/LibreOffice.app/Contents/MacOS/soffice"


def _repo_libreoffice_path() -> Optional[str]:
    """`libreoffice_path` from the monorepo config, or None (standalone checkout)."""
    try:
        from excel_cli_agent.repo_config import repo_value
    except ImportError:
        return None
    return repo_value("libreoffice_path")


def resolve_soffice() -> Optional[Tuple[str, str]]:
    """Locate the soffice binary as (path, source), or None.

    Resolution order:
      1. LIBREOFFICE_PATH env var (explicit override)
      2. `libreoffice_path` in <MBABenchV2>/config/config.yaml
      3. `soffice` on PATH (Linux: apt-get install libreoffice-calc)
      4. macOS default app bundle location

    An explicitly configured path (1 or 2) is returned as-is even if it does
    not exist — the caller reports it as a config error rather than silently
    falling through to a different binary.
    """
    env_path = os.environ.get("LIBREOFFICE_PATH")
    if env_path:
        return env_path, "LIBREOFFICE_PATH env var"
    repo_path = _repo_libreoffice_path()
    if repo_path:
        return repo_path, "libreoffice_path in <MBABenchV2>/config/config.yaml"
    on_path = shutil.which("soffice")
    if on_path:
        return on_path, "soffice on PATH"
    if sys.platform == "darwin" and os.path.exists(_MACOS_SOFFICE):
        return _MACOS_SOFFICE, "macOS default app location"
    return None


class LibreOfficeCalcEngine:
    """Recalculates xlsx files by shelling out to `soffice --convert-to`.

    One conversion subprocess per recalculate() call. A persistent
    -env:UserInstallation profile directory — one per engine instance, i.e.
    per MCP server process — keeps every launch after the first warm, and
    isolates concurrent batches from each other (soffice locks its profile,
    so two processes must never share one).
    """

    def __init__(self, recalc_timeout: float = 60.0):
        self.recalc_timeout = recalc_timeout
        self.soffice_path: Optional[str] = None
        self.soffice_source: Optional[str] = None
        self.soffice_version: Optional[str] = None
        self._profile_dir: Optional[str] = None
        self._started = False
        self._lock = threading.Lock()

    def start(self):
        """Resolve the binary, build the profile, and prove conversion works.

        Raises RuntimeError when no working LibreOffice is available. The
        caller decides whether that is fatal (server.py: fatal unless
        --no-libreoffice / --allow-recalc-fallback).
        """
        resolved = resolve_soffice()
        if resolved is None:
            raise RuntimeError(
                "No LibreOffice binary found. Install LibreOffice "
                "(Linux: apt-get install libreoffice-calc; macOS: install "
                "LibreOffice.app), or point libreoffice_path in "
                "<MBABenchV2>/config/config.yaml (or LIBREOFFICE_PATH) at the "
                "soffice binary."
            )
        path, source = resolved
        if not (os.path.isfile(path) and os.access(path, os.X_OK)):
            raise RuntimeError(
                f"LibreOffice binary from {source} is not executable: {path}"
            )
        self.soffice_path = path
        self.soffice_source = source

        self._profile_dir = tempfile.mkdtemp(prefix="lo_profile_")
        self.soffice_version = self._read_version()

        # Warmup conversion: pays the one-time profile bootstrap before the
        # agent's first write, and proves headless conversion actually
        # produces cached values (a --version probe would not catch a broken
        # install).
        self._warmup()

        self._started = True
        logger.info(
            f"LibreOffice recalc engine ready: {path} "
            f"({self.soffice_version or 'version unknown'})"
        )

    def stop(self):
        """Release the profile directory. Call at MCP server shutdown."""
        if self._profile_dir:
            shutil.rmtree(self._profile_dir, ignore_errors=True)
            self._profile_dir = None
        self._started = False
        logger.info("LibreOffice engine stopped")

    def recalculate(self, file_path: str) -> dict:
        """
        Recalculate all formulas in the given .xlsx file.

        The file is replaced in-place with cached formula values embedded.
        After this call, openpyxl.load_workbook(data_only=True) will return
        computed values for all formulas.

        Returns:
            dict with keys: success (bool), duration_ms (float), error (str|None)
        """
        if not self._started:
            return {"success": False, "duration_ms": 0, "error": "LibreOffice engine not started"}
        with self._lock:
            return self._convert(file_path)

    @property
    def is_running(self) -> bool:
        # No persistent process any more; "running" means start() succeeded.
        return self._started

    def info(self) -> dict:
        """Engine provenance, recorded in extra_configs and transcripts."""
        return {
            "engine": "libreoffice",
            "soffice_path": self.soffice_path,
            "soffice_source": self.soffice_source,
            "soffice_version": self.soffice_version,
        }

    def _convert(self, file_path: str) -> dict:
        """Run one soffice conversion; on success the recalculated file has
        replaced the original. Returns {success, duration_ms, error}."""
        src = Path(file_path)
        start_time = time.monotonic()
        # Engine-owned temp dir, never inside the workspace, so a failed run
        # can't leak files into the attempt's uploaded artifacts.
        outdir = tempfile.mkdtemp(prefix="lo_recalc_")
        try:
            result = subprocess.run(
                [
                    self.soffice_path,
                    "--headless",
                    "--norestore",
                    "--nologo",
                    "--calc",
                    f"-env:UserInstallation=file://{self._profile_dir}",
                    "--convert-to",
                    "xlsx",
                    "--outdir",
                    outdir,
                    str(src),
                ],
                capture_output=True,
                text=True,
                timeout=self.recalc_timeout,
            )
            converted = Path(outdir) / (src.stem + ".xlsx")
            # soffice can exit 0 without writing anything (filter errors only
            # reach stderr) — the output file is the real success signal.
            if (
                result.returncode != 0
                or not converted.is_file()
                or converted.stat().st_size == 0
            ):
                duration = (time.monotonic() - start_time) * 1000
                error_msg = (
                    result.stderr.strip()
                    or result.stdout.strip()
                    or f"soffice exited {result.returncode} without producing output"
                )
                return {"success": False, "duration_ms": round(duration, 1), "error": error_msg}
            # shutil.move handles a cross-filesystem workspace (copy+rm);
            # same-filesystem it is an atomic rename.
            shutil.move(str(converted), str(src))
            duration = (time.monotonic() - start_time) * 1000
            return {"success": True, "duration_ms": round(duration, 1), "error": None}
        except subprocess.TimeoutExpired:
            duration = (time.monotonic() - start_time) * 1000
            return {
                "success": False,
                "duration_ms": round(duration, 1),
                "error": f"Recalculation timed out after {self.recalc_timeout}s",
            }
        except Exception as e:
            duration = (time.monotonic() - start_time) * 1000
            return {"success": False, "duration_ms": round(duration, 1), "error": str(e)}
        finally:
            shutil.rmtree(outdir, ignore_errors=True)

    def _read_version(self) -> Optional[str]:
        try:
            result = subprocess.run(
                [
                    self.soffice_path,
                    f"-env:UserInstallation=file://{self._profile_dir}",
                    "--version",
                ],
                capture_output=True,
                text=True,
                timeout=30,
            )
            lines = (result.stdout or result.stderr).strip().splitlines()
            return lines[0].strip() if lines else None
        except Exception:
            return None

    def _warmup(self):
        import openpyxl  # server dependency; imported here to keep module import light

        warmup_dir = tempfile.mkdtemp(prefix="lo_warmup_")
        try:
            path = Path(warmup_dir) / "warmup.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws["A1"] = 1
            ws["A2"] = 2
            ws["A3"] = "=SUM(A1:A2)"
            wb.save(path)
            wb.close()

            result = self._convert(str(path))
            if not result["success"]:
                raise RuntimeError(f"LibreOffice warmup conversion failed: {result['error']}")

            check = openpyxl.load_workbook(path, data_only=True)
            value = check.active["A3"].value
            check.close()
            if value != 3:
                raise RuntimeError(
                    f"LibreOffice warmup conversion saved no cached value "
                    f"(expected 3 for =SUM(1,2), got {value!r}) — the install "
                    f"at {self.soffice_path} is broken."
                )
        finally:
            shutil.rmtree(warmup_dir, ignore_errors=True)
