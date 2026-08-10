"""Workbook I/O: loading, saving, file path resolution, auto-fit."""
import os
import re
from pathlib import Path

import openpyxl
from openpyxl.workbook import Workbook

from . import shared_state


def _get_file_path(filename: str) -> Path:
    """Get full path to Excel file."""
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    return shared_state.STORAGE_PATH / filename


def _auto_fit_columns(wb: Workbook) -> None:
    """Auto-fit column widths based on cell content length."""
    for ws in wb.worksheets:
        for col_cells in ws.columns:
            max_length = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                if cell.value is not None:
                    cell_len = len(str(cell.value))
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        cell_len = min(cell_len, 15)
                    max_length = max(max_length, cell_len)
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[col_letter].width = adjusted_width


def _save_workbook_sync(wb: Workbook, file_path: Path) -> None:
    """Save workbook with fsync to ensure data is written to disk."""
    _auto_fit_columns(wb)
    wb.save(file_path)
    with open(file_path, 'r+b') as f:
        os.fsync(f.fileno())
    dir_fd = os.open(str(file_path.parent), os.O_RDONLY | os.O_DIRECTORY)
    try:
        os.fsync(dir_fd)
    finally:
        os.close(dir_fd)


def _slugify(text: str) -> str:
    """Create a filesystem-safe slug from a title."""
    slug = re.sub(r"[^A-Za-z0-9\-_. ]+", "", text).strip().lower().replace(" ", "-")
    return slug[:80] if slug else "issue"


def _load_workbook(filename: str) -> Workbook:
    """Load Excel workbook (raw, formulas visible)."""
    file_path = _get_file_path(filename)
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file '{filename}' not found")
    return openpyxl.load_workbook(file_path, data_only=False)


def _load_workbook_view(filename: str) -> Workbook:
    """Load Excel workbook returning last calculated values (view mode)."""
    file_path = _get_file_path(filename)
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file '{filename}' not found")
    return openpyxl.load_workbook(file_path, data_only=True)
