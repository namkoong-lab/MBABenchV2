"""Worksheet structure analysis: used ranges, blocks, headers, tasks, references."""
import re
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .type_inference import _infer_type


def _actual_used_range(ws: Worksheet, max_rows: Optional[int] = None, max_cols: Optional[int] = None) -> Optional[Tuple[int, int, int, int]]:
    """Compute the actual used rectangular bounds of non-empty cells.
    Returns (min_row, min_col, max_row, max_col) or None if empty.
    Scans up to max_rows x max_cols if provided to limit work.
    """
    min_r, min_c, max_r, max_c = None, None, None, None
    rows = ws.max_row if max_rows is None else min(ws.max_row, max_rows)
    cols = ws.max_column if max_cols is None else min(ws.max_column, max_cols)
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            v = ws.cell(row=r, column=c).value
            if v not in (None, ""):
                min_r = r if min_r is None else min(min_r, r)
                min_c = c if min_c is None else min(min_c, c)
                max_r = r if max_r is None else max(max_r, r)
                max_c = c if max_c is None else max(max_c, c)
    if min_r is None:
        return None
    return (min_r, min_c, max_r, max_c)


def _contiguous_blocks(ws: Worksheet, bounds: Tuple[int, int, int, int]) -> List[Dict[str, Any]]:
    """Find simple contiguous non-empty rectangular blocks within bounds."""
    min_r, min_c, max_r, max_c = bounds
    visited = set()
    blocks: List[Dict[str, Any]] = []

    def expand(r0, c0):
        r = r0
        c = c0
        c2 = c
        while c2 <= max_c and ws.cell(row=r, column=c2).value not in (None, ""):
            c2 += 1
        c2 -= 1
        r2 = r
        while r2 <= max_r:
            ok = True
            for cc in range(c, c2 + 1):
                if ws.cell(row=r2, column=cc).value in (None, ""):
                    ok = False
                    break
            if not ok:
                break
            r2 += 1
        r2 -= 1
        return r, c, r2, c2

    for r in range(min_r, max_r + 1):
        for c in range(min_c, max_c + 1):
            if (r, c) in visited:
                continue
            v = ws.cell(row=r, column=c).value
            if v in (None, ""):
                continue
            r1, c1, r2, c2 = expand(r, c)
            for rr in range(r1, r2 + 1):
                for cc in range(c1, c2 + 1):
                    visited.add((rr, cc))
            blocks.append({
                "top": r1, "left": c1, "bottom": r2, "right": c2,
                "range": f"{openpyxl.utils.get_column_letter(c1)}{r1}:{openpyxl.utils.get_column_letter(c2)}{r2}"
            })
    return blocks


def _iter_sheet_cells(ws: Worksheet, bounds: Tuple[int, int, int, int], row_cap: Optional[int] = None, col_cap: Optional[int] = None):
    """Yield (row, col, cell) within bounds with optional caps."""
    min_r, min_c, max_r, max_c = bounds
    max_r_eff = max_r if row_cap is None else min(max_r, min_r + row_cap - 1)
    max_c_eff = max_c if col_cap is None else min(max_c, min_c + col_cap - 1)
    for r in range(min_r, max_r_eff + 1):
        for c in range(min_c, max_c_eff + 1):
            yield r, c, ws.cell(row=r, column=c)


def _find_task_like_text(ws: Worksheet, bounds: Tuple[int, int, int, int], row_cap: int = 600, col_cap: int = 60) -> List[Dict[str, Any]]:
    """Identify task-like lines in a worksheet."""
    tasks: List[Dict[str, Any]] = []
    patterns = [
        (re.compile(r"^(?:task|question|requirements?)\b", re.I), "label"),
        (re.compile(r"^(?:q\s*\d+\b)", re.I), "question"),
        (re.compile(r"^\d+\s*[\.)]\s+"), "numbered"),
        (re.compile(r"^[\-\*]\s+"), "bullet"),
    ]
    min_r, min_c, max_r, max_c = bounds
    max_r_eff = min(max_r, min_r + max(1, row_cap) - 1)
    max_c_eff = min(max_c, min_c + max(1, col_cap) - 1)
    for r in range(min_r, max_r_eff + 1):
        row_texts = []
        for c in range(min_c, max_c_eff + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                sv = v.strip()
                if sv:
                    row_texts.append((c, sv))
                    for rx, kind in patterns:
                        if rx.search(sv):
                            tasks.append({
                                "cell": f"{openpyxl.utils.get_column_letter(c)}{r}",
                                "text": sv[:300],
                                "kind": kind,
                            })
        if len(row_texts) >= 2:
            joined = " ".join(t for _, t in row_texts)
            for rx, kind in patterns:
                if rx.search(joined):
                    c0 = row_texts[0][0]
                    tasks.append({
                        "cell": f"{openpyxl.utils.get_column_letter(c0)}{r}",
                        "text": joined[:300],
                        "kind": kind,
                        "joined": True,
                    })
    seen = set()
    unique: List[Dict[str, Any]] = []
    for t in tasks:
        key = (t["cell"].split(str(int(t["cell"][1:])))[0], t["kind"], t["text"][:80])
        if key in seen:
            continue
        seen.add(key)
        unique.append(t)
    return unique[:500]


def _find_text_sheet_references(ws: Worksheet, bounds: Tuple[int, int, int, int], sheet_names: List[str], current_sheet: str, row_cap: int = 800, col_cap: int = 80) -> List[Dict[str, Any]]:
    """Find text cells that appear to reference other sheet names."""
    refs: List[Dict[str, Any]] = []
    if not sheet_names:
        return refs
    normalized = [(s, s.lower()) for s in sheet_names if s != current_sheet]
    min_r, min_c, max_r, max_c = bounds
    max_r_eff = min(max_r, min_r + max(1, row_cap) - 1)
    max_c_eff = min(max_c, min_c + max(1, col_cap) - 1)
    for r in range(min_r, max_r_eff + 1):
        for c in range(min_c, max_c_eff + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip():
                lv = v.lower()
                for original, low in normalized:
                    if low in lv:
                        refs.append({
                            "from": f"{openpyxl.utils.get_column_letter(c)}{r}",
                            "to_sheet": original,
                            "context": v[:200],
                            "kind": "text",
                        })
                        break
    out: List[Dict[str, Any]] = []
    seen = set()
    for r in refs:
        key = (r.get("from"), r.get("to_sheet"))
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out[:2000]


def _tokenize_words(text: str) -> List[str]:
    """Simple tokenizer: lowercase split on non-alphanum, drop short tokens; include naive singular forms."""
    toks = re.split(r"[^A-Za-z0-9]+", (text or "").lower())
    out: List[str] = []
    for t in toks:
        if len(t) < 2:
            continue
        out.append(t)
        if len(t) > 3 and t.endswith('s'):
            out.append(t[:-1])
    return out


def _collect_sheet_keywords(ws: Worksheet, bounds: Tuple[int, int, int, int], blocks: List[Dict[str, Any]], row_cap: int = 60, col_cap: int = 40) -> List[str]:
    """Collect keyword candidates from sheet title, header names, and early visible text cells."""
    words: Dict[str, int] = {}
    for tok in _tokenize_words(ws.title):
        words[tok] = words.get(tok, 0) + 3
    for b in blocks[:10]:
        for col in b.get("columns", [])[:20]:
            name = col.get("name")
            if name is None:
                continue
            for tok in _tokenize_words(str(name)):
                words[tok] = words.get(tok, 0) + 2
    min_r, min_c, max_r, max_c = bounds
    max_r_eff = min(max_r, min_r + max(1, row_cap) - 1)
    max_c_eff = min(max_c, min_c + max(1, col_cap) - 1)
    for r in range(min_r, max_r_eff + 1):
        for c in range(min_c, max_c_eff + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                for tok in _tokenize_words(v):
                    words[tok] = words.get(tok, 0) + 1
    top = sorted(words.items(), key=lambda kv: (-kv[1], kv[0]))[:60]
    return [w for w, _ in top]


def _extract_references_from_formula(formula: str, current_sheet: str) -> List[Dict[str, Any]]:
    """Extract cell/range references from a formula string."""
    refs: List[Dict[str, Any]] = []
    f = formula or ""
    sheet_ref_re = re.compile(r"(?:'([^']+)'|([A-Za-z0-9_ .\-]+))!\$?[A-Za-z]{1,3}\$?\d+(?::\$?[A-Za-z]{1,3}\$?\d+)?")
    for m in sheet_ref_re.finditer(f):
        sheet = m.group(1) or (m.group(2).strip() if m.group(2) else None)
        rng = m.group(0)
        if sheet and rng:
            bang = rng.find('!')
            rng_only = rng[bang + 1:] if bang != -1 else rng
            refs.append({"ref_sheet": sheet, "ref_range": rng_only})
    same_re = re.compile(r"\$?[A-Za-z]{1,3}\$?\d+(?::\$?[A-Za-z]{1,3}\$?\d+)?")
    for m in same_re.finditer(f):
        if '!' in m.group(0):
            continue
        refs.append({"ref_sheet": current_sheet, "ref_range": m.group(0)})
    out: List[Dict[str, Any]] = []
    seen = set()
    for r in refs:
        key = (r["ref_sheet"], r["ref_range"])
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out[:1000]


def _detect_header_row(ws: Worksheet, block: Dict[str, Any]) -> Optional[int]:
    """Heuristic to detect a header row within a block: first row with majority strings."""
    r1, c1, r2, c2 = block["top"], block["left"], block["bottom"], block["right"]
    for r in range(r1, min(r1 + 3, r2 + 1)):
        vals = [ws.cell(row=r, column=c).value for c in range(c1, c2 + 1)]
        non_empty = [v for v in vals if v not in (None, "")]
        if not non_empty:
            continue
        strings = sum(1 for v in non_empty if isinstance(v, str))
        if strings >= max(1, int(0.6 * len(non_empty))):
            return r
    return None


def _summarize_columns(ws: Worksheet, header_row: int, c1: int, c2: int, r2: int) -> List[Dict[str, Any]]:
    """Summarize column types and sample values for a headered block."""
    cols: List[Dict[str, Any]] = []
    for c in range(c1, c2 + 1):
        name = ws.cell(row=header_row, column=c).value
        non_null = 0
        types: Dict[str, int] = {}
        examples: List[Any] = []
        for r in range(header_row + 1, r2 + 1):
            v = ws.cell(row=r, column=c).value
            if v in (None, ""):
                continue
            non_null += 1
            t = _infer_type(v)
            types[t] = types.get(t, 0) + 1
            if len(examples) < 3:
                examples.append(v)
        inferred = max(types.items(), key=lambda kv: kv[1])[0] if types else "unknown"
        cols.append({
            "name": name,
            "column_index": c,
            "inferred_type": inferred,
            "non_null": non_null,
            "examples": examples,
        })
    return cols
