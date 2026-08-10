"""Cell reference validation for formulas."""
import re
from typing import List

from openpyxl.workbook import Workbook


def _validate_cell_references(formula: str, workbook: Workbook, current_worksheet: str = None) -> List[str]:
    """
    Validate that cell references in formulas point to appropriate data types.

    Prevents formulas from referencing text cells in mathematical operations.
    Returns list of validation error messages.
    """
    errors = []

    # Pattern to match cell references: A1, B2, 'Sheet Name'!C3, Sheet!D4
    cell_ref_pattern = r"(?:'([^']+)'!)?([A-Z]+[0-9]+)(?![A-Z0-9])"
    matches = re.findall(cell_ref_pattern, formula)

    for sheet_name, cell_address in matches:
        try:
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    continue
                ws = workbook[sheet_name]
            else:
                if current_worksheet and current_worksheet in workbook.sheetnames:
                    ws = workbook[current_worksheet]
                    cell_value = ws[cell_address].value

                    if cell_value is not None and isinstance(cell_value, str):
                        if cell_value.startswith('='):
                            continue

                        formula_context = formula.replace("'", "").replace("!", "")
                        math_operators = ['*', '/', '+', '-', '^', '(', ')']

                        cell_pos = formula_context.find(cell_address)
                        if cell_pos >= 0:
                            start_check = max(0, cell_pos - 3)
                            end_check = min(len(formula_context), cell_pos + len(cell_address) + 3)
                            context = formula_context[start_check:end_check]

                            has_math_operator = any(op in context for op in math_operators)

                            if has_math_operator:
                                errors.append(f"Cell {cell_address} contains text '{cell_value}' but is used in mathematical formula in current worksheet '{current_worksheet}'")
                else:
                    found_problematic_ref = False
                    for ws_name in workbook.sheetnames:
                        ws = workbook[ws_name]
                        cell_value = ws[cell_address].value

                        if cell_value is not None and isinstance(cell_value, str):
                            if cell_value.startswith('='):
                                continue

                            formula_context = formula.replace("'", "").replace("!", "")
                            math_operators = ['*', '/', '+', '-', '^', '(', ')']

                            cell_pos = formula_context.find(cell_address)
                            if cell_pos >= 0:
                                start_check = max(0, cell_pos - 3)
                                end_check = min(len(formula_context), cell_pos + len(cell_address) + 3)
                                context = formula_context[start_check:end_check]

                                has_math_operator = any(op in context for op in math_operators)

                                if has_math_operator:
                                    errors.append(f"Cell {cell_address} contains text '{cell_value}' but is used in mathematical formula")
                                    found_problematic_ref = True
                                    break

                    if found_problematic_ref:
                        continue

            if sheet_name:
                cell_value = ws[cell_address].value
                if cell_value is not None and isinstance(cell_value, str):
                    if cell_value.startswith('='):
                        continue

                    formula_context = formula.replace("'", "").replace("!", "")
                    math_operators = ['*', '/', '+', '-', '^', '(', ')']

                    cell_pos = formula_context.find(cell_address)
                    if cell_pos >= 0:
                        start_check = max(0, cell_pos - 3)
                        end_check = min(len(formula_context), cell_pos + len(cell_address) + 3)
                        context = formula_context[start_check:end_check]

                        has_math_operator = any(op in context for op in math_operators)

                        if has_math_operator:
                            errors.append(f"Cell {sheet_name}!{cell_address} contains text '{cell_value}' but is used in mathematical formula")

        except Exception:
            continue

    return errors
